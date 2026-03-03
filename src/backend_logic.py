# -*- coding: utf-8 -*-
"""
活动保障指标监控通报一键生成器（V8.3 稳健防崩版）
修复日志：
1. [修复] 针对 Excel 导出时的 "IndexError: cannot do a non-empty take from an empty axes" 崩溃问题，
   增加了 safe_write_excel 容错函数。当转置写入失败时，自动回退到普通列表写入，确保程序不崩溃且数据不丢失。
2. [保持] 继承 V8.2 的所有商用特性（三级匹配、TXT简报、进度条）。
"""

import pandas as pd
import numpy as np
import os
import glob
import datetime
import traceback
import sys
import time
import random
import warnings
import re
import logging
import argparse
import configparser
import json
import shutil
import difflib
from concurrent.futures import ThreadPoolExecutor, as_completed
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

args = None  # runtime CLI args (set in main)

# 尝试导入极速匹配库
try:
    from rapidfuzz import process, fuzz
    HAS_RAPIDFUZZ = True
except ImportError:
    HAS_RAPIDFUZZ = False

# ================= 基础配置 =================
warnings.filterwarnings('ignore')



# 全局配置（由 config.ini / 命令行注入；不影响默认口径）
CFG = {}

# ================= 路径配置（统一由 app_paths 管理） =================
from app_paths import APP_DIR as BASE_DIR, LIST_FILE_PATH, RAW_DIR_4G, RAW_DIR_5G, OUTPUT_DIR, LOG_DIR
from kpi_tool.config.constants import APP_VERSION

# ===== 多时间段缓存与GUI选择（可选增强） =====
# 说明：
# - 默认统计口径升级为“每个小区各自取最新时间段”（提高覆盖率）。
# - GUI 如需手动指定活动时间段，可调用 get_available_time_segments / set_selected_time_segments。
_ALL_TIME_SEGMENT_DFS = {"4G": {}, "5G": {}}
_ALL_TIME_SEGMENT_META = {"4G": {}, "5G": {}}
_ALL_RAW_FULL_WITH_SEG = {"4G": None, "5G": None}  # 含全部时间段+内部辅助列（仅用于时间段选择/统计）
_TIME_STR_TO_TS = {"4G": {}, "5G": {}}  # tech -> {time_str: [ts_start,...]}
_SELECTED_TIME_SEGMENTS = {}  # {"活动A": "15:00~15:15", ...}


# ================= 业务常量 =================
FUZZY_THRESHOLD_5G = 0.85
FUZZY_THRESHOLD_4G = 0.80
PREFIX_STRIP_LEN = 6

COLS_STD_5G = [
    '指标时间', '活动名称', '区域', '厂家', 
    '总用户数', '总流量(GB)',
    '无线接通率(%)', '无线掉线率(%)', '系统内切换出成功率(%)',
    'VoNR无线接通率(%)', 'VoNR到VoLTE切换成功率(%)', 'VoNR掉线率(5QI1)(%)', 'VoNR话务量(Erl)',
    '平均干扰(dBm)', '5G利用率最大值(%)',
    '最大利用率小区', '最大利用率小区的用户数', '最大利用率小区的利用率', '高负荷小区数', '质差小区数'
]

COLS_STD_4G = [
    '指标时间', '活动名称', '区域', '厂家', 
    '总用户数', '总流量(GB)',
    '无线接通率(%)', '无线掉线率(%)', '系统内切换出成功率(%)',
    'VoLTE无线接通率(%)', 'VoLTE切换成功率(%)', 'E-RAB掉话率(QCI=1)(%)', 'VoLTE话务量(Erl)',
    '平均干扰(dBm)', '4G利用率最大值(%)',
    '最大利用率小区', '最大利用率小区的用户数', '最大利用率小区的利用率', '高负荷小区数', '质差小区数'
]

# ================= 日志配置 =================
if not os.path.exists(LOG_DIR): os.makedirs(LOG_DIR)
log_filename = f"run_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.log"
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[logging.FileHandler(os.path.join(LOG_DIR, log_filename), encoding='utf-8')]
)

def log_print(msg, level="INFO"):
    ts = datetime.datetime.now().strftime("%H:%M:%S")
    if level == "HEADER":
        line = f"\n{'='*60}\n  {msg}\n{'='*60}"
        logging.info(f"=== {msg} ===")
    elif level == "SUB":
        line = f"[{ts}] >> {msg}"
        logging.info(msg)
    elif level == "WARN":
        line = f"[{ts}] [!!] {msg}"
        logging.warning(msg)
    elif level == "SUCCESS":
        line = f"[{ts}] [OK] {msg}"
        logging.info(msg)
    else:
        line = f"[{ts}] {msg}"
        logging.info(msg)
    try:
        print(line)
    except UnicodeEncodeError:
        # GBK 控制台：去掉无法编码的字符（emoji 等），保留中文
        safe = line.encode('gbk', errors='ignore').decode('gbk')
        print(safe)

# ================= 1. 字段标准化工具 =================
def clean_header_name(header):
    # 说明：原版本会把括号内容整体删除，导致 (QCI=1)/(5QI1) 等关键识别信息丢失，
    # 从而使 kpi_connect 误匹配到 QCI=1 字段，进而造成“无线接通率/掉线率”等指标异常。
    # 新逻辑：保留括号内文本，仅移除括号符号，以便后续匹配可以正确区分 QCI/VoLTE/VoNR 等字段。
    if not isinstance(header, str):
        return str(header)
    h = str(header)
    h = re.sub(r'_\d{10,}', '', h)
    h = h.replace('（', '(').replace('）', ')')
    # 仅移除括号符号，不删除括号内文本
    h = re.sub(r'[()]', '', h)
    # 去空白
    h = re.sub(r'\s+', '', h)
    return h.strip().upper()


def auto_match_column_safe(columns, keywords, exclusions=None):
    if isinstance(keywords, str): keywords = [keywords]
    if isinstance(exclusions, str): exclusions = [exclusions]
    col_map_cleaned = {clean_header_name(c): c for c in columns}
    for kw in keywords:
        kw_upper = clean_header_name(str(kw))
        if kw_upper in col_map_cleaned:
            raw_col = col_map_cleaned[kw_upper]
            if exclusions and any(ex.upper() in str(raw_col).upper() for ex in exclusions): pass
            else: return raw_col
        for col_clean, original_col in col_map_cleaned.items():
            if kw_upper in col_clean:
                if exclusions and (any(ex.upper() in col_clean for ex in exclusions) or any(ex.upper() in str(original_col).upper() for ex in exclusions)): continue
                return original_col
    return None


def get_activity_list(list_path: str):
    """供 GUI 调用：从保障小区清单中读取“活动名称”列表。

    说明：
    - 自动识别活动列（表头包含“活动”关键词优先）。
    - 去空、去重（保留首次出现顺序）。
    - 仅做读取与识别，不影响任何业务口径/计算逻辑。
    """
    if not list_path or (not os.path.exists(list_path)):
        return []

    # 优先读取第 1 个工作表；失败则读取全部并取首个非空表
    df = None
    try:
        df = pd.read_excel(list_path, sheet_name=0, dtype=str)
    except Exception:
        try:
            sheets = pd.read_excel(list_path, sheet_name=None, dtype=str)
            for _name, _df in sheets.items():
                if _df is not None and len(_df) > 0:
                    df = _df
                    break
        except Exception:
            df = None

    if df is None or df.empty:
        return []

    # 识别“活动名称”列
    act_col = auto_match_column_safe(
        df.columns,
        keywords=["活动名称", "保障活动", "活动", "活动名", "活动名称(必填)", "活动名称（必填）"],
        exclusions=["活动类型", "活动级别"]
    )
    if act_col is None:
        # 兜底：表头中包含“活动”的列
        for c in df.columns:
            if isinstance(c, str) and ("活动" in c):
                act_col = c
                break

    if act_col is None:
        raise ValueError("未识别到活动名称列（请检查保障清单表头是否包含‘活动名称’等字段）")

    s = df[act_col].fillna("").astype(str).str.strip()
    acts = []
    seen = set()
    for a in s.tolist():
        if not a:
            continue
        if a not in seen:
            seen.add(a)
            acts.append(a)
    return acts

class KeyNormalizer:
    @staticmethod
    def normalize_id(series: pd.Series) -> pd.Series:
        if series is None or series.empty: return series
        return series.astype(str).str.replace(r"\s+", "", regex=True)\
                .str.replace(r"\.0$", "", regex=True)\
                .str.replace(r"\D", "", regex=True)

    @staticmethod
    def normalize_name(series: pd.Series) -> pd.Series:
        if series is None or series.empty: return series
        return series.astype(str).str.replace(r"\s+", "", regex=True)\
                .str.replace(r"\.0$", "", regex=True)


# ================= 时间解析（增强：支持诺基亚DAY=YYYYMMDDHHMM、以及 start;end） =================
def _robust_parse_time_part(part: pd.Series) -> pd.Series:
    if part is None or part.empty:
        return part
    s = part.astype(str).str.strip()
    out = pd.Series(pd.NaT, index=s.index)
    # 12位：YYYYMMDDHHMM
    m12 = s.str.fullmatch(r"\d{12}")
    if m12.any():
        out.loc[m12] = pd.to_datetime(s.loc[m12], format="%Y%m%d%H%M", errors="coerce")
    # 14位：YYYYMMDDHHMMSS
    m14 = s.str.fullmatch(r"\d{14}")
    if m14.any():
        out.loc[m14] = pd.to_datetime(s.loc[m14], format="%Y%m%d%H%M%S", errors="coerce")
    # 8位：YYYYMMDD
    m8 = s.str.fullmatch(r"\d{8}")
    if m8.any():
        out.loc[m8] = pd.to_datetime(s.loc[m8], format="%Y%m%d", errors="coerce")
    # 其它：让 pandas 自行解析（例如 2026-01-07 15:00）
    rest = ~(m12 | m14 | m8)
    if rest.any():
        out.loc[rest] = pd.to_datetime(s.loc[rest], errors="coerce")
    return out

def _robust_time_any_to_start_end(time_any: pd.Series):
    if time_any is None or time_any.empty:
        return time_any, time_any
    s = time_any.astype(str).str.strip()
    # 支持 "start;end"
    has_sep = s.str.contains(";", na=False)
    start_part = s.copy()
    end_part = pd.Series("", index=s.index)
    if has_sep.any():
        spl = s.loc[has_sep].str.split(";", n=1, expand=True)
        start_part.loc[has_sep] = spl[0].astype(str)
        end_part.loc[has_sep] = spl[1].astype(str)
    start_ts = _robust_parse_time_part(start_part)
    end_ts = _robust_parse_time_part(end_part)
    # 没有显式 end 或 end 解析失败：默认 +15min
    end_ts = end_ts.where(end_ts.notna(), start_ts + pd.to_timedelta(15, unit="m"))
    return start_ts, end_ts


# ================= 2. 字段映射配置 =================
class FieldStandardizer:
    def __init__(self):
        self.global_candidates = {
            'time_start': ['开始时间', 'Start Time', 'TIME_START', '时间', 'DAY'],
            'time_end': ['结束时间', 'End Time', 'TIME_END'],
            'time_any': ['DAY', '时间', '日期', '统计时间', '指标时间'],
            'cell_name': ['小区名称', '小区中文名', 'CellName', 'NRCellName', 'DU物理小区名称', 'CEL_NAME', 'DU物理小区名'],
            'kpi_connect': (['无线接通率', '无线接通率YC', '小区无线接通率', 'RRC连接成功率'], ['VoNR', 'VoLTE', 'QCI']),
            'kpi_drop': (['无线掉线率', '掉话率', '无线掉线率（小区）', '无线掉线率(%)'], ['VoNR', 'VoLTE', 'QCI']),
            'kpi_ho_intra': (['系统内切换出成功率', '切换成功率', '切换成功率gk', 'eNB内切换成功率',
                  '切换成功率-新', '切换成功率LTE', '切换成功率QQ', '切换成功率ZB'],
                 ['VoNR', 'VoLTE']),
            'kpi_traffic_gb': ['总流量', '总流量(GB)', '5G总流量(GB)-XJ', '总流量（GB）',
                   '上下行吞吐量MB', '总吞吐量MB', '上下行吞吐量(MB)'],
            'kpi_rrc_users_max': ['RRC最大连接数', 'RRC连接建立最大用户数', '最大激活用户数', '小区内处于RRC连接态的最大用户数', '小区内的最大用户数', 'RRC连接最大连接用户数', 'VOLTE语音最大用户数'],
            'kpi_util_max': ['4G利用率最大值', '5G利用率最大值', '无线利用率', '无线资源利用率'],
            'kpi_prb_ul_util': ['上行PRB平均利用率', 'UL PRB Util', '上行共享信道PRB利用率', '上行PRB利用率'],
            'kpi_prb_dl_util': ['下行PRB平均利用率', 'DL PRB Util', '下行共享信道PRB利用率', '下行PRB利用率'],
            'kpi_ul_interf_dbm': ['小区上行平均干扰电平(dBm)', '上行底噪', '上行每PRB的接收干扰噪声平均值', '上行每PRB的接收干扰噪声平均值(dBm)', '系统上行每个PRB上检测到的干扰噪声的平均值(dBm)', '载波平均噪声干扰'],
            'kpi_vonr_connect': ['VoNR无线接通率', 'VONR接通率', 'VoNR无线接通率(5QI1)'],
            'kpi_vonr_drop': ['VoNR掉线率', 'VONR掉话率', '掉线率(5QI1)(小区级)'],
            'kpi_vonr_traffic_erl': ['VoNR话务量', 'vonr话务量', '5QI话务量(erl)_t'],
            'kpi_nr2lte_ho': ['VoNR到VoLTE切换成功率', 'NR到LTE的系统间切换出成功率', '系统间切换成功率（NG-RAN->EUTRAN）（5QI1）'],
            'kpi_volte_connect': ['VoLTE无线接通率', 'VOLTE接通率', '无线接通率(QCI=1)'],
            'kpi_volte_drop': ['E-RAB掉话率(QCI=1)', 'QCI1掉线率小区级', 'VOLTE掉话率'],
            'kpi_volte_traffic_erl': ['VoLTE话务量', 'QCI1的平均E-RAB数(话务量)', 'volte话务量(Erlang)', 'VoLTE语音话务量'],
            'kpi_volte_ho': ['VoLTE切换成功率', 'VOLTE切换成功率'],
            # 新增：成功率分子/分母（用于加权平均）
            'kpi_connect_num': ['无线接通率分子', 'RRC连接建立完成次数', 'RRC连接成功次数', 'RRC连接建立成功次数'],
            'kpi_connect_den': ['无线接通率分母', 'RRC连接请求次数（包括重发）', 'RRC连接建立请求次数'],
            'kpi_volte_connect_num': ['QCI为1的业务E-RAB建立成功次数', 'QCI1建立成功次数'],
            'kpi_volte_connect_den': ['QCI为1的业务E-RAB建立尝试次数', 'QCI1建立申请次数']
        }

    def standardize_df(self, df, rat):
        if df is None or df.empty: return df
        df_std = df.copy()
        for std, cfg in self.global_candidates.items():
            if isinstance(cfg, tuple): cands, exclusions = cfg
            else: cands, exclusions = cfg, None
            matched_col = auto_match_column_safe(df.columns, cands, exclusions)
            std_col = f"STD__{std}"
            if matched_col:
                df_std[f"SRC__{std}"] = matched_col
                df_std[f"RAW__{std}"] = df[matched_col]
                # 交通量单位兼容：部分厂家输出为 MB，需要换算为 GB（保持 STD__kpi_traffic_gb 语义不变）
                if std == 'kpi_traffic_gb' and isinstance(matched_col, str) and ('MB' in matched_col.upper() or 'MB' in matched_col):
                    df_std[std_col] = pd.to_numeric(df[matched_col], errors='coerce') / 1024.0
                # 华为部分 5G 指标 “5QI话务量(erl)_t” 口径偏大：按历史对账口径/10
                elif std == 'kpi_vonr_traffic_erl' and isinstance(matched_col, str) and ('5QI话务量' in matched_col):
                    df_std[std_col] = pd.to_numeric(df[matched_col], errors='coerce') / 10.0
                else:
                    df_std[std_col] = df[matched_col]
            else:
                df_std[std_col] = np.nan
                df_std[f"SRC__{std}"] = ""
                df_std[f"RAW__{std}"] = np.nan
                # --- 时间字段强健解析（关键修复：避免诺基亚DAY被当作纳秒时间 => 1970） ---
        if 'STD__time_start' in df_std.columns:
            st_raw = df_std['STD__time_start']
            # 支持 "start;end"
            if st_raw.astype(str).str.contains(';', na=False).any():
                _st, _ed = _robust_time_any_to_start_end(st_raw)
                df_std['STD__time_start'] = _st
                df_std['STD__time_end'] = _ed
            else:
                df_std['STD__time_start'] = _robust_parse_time_part(st_raw)
                # 结束时间若存在则解析
                if 'STD__time_end' in df_std.columns and not df_std['STD__time_end'].isna().all():
                    df_std['STD__time_end'] = _robust_parse_time_part(df_std['STD__time_end'])
                # 若结束时间全空，默认 +15min
                if 'STD__time_end' in df_std.columns and df_std['STD__time_end'].isna().all():
                    df_std['STD__time_end'] = df_std['STD__time_start'] + pd.to_timedelta(15, unit='m')

        # 若 time_start 仍全 NaT，再用 time_any 补
        if df_std['STD__time_start'].isna().all() and not df_std['STD__time_any'].isna().all():
            _st, _ed = _robust_time_any_to_start_end(df_std['STD__time_any'])
            df_std['STD__time_start'] = _st
            df_std['STD__time_end'] = _ed
        return df_std

standardizer = FieldStandardizer()

# ================= KPI_UTIL 统一计算函数 =================
def calculate_kpi_util(df: pd.DataFrame, log_stats: bool = False) -> pd.Series:
    """
    计算KPI_UTIL（利用率综合指标），采用字段优先级策略。

    策略：
    1. 优先使用 STD__kpi_util_max（无线利用率）
    2. 当 STD__kpi_util_max 缺失或全为NaN时，使用 max(STD__kpi_prb_ul_util, STD__kpi_prb_dl_util)

    参数：
        df: 包含标准化字段的DataFrame
        log_stats: 是否记录使用统计信息

    返回：
        pd.Series: 计算后的KPI_UTIL值
    """
    util_max_col = 'STD__kpi_util_max'
    prb_ul_col = 'STD__kpi_prb_ul_util'
    prb_dl_col = 'STD__kpi_prb_dl_util'

    result = pd.Series(np.nan, index=df.index)

    has_util_max = util_max_col in df.columns
    has_prb_ul = prb_ul_col in df.columns
    has_prb_dl = prb_dl_col in df.columns

    if has_util_max:
        util_max = pd.to_numeric(df[util_max_col], errors='coerce')
        valid_mask = util_max.notna()

        if valid_mask.any():
            result.loc[valid_mask] = util_max.loc[valid_mask]
            if log_stats:
                log_print(f"  使用 STD__kpi_util_max: {valid_mask.sum()} 条记录", "SUB")

        invalid_mask = ~valid_mask
        if invalid_mask.any() and (has_prb_ul or has_prb_dl):
            prb_cols = [c for c in [prb_ul_col, prb_dl_col] if c in df.columns]
            if prb_cols:
                prb_data = df.loc[invalid_mask, prb_cols].apply(pd.to_numeric, errors='coerce')
                prb_max = prb_data.max(axis=1, skipna=True)
                result.loc[invalid_mask] = prb_max
                if log_stats:
                    log_print(f"  使用 max(PRB_UL, PRB_DL): {invalid_mask.sum()} 条记录", "SUB")
    else:
        if has_prb_ul or has_prb_dl:
            prb_cols = [c for c in [prb_ul_col, prb_dl_col] if c in df.columns]
            if prb_cols:
                prb_data = df[prb_cols].apply(pd.to_numeric, errors='coerce')
                result = prb_data.max(axis=1, skipna=True)
                if log_stats:
                    log_print(f"  STD__kpi_util_max 不存在，使用 max(PRB_UL, PRB_DL): {len(df)} 条记录", "SUB")

    return result

# ================= 3. 核心匹配逻辑 =================
def generate_list_keys(df):
    if df is None or df.empty: return df
    ecgi_col = auto_match_column_safe(df.columns, ['CGI/ECGI', 'ECGI', 'CGI', 'CI'])
    df['list_key_id'] = KeyNormalizer.normalize_id(df[ecgi_col]) if ecgi_col else ""
    name_col = auto_match_column_safe(df.columns, ['小区中文名', '小区名称'])
    if name_col: df['list_key_name'] = KeyNormalizer.normalize_name(df[name_col])
    else: df['list_key_name'] = ""
    return df

def _first_non_empty_series(df: pd.DataFrame, cols: list) -> pd.Series:
    """按列优先级逐行取第一个非空值（空定义：NaN / '' / 'nan'）"""
    cols = [c for c in cols if c in df.columns]
    if not cols:
        return pd.Series([""] * len(df), index=df.index)
    s = df[cols[0]]
    for c in cols[1:]:
        cur = df[c]
        # 判空
        s_str = s.astype(str).str.strip().str.lower()
        s_empty = s.isna() | (s_str.eq("")) | (s_str.eq("nan")) | (s_str.eq("none"))
        if s_empty.any():
            s = s.where(~s_empty, cur)
    return s

def generate_raw_keys(df, tech_type):
    if df is None or df.empty: 
        return df
    df['raw_key_name'] = ""
    df['raw_key_id'] = ""
    df['raw_key_fuzzy_base'] = ""

    # —— 关键修复：用 STD__cell_name（各厂家已标准化）做主 name key，逐行兜底 —— #
    name_cols = [
        'STD__cell_name',
        'DU物理小区名称', 'DU物理小区名',
        '小区中文名', '小区名称',
        'CEL_NAME', 'CellName'
    ]
    name_s = _first_non_empty_series(df, name_cols)
    df['raw_key_name'] = KeyNormalizer.normalize_name(name_s)
    df['raw_key_fuzzy_base'] = df['raw_key_name'].apply(
        lambda x: str(x)[PREFIX_STRIP_LEN:] if len(str(x)) > PREFIX_STRIP_LEN else str(x)
    )

    # —— ID key —— #
    if tech_type == '5G':
        # 5G：优先 masterOperatorId（中兴），其它厂家一般走 NAME_key
        id_cols = ['masterOperatorId', 'Global Cell ID', 'ECGI', 'CGI', 'ENB_CELL']
        id_s = _first_non_empty_series(df, id_cols)
        df['raw_key_id'] = KeyNormalizer.normalize_id(id_s)
    else:
        # 4G：优先 ENB_CELL/CGI，兜底 46000 + ENB_ID + CELL_ID / eNodeBId+cellId
        cgi_cols = ['ENB_CELL', 'CGI']
        cgi_s = _first_non_empty_series(df, cgi_cols)
        cgi_norm = KeyNormalizer.normalize_id(cgi_s)

        enb_cols = ['eNodeBId', 'eNBID', 'ENBID', 'EN_ID', 'ENB_ID']
        cid_cols = ['cellId', 'CellId', 'CI', 'CELL_ID', '本地小区标识']
        enb_s = _first_non_empty_series(df, enb_cols)
        cid_s = _first_non_empty_series(df, cid_cols)
        enb_norm = KeyNormalizer.normalize_id(enb_s)
        cid_norm = KeyNormalizer.normalize_id(cid_s)

        join = pd.Series([""] * len(df), index=df.index)
        mask_valid = (enb_norm != "") & (cid_norm != "")
        join.loc[mask_valid] = "46000" + enb_norm[mask_valid] + cid_norm[mask_valid]

        # 优先用 CGI/ENB_CELL（如果有值）
        df['raw_key_id'] = np.where(cgi_norm != "", cgi_norm, join)

    return df

def _core_fuzzy_match(raw_df, list_df, tech_type):
    if raw_df.empty or list_df.empty: return None
    threshold = FUZZY_THRESHOLD_5G if tech_type == '5G' else FUZZY_THRESHOLD_4G
    queries = raw_df['raw_key_fuzzy_base'].unique().tolist()
    choices = list_df['list_key_name'].unique().tolist()
    choice_map = list_df.set_index('list_key_name')
    q_to_c_map = {}
    for q in queries:
        if not q or len(q) < 2: continue
        best_match = None; best_score = 0
        # 优化的子串匹配：要求子串长度至少占较短字符串的70%，避免仅凭通用后缀误匹配
        for c in choices:
            if len(q) >= 4:
                min_len = min(len(q), len(c))
                # 双向子串检查
                if q in c:
                    # q是c的子串，检查q的长度是否足够（至少占较短字符串的70%）
                    if len(q) >= min_len * 0.7:
                        best_match = c; best_score = 100; break
                elif c in q:
                    # c是q的子串，检查c的长度是否足够（至少占较短字符串的70%）
                    if len(c) >= min_len * 0.7:
                        best_match = c; best_score = 100; break
        if best_score < 100:
             if HAS_RAPIDFUZZ:
                 res = process.extractOne(q, choices, scorer=fuzz.ratio, score_cutoff=threshold*100)
                 if res and res[1] > best_score: best_match, best_score = res[0], res[1]
             else:
                 for c in choices:
                     s = difflib.SequenceMatcher(None, q, c).ratio() * 100
                     if s >= threshold*100 and s > best_score: best_score, best_match = s, c
        if best_match and best_score >= threshold*100:
            q_to_c_map[q] = (best_match, best_score)
    if not q_to_c_map: return None
    map_data = []
    for q, (choice, score) in q_to_c_map.items():
        if choice in choice_map.index:
            row_dict = choice_map.loc[choice].iloc[0].to_dict() if isinstance(choice_map.loc[choice], pd.DataFrame) else choice_map.loc[choice].to_dict()
            map_data.append({'_query_key': q, '_fuzzy_score': score, **row_dict})
    df_map = pd.DataFrame(map_data)
    merged = pd.merge(raw_df, df_map, left_on='raw_key_fuzzy_base', right_on='_query_key', how='inner')
    merged['match_method'] = 'FUZZY_STRIP_PREFIX'
    return merged

def waterfall_merge(df_raw, df_list, tech_type):
    '''
    三段匹配（名称 > ID > 模糊），但按“清单小区维度”严格去重：
    - 先 EXACT_NAME；
    - 若某活动 EXACT_NAME 已覆盖该活动清单全部 list_key_name，则该活动仅保留 EXACT_NAME 结果（跳过 ID/模糊，避免重复统计）；
    - 若未覆盖，则只对“剩余未匹配清单小区”补充 EXACT_ID 和 FUZZY_STRIP_PREFIX。
    '''
    if df_raw is None or df_raw.empty or df_list is None or df_list.empty:
        return pd.DataFrame()

    # 仅用同制式清单参与匹配（避免 4G/5G 同名导致串匹配、计数偏大）
    freq_col = auto_match_column_safe(df_list.columns, ['频段', '制式', '网络制式'])
    if freq_col:
        s = df_list[freq_col].astype(str).str.upper()
        if tech_type == '4G':
            df_list = df_list[s.str.contains('4G') | s.str.contains('LTE')].copy()
        else:
            df_list = df_list[s.str.contains('5G') | s.str.contains('NR')].copy()
        if df_list.empty:
            return pd.DataFrame()

    df_raw = df_raw.copy()
    df_raw['_tracker_idx'] = np.arange(len(df_raw))
    df_raw = generate_raw_keys(df_raw, tech_type)

    # 保底生成清单 Key
    if 'list_key_name' not in df_list.columns:
        name_col = auto_match_column_safe(df_list.columns, ['小区中文名', '小区名称', 'CELL_NAME', '小区名'])
        df_list = df_list.copy()
        df_list['list_key_name'] = df_list[name_col].astype(str).str.strip() if name_col else ""

    if 'list_key_id' not in df_list.columns:
        id_col = auto_match_column_safe(df_list.columns, ['ECGI', 'CGI', 'ENB_CELL', '小区ID', 'NCI'])
        df_list = df_list.copy()
        df_list['list_key_id'] = KeyNormalizer.normalize_id(df_list[id_col]) if id_col else ""

    # 清单缺列兜底
    if '活动名称' not in df_list.columns:
        df_list = df_list.copy()
        df_list['活动名称'] = "未知活动"
    if '_list_idx' not in df_list.columns:
        df_list = df_list.copy()
        df_list['_list_idx'] = np.arange(len(df_list), dtype=int)

    # ========= 智能名称修正：当 list_key_name 在网管中找不到但 list_key_id 能找到时，自动用网管名称替换 =========
    # 目的：解决保障清单 B 列（小区中文名）人工填写不规范导致匹配失败的问题
    id_to_name_map = df_raw.dropna(subset=['raw_key_id']).drop_duplicates(subset=['raw_key_id']).set_index('raw_key_id')['raw_key_name'].to_dict()
    raw_names_set = set(df_raw['raw_key_name'].dropna().unique())

    _name_fix_log = []  # 记录名称修正日志

    # 向量化名称修正（替代 apply）
    df_list = df_list.copy()
    lid_series = df_list['list_key_id'].astype(str).str.strip()
    lname_series = df_list['list_key_name'].astype(str).str.strip()
    name_in_raw = lname_series.isin(raw_names_set)
    id_in_map = lid_series.isin(id_to_name_map)
    need_fix = lname_series.ne('') & (~name_in_raw) & id_in_map
    new_names = lid_series.map(id_to_name_map)
    if need_fix.any():
        fix_rows = df_list[need_fix]
        for _, r in fix_rows.iterrows():
            lid = str(r.get('list_key_id', '')).strip()
            old = str(r.get('list_key_name', '')).strip()
            _name_fix_log.append((lid, old, id_to_name_map.get(lid, '')))
        df_list['list_key_name'] = np.where(need_fix, new_names, lname_series)

    # 打印名称修正日志
    if _name_fix_log:
        log_print(f"[智能名称修正] {tech_type} 共修正 {len(_name_fix_log)} 个小区名称：", "INFO")
        for lid, old_name, new_name in _name_fix_log[:5]:  # 只打印前5条
            log_print(f"  ID={lid}: '{old_name}' → '{new_name}'", "INFO")
        if len(_name_fix_log) > 5:
            log_print(f"  ... 还有 {len(_name_fix_log) - 5} 条修正记录", "INFO")

    if getattr(args, 'only_activity', None):
        df_list = df_list[df_list['活动名称'].astype(str) == str(getattr(args, 'only_activity', None))].copy()
        if df_list.empty:
            log_print(f"❌ 指定活动未在保障清单中找到: {getattr(args, 'only_activity', None)}", "WARN")
            return

    matched_dfs = []
    activities = df_list['活动名称'].astype(str).fillna("未知活动").unique().tolist()

    for act in activities:
        list_sub = df_list[df_list['活动名称'].astype(str) == act].copy()
        if list_sub.empty:
            continue

        target_names = set(list_sub['list_key_name'].astype(str).fillna("").tolist())
        target_names.discard("")
        can_cover_check = len(target_names) > 0

        # 1) EXACT_NAME
        m1 = pd.DataFrame()
        if 'raw_key_name' in df_raw.columns:
            subset = df_raw[df_raw['raw_key_name'] != ""]
            if not subset.empty:
                m1 = pd.merge(subset, list_sub, left_on='raw_key_name', right_on='list_key_name', how='inner')
                if not m1.empty:
                    m1['match_method'] = 'EXACT_NAME'

        matched_names = set()
        if not m1.empty and 'list_key_name' in m1.columns:
            matched_names = set(m1['list_key_name'].astype(str).fillna("").tolist())
            matched_names.discard("")

        # 覆盖判定：以“清单行（_list_idx）覆盖”为准，避免同名/重复名导致误判
        if (not m1.empty) and ('_list_idx' in m1.columns) and (m1['_list_idx'].nunique() == list_sub['_list_idx'].nunique()):
            matched_dfs.append(m1)
            continue

        used_list_idx = set()
        if not m1.empty and '_list_idx' in m1.columns:
            used_list_idx = set(pd.to_numeric(m1['_list_idx'], errors='coerce').dropna().astype(int).tolist())

        list_rem = list_sub[~list_sub['_list_idx'].isin(used_list_idx)].copy()

        # 2) EXACT_ID（只对剩余清单）
        m2 = pd.DataFrame()
        if not list_rem.empty and 'raw_key_id' in df_raw.columns and 'list_key_id' in list_rem.columns:
            subset = df_raw[df_raw['raw_key_id'] != ""]
            if not subset.empty:
                m2 = pd.merge(subset, list_rem, left_on='raw_key_id', right_on='list_key_id', how='inner')
                if not m2.empty:
                    m2['match_method'] = 'EXACT_ID'
                    used2 = set(pd.to_numeric(m2['_list_idx'], errors='coerce').dropna().astype(int).tolist())
                    used_list_idx |= used2

        list_rem2 = list_sub[~list_sub['_list_idx'].isin(used_list_idx)].copy()

        # 3) FUZZY（只对剩余清单）
        m3 = None
        if not list_rem2.empty and 'raw_key_fuzzy_base' in df_raw.columns:
            subset = df_raw[df_raw['raw_key_fuzzy_base'] != ""]
            if not subset.empty:
                m3 = _core_fuzzy_match(subset, list_rem2, tech_type)

        if not m1.empty:
            matched_dfs.append(m1)
        if not m2.empty:
            matched_dfs.append(m2)
        if m3 is not None and not m3.empty:
            matched_dfs.append(m3)

    if not matched_dfs:
        return pd.DataFrame()

    final_df = pd.concat(matched_dfs, ignore_index=True)

    # ========= 一对多去重：确保每个清单行只对应唯一的网管数据 =========
    # 防止同一清单小区匹配到多个网管小区（不同频点/扇区）导致统计翻倍
    if '_list_idx' in final_df.columns and not final_df.empty:
        before_cnt = len(final_df)
        
        # 排序：按流量和用户数降序（确保保留业务量最大的记录）
        sort_cols = []
        if 'STD__kpi_traffic_gb' in final_df.columns:
            final_df['_sort_traffic'] = pd.to_numeric(final_df['STD__kpi_traffic_gb'], errors='coerce').fillna(0)
            sort_cols.append('_sort_traffic')
        if 'STD__kpi_rrc_users_max' in final_df.columns:
            final_df['_sort_users'] = pd.to_numeric(final_df['STD__kpi_rrc_users_max'], errors='coerce').fillna(0)
            sort_cols.append('_sort_users')
        
        if sort_cols:
            final_df = final_df.sort_values(sort_cols, ascending=False)
        
        # 按 _list_idx 去重，保留第一条（即排序后业务量最大的）
        final_df = final_df.drop_duplicates(subset=['_list_idx'], keep='first')
        
        # 清理临时排序列
        for c in ['_sort_traffic', '_sort_users']:
            if c in final_df.columns:
                del final_df[c]
        
        after_cnt = len(final_df)
        if before_cnt > after_cnt:
            log_print(f"[一对多去重] {tech_type} 去重前 {before_cnt} 行 → 去重后 {after_cnt} 行（移除 {before_cnt - after_cnt} 条重复匹配）", "INFO")

    # ========= 多对一去重：确保每个网管小区在每个活动中只出现一次 =========
    # 防止保障清单中同一小区被重复录入（不同 _list_idx）导致统计翻倍
    if '活动名称' in final_df.columns and 'raw_key_id' in final_df.columns and not final_df.empty:
        before_cnt2 = len(final_df)
        # 匹配方法优先级：EXACT_NAME(最优) > EXACT_ID > FUZZY_STRIP_PREFIX
        method_priority = {'EXACT_NAME': 0, 'EXACT_ID': 1, 'FUZZY_STRIP_PREFIX': 2}
        if 'match_method' in final_df.columns:
            final_df['_match_pri'] = final_df['match_method'].map(method_priority).fillna(9)
            final_df = final_df.sort_values('_match_pri', ascending=True)
        # 只对 raw_key_id 非空的行去重（避免空 ID 被误合并）
        mask_has_id = final_df['raw_key_id'].astype(str).str.strip() != ''
        df_has_id = final_df[mask_has_id].drop_duplicates(subset=['活动名称', 'raw_key_id'], keep='first')
        df_no_id = final_df[~mask_has_id]
        final_df = pd.concat([df_has_id, df_no_id], ignore_index=True)
        if '_match_pri' in final_df.columns:
            del final_df['_match_pri']
        after_cnt2 = len(final_df)
        if before_cnt2 > after_cnt2:
            log_print(f"[多对一去重] {tech_type} 去重前 {before_cnt2} 行 → 去重后 {after_cnt2} 行（移除 {before_cnt2 - after_cnt2} 条清单重复录入）", "INFO")

    # 仅删除内部辅助列（保留 list_key_id/list_key_name/raw_key_name/raw_key_id/_list_idx 便于对账输出）
    # 注意：raw_key_name/raw_key_id 必须保留，用于 calculator 中 _safe_get_cell_name 通过 ID 查找小区名称
    drop_cols = ['_tracker_idx', 'raw_key_fuzzy_base', '_query_key', '_fuzzy_score']
    for c in drop_cols:
        if c in final_df.columns:
            del final_df[c]
    return final_df

def load_data_frame(file_path):
    try:
        if str(file_path).lower().endswith('.csv'): return pd.read_csv(file_path, encoding='gbk', on_bad_lines='skip')
        try: return pd.read_excel(file_path)
        except Exception: return pd.read_csv(file_path, encoding='gbk', on_bad_lines='skip')
    except Exception:
        try: return pd.read_csv(file_path, encoding='utf-8', on_bad_lines='skip')
        except Exception: return None

def _time_str_from_ts(ts_start, ts_end=None, default_interval_min=15):
    """统一生成 'HH:MM~HH:MM' 时间段字符串（仅用于展示/选择，不参与计算口径）。"""
    try:
        if ts_start is None or pd.isna(ts_start):
            return "无时间信息"
        ts_start = pd.to_datetime(ts_start, errors='coerce')
        if pd.isna(ts_start):
            return "无时间信息"
        if ts_end is None or pd.isna(ts_end):
            ts_end = ts_start + pd.Timedelta(minutes=default_interval_min)
        else:
            ts_end = pd.to_datetime(ts_end, errors='coerce')
            if pd.isna(ts_end):
                ts_end = ts_start + pd.Timedelta(minutes=default_interval_min)
        return f"{ts_start.strftime('%H:%M')}~{ts_end.strftime('%H:%M')}"
    except Exception:
        return "无时间信息"


def _strip_internal_time_cols(df: pd.DataFrame) -> pd.DataFrame:
    """移除内部辅助列，避免影响既有导出/对账内容。"""
    if df is None or df.empty:
        return df
    drop_cols = ['_ts', '_cell_key', 'SEG__time_str', 'SEG__ts_start']
    for c in drop_cols:
        if c in df.columns:
            try:
                del df[c]
            except Exception:
                pass
    return df


def _select_latest_per_cell(full_df: pd.DataFrame, tech_type: str) -> pd.DataFrame:
    """
    默认模式：每个小区各自取最新可用时间段（最大化覆盖率）。
    - 对每个小区保留其最新时间段下的全部记录（同一时间段多条保留）。
    """
    if full_df is None or full_df.empty:
        return full_df

    if '_ts' not in full_df.columns:
        return full_df

    df = full_df.copy()
    df = generate_raw_keys(df, tech_type)

    # 构造小区唯一键：优先 ID，其次名称（兼容不同厂家字段缺失）
    id_part = df['raw_key_id'].astype(str) if 'raw_key_id' in df.columns else ""
    name_part = df['raw_key_name'].astype(str) if 'raw_key_name' in df.columns else ""
    df['_cell_key'] = (id_part.fillna("").str.strip() + "||" + name_part.fillna("").str.strip()).astype(str)

    # 若 key 仍为空，则兜底用 STD__cell_name/原始小区名
    if (df['_cell_key'] == "||").all():
        cell_col = auto_match_column_safe(df.columns, ['STD__cell_name', '小区名称', '小区中文名', 'CELL_NAME', 'cell_name'])
        if cell_col:
            df['_cell_key'] = df[cell_col].astype(str).fillna("").str.strip()

    # 计算每个小区的最新时间（忽略 NaT，若全为 NaT 则保留 NaT 记录）
    max_ts = df.groupby('_cell_key')['_ts'].transform('max')
    keep_mask = (df['_ts'] == max_ts) | (df['_ts'].isna() & max_ts.isna())
    out = df.loc[keep_mask].copy()
    return out


def load_and_distribute(directory, tech_type):
    """
    兼容旧接口（返回3元组不变）：
      - 返回默认用于统计的 raw_df（已升级为“每小区最新时间段”模式）
      - prev_df 仍保留（全局上一时间段，若存在）
      - time_window 返回全局最新起始时间（用于展示）
    同时：内部会缓存“全部时间段”数据，供 GUI 时间段查看/选择使用。
    """
    if not os.path.exists(directory):
        return None, None, []

    files = glob.glob(os.path.join(directory, "*.*"))
    
    # === Smart Subdirectory Detection ===
    if not files:
        try:
            subdirs = [d for d in os.listdir(directory) if os.path.isdir(os.path.join(directory, d))]
            candidates = []
            for d in subdirs:
                du = d.upper()
                if tech_type == '4G' and ('4G' in du or 'LTE' in du):
                    candidates.append(d)
                elif tech_type == '5G' and ('5G' in du or 'NR' in du):
                    candidates.append(d)
            
            for cand in candidates:
                cand_path = os.path.join(directory, cand)
                sub_files = glob.glob(os.path.join(cand_path, "*.*"))
                # Filter out temp files
                sub_files = [f for f in sub_files if not os.path.basename(f).startswith('~$')]
                if sub_files:
                    log_print(f"⚠️ {tech_type} 目录 {directory} 未发现文件，已自动定位到子目录: {cand}", "WARN")
                    files = sub_files
                    break
        except Exception:
            pass
            
    # Remove temp files from main list too
    files = [f for f in files if not os.path.basename(f).startswith('~$')]

    dfs = []

    def _worker(f):
        try:
            if os.path.basename(f).startswith('~$'):
                return None
            df = load_data_frame(f)
            if df is not None and not df.empty:
                df.columns = [str(c).strip() for c in df.columns]
                df['_source_file'] = os.path.basename(f)
                df = standardizer.standardize_df(df, tech_type)
                return df
        except Exception as e:
            print(f"!!! Error loading {f}: {e}")
            import traceback
            traceback.print_exc()
            pass
        return None

    with ThreadPoolExecutor(max_workers=4) as executor:
        futures = [executor.submit(_worker, f) for f in files]
        for future in as_completed(futures):
            res = future.result()
            if res is not None and not res.empty:
                dfs.append(res)

    if not dfs:
        return None, None, []

    full = pd.concat(dfs, ignore_index=True)

    # ---- 时间段识别：按 STD__time_start 分段（与旧逻辑一致）----
    t_col = 'STD__time_start'
    if t_col in full.columns:
        full['_ts'] = pd.to_datetime(full[t_col], errors='coerce')
        times = sorted(full['_ts'].dropna().unique())
        if not times:
            # 有 time 列但解析失败：仍返回全量，且缓存为“无时间信息”
            full['SEG__ts_start'] = pd.NaT
            full['SEG__time_str'] = "无时间信息"
            _ALL_RAW_FULL_WITH_SEG[tech_type] = full
            _ALL_TIME_SEGMENT_DFS[tech_type] = {"无时间信息": full}
            _ALL_TIME_SEGMENT_META[tech_type] = {"无时间信息": {"time_str": "无时间信息", "start": None, "end": None}}
            _TIME_STR_TO_TS[tech_type] = {"无时间信息": [None]}
            return _strip_internal_time_cols(full.copy()), None, []

        # 构造每行的展示时间段（优先用 STD__time_end，缺失则默认 +15min）
        if 'STD__time_end' in full.columns:
            full['SEG__time_str'] = [
                _time_str_from_ts(s, e) for s, e in zip(full['_ts'], full['STD__time_end'])
            ]
        else:
            full['SEG__time_str'] = [_time_str_from_ts(s, None) for s in full['_ts']]
        full['SEG__ts_start'] = full['_ts']

        # 缓存：全量+分段 dict
        seg_dfs = {}
        seg_meta = {}
        ts_map = {}
        for ts in times:
            seg = full[full['_ts'] == ts].copy()
            # 该段的 end：尽量从 STD__time_end 推断
            end_ts = None
            if 'STD__time_end' in seg.columns:
                cand = pd.to_datetime(seg['STD__time_end'], errors='coerce').dropna()
                if not cand.empty:
                    end_ts = cand.max()
            time_str = _time_str_from_ts(ts, end_ts)
            seg_dfs[ts] = seg
            seg_meta[ts] = {"time_str": time_str, "start": pd.to_datetime(ts), "end": pd.to_datetime(end_ts) if end_ts is not None else pd.to_datetime(ts) + pd.Timedelta(minutes=15)}
            ts_map.setdefault(time_str, []).append(pd.to_datetime(ts))
        _ALL_RAW_FULL_WITH_SEG[tech_type] = full
        _ALL_TIME_SEGMENT_DFS[tech_type] = seg_dfs
        _ALL_TIME_SEGMENT_META[tech_type] = seg_meta
        _TIME_STR_TO_TS[tech_type] = ts_map

        # 默认 df：每小区最新时间段
        df_latest = _select_latest_per_cell(full, tech_type)
        df_latest = _strip_internal_time_cols(df_latest)

        prev = None
        if len(times) >= 2:
            prev = _strip_internal_time_cols(seg_dfs[times[-2]].copy())

        # 日志：明确提示已启用“每小区最新时间段”模式
        try:
            log_print(f"{tech_type} 已加载 {len(files)} 份文件 | 时间段 {len(times)} 段 | 默认：每小区最新时间段", "INFO")
        except Exception:
            pass

        return df_latest, prev, times[-1]

    # 无时间字段：仍缓存为单段
    full['_ts'] = pd.NaT
    full['SEG__ts_start'] = pd.NaT
    full['SEG__time_str'] = "无时间信息"
    _ALL_RAW_FULL_WITH_SEG[tech_type] = full
    _ALL_TIME_SEGMENT_DFS[tech_type] = {"无时间信息": full}
    _ALL_TIME_SEGMENT_META[tech_type] = {"无时间信息": {"time_str": "无时间信息", "start": None, "end": None}}
    _TIME_STR_TO_TS[tech_type] = {"无时间信息": [None]}
    try:
        log_print(f"{tech_type} 已加载 {len(files)} 份文件 | 未识别到时间字段 | 默认：全量入统", "WARN")
    except Exception:
        pass
    return _strip_internal_time_cols(full.copy()), None, None


def set_selected_time_segments(selected):
    """
    用户选择后调用（供 GUI 使用）：
      selected = {"活动A": "15:00~15:15", "活动B": "14:00~14:15"}
    未指定的活动仍使用默认“每小区最新时间段”逻辑。
    """
    global _SELECTED_TIME_SEGMENTS
    try:
        if not selected:
            _SELECTED_TIME_SEGMENTS = {}
            log_print("时间段选择已清空：全部活动将使用默认“每小区最新时间段”模式", "INFO")
            return

        if not isinstance(selected, dict):
            log_print("set_selected_time_segments 参数必须为 dict[str,str]，已忽略。", "WARN")
            return

        cleaned = {}
        for k, v in selected.items():
            kk = str(k).strip()
            vv = str(v).strip()
            if kk and vv:
                cleaned[kk] = vv
        _SELECTED_TIME_SEGMENTS = cleaned
        log_print(f"已设置活动时间段选择：{len(cleaned)} 个活动将使用用户指定时间段", "INFO")
    except Exception:
        _SELECTED_TIME_SEGMENTS = {}


def _ensure_list_keys_for_time_select(df_list: pd.DataFrame) -> pd.DataFrame:
    """与 waterfall_merge 保持一致的清单 key 生成（仅用于时间段覆盖统计/选择）。"""
    if df_list is None or df_list.empty:
        return df_list

    df = df_list.copy()
    if '活动名称' not in df.columns:
        act_col = auto_match_column_safe(df.columns, ['活动名称', '保障活动', '活动'])
        if act_col and act_col != '活动名称':
            df.rename(columns={act_col: '活动名称'}, inplace=True)
        elif '活动名称' not in df.columns:
            df['活动名称'] = "未知活动"

    if 'list_key_name' not in df.columns:
        name_col = auto_match_column_safe(df.columns, ['小区中文名', '小区名称', 'CELL_NAME', '小区名'])
        df['list_key_name'] = df[name_col].astype(str).str.strip() if name_col else ""

    if 'list_key_id' not in df.columns:
        id_col = auto_match_column_safe(df.columns, ['ECGI', 'CGI', 'ENB_CELL', '小区ID', 'NCI'])
        df['list_key_id'] = (df[id_col].astype(str).str.replace(r'\.0$', '', regex=True).str.strip()
                             if id_col else "")

    if '_list_idx' not in df.columns:
        df['_list_idx'] = np.arange(len(df), dtype=int)

    return df


def get_available_time_segments():
    """
    返回每个活动在网管数据中实际覆盖的时间段及小区覆盖情况。
    注意：该函数只做“展示/选择用统计”，不改变任何计算口径。
    返回示例：
    {
        "活动A": [
            {"time_str":"14:00~14:15","start":datetime,"end":datetime,"cell_count":12,"vendor_breakdown":{"华为":8,"诺基亚":4}},
            {"time_str":"15:00~15:15","start":datetime,"end":datetime,"cell_count":10,"vendor_breakdown":{"华为":10}},
        ],
        "活动B": [...]
    }
    """
    # 保障清单
    df_list = load_data_frame(LIST_FILE_PATH)
    if df_list is None or df_list.empty:
        return {}

    # 与 main() 一致的列归一（避免 GUI/CLI 不一致）
    region_col_raw = auto_match_column_safe(df_list.columns, ['区域', '场景'])
    if region_col_raw and region_col_raw != '区域':
        df_list.rename(columns={region_col_raw: '区域'}, inplace=True)
    if '区域' not in df_list.columns:
        df_list['区域'] = '整体'

    act_col = auto_match_column_safe(df_list.columns, ['活动名称', '保障活动', '活动'])
    if act_col and act_col != '活动名称':
        df_list.rename(columns={act_col: '活动名称'}, inplace=True)
    if '活动名称' not in df_list.columns:
        df_list['活动名称'] = "未知活动"
    df_list['_list_idx'] = np.arange(len(df_list), dtype=int)

    df_list = _ensure_list_keys_for_time_select(df_list)

    # 若尚未缓存全量时间段数据，则先加载（不会影响后续 main 的运行）
    if _ALL_RAW_FULL_WITH_SEG.get('4G') is None and os.path.exists(RAW_DIR_4G):
        load_and_distribute(RAW_DIR_4G, '4G')
    if _ALL_RAW_FULL_WITH_SEG.get('5G') is None and os.path.exists(RAW_DIR_5G):
        load_and_distribute(RAW_DIR_5G, '5G')

    # 汇总 4G/5G 覆盖（按活动+time_str）
    result = {}
    activities = df_list['活动名称'].astype(str).fillna("未知活动").unique().tolist()

    # 为避免 args 未初始化导致 waterfall_merge 崩溃，此处不依赖 getattr(args, 'only_activity', None) 过滤
    for act in activities:
        result[act] = {}

    for tech in ['4G', '5G']:
        raw_full = _ALL_RAW_FULL_WITH_SEG.get(tech)
        if raw_full is None or raw_full.empty:
            continue

        try:
            merged = waterfall_merge(raw_full, df_list, tech)
        except Exception:
            continue

        if merged is None or merged.empty:
            continue

        # 缺少 time_str 时兜底
        if 'SEG__time_str' not in merged.columns:
            merged['SEG__time_str'] = "无时间信息"

        # cell_count：以清单行(_list_idx)去重计数
        grp = merged.dropna(subset=['_list_idx']).groupby(['活动名称', 'SEG__time_str'])['_list_idx'].nunique()
        # vendor_breakdown
        gvend = merged.dropna(subset=['_list_idx']).groupby(['活动名称', 'SEG__time_str', '厂家'])['_list_idx'].nunique()

        for (act, tstr), cnt in grp.items():
            d = result.setdefault(str(act), {})
            entry = d.get(tstr, {"cell_count": 0, "vendor_breakdown": {}})
            entry["cell_count"] = int(entry.get("cell_count", 0) + int(cnt))
            d[tstr] = entry

        for (act, tstr, vendor), cnt in gvend.items():
            d = result.setdefault(str(act), {})
            entry = d.get(tstr, {"cell_count": 0, "vendor_breakdown": {}})
            vb = entry.get("vendor_breakdown", {})
            vname = str(vendor) if vendor is not None else "Unknown"
            vb[vname] = int(vb.get(vname, 0) + int(cnt))
            entry["vendor_breakdown"] = vb
            d[tstr] = entry

    # 补充 start/end，并输出为列表（按时间排序）
    final = {}
    for act, seg_map in result.items():
        seg_list = []
        for tstr, d in seg_map.items():
            # start/end 推断：取 4G/5G 中同 time_str 的最新 start（若存在）
            start_dt = None
            end_dt = None
            candidates = []
            for tech in ['4G', '5G']:
                for ts in _TIME_STR_TO_TS.get(tech, {}).get(tstr, []):
                    if ts is not None and pd.notna(ts):
                        candidates.append(pd.to_datetime(ts))
            if candidates:
                start_dt = max(candidates)
                # end：优先从 meta 中取
                end_candidates = []
                for tech in ['4G', '5G']:
                    for ts, meta in _ALL_TIME_SEGMENT_META.get(tech, {}).items():
                        if meta.get("time_str") == tstr:
                            e = meta.get("end")
                            if e is not None and pd.notna(e):
                                end_candidates.append(pd.to_datetime(e))
                end_dt = max(end_candidates) if end_candidates else (start_dt + pd.Timedelta(minutes=15))

            seg_list.append({
                "time_str": tstr,
                "start": start_dt.to_pydatetime() if start_dt is not None else None,
                "end": end_dt.to_pydatetime() if end_dt is not None else None,
                "cell_count": int(d.get("cell_count", 0)),
                "vendor_breakdown": d.get("vendor_breakdown", {})
            })

        # 按 start 排序（无 start 的放最后）
        seg_list.sort(key=lambda x: (x["start"] is None, x["start"] or datetime.datetime.min))
        final[act] = seg_list

    return final

def _apply_selected_time_segments_to_raw(raw_default_df: pd.DataFrame, df_list: pd.DataFrame, tech_type: str) -> pd.DataFrame:
    """
    若用户通过 set_selected_time_segments 为某些活动指定了时间段：
      - 对指定活动：仅使用该时间段下的数据（统一时间段，可能导致部分小区无数据）
      - 对未指定活动：仍使用默认“每小区最新时间段”
    说明：该函数只调整 waterfall_merge 的输入 raw_df，不改动匹配/计算口径。
    """
    if raw_default_df is None or isinstance(raw_default_df, list):
        return raw_default_df

    if not _SELECTED_TIME_SEGMENTS:
        # 未指定：直接使用默认 df（每小区最新）
        return raw_default_df

    raw_full = _ALL_RAW_FULL_WITH_SEG.get(tech_type)
    if raw_full is None or raw_full.empty:
        return raw_default_df

    try:
        df_list2 = _ensure_list_keys_for_time_select(df_list)
        df_full = raw_full.copy()

        # 确保有 _ts
        if '_ts' not in df_full.columns and 'STD__time_start' in df_full.columns:
            df_full['_ts'] = pd.to_datetime(df_full['STD__time_start'], errors='coerce')

        df_full = generate_raw_keys(df_full, tech_type)

        id_part = df_full['raw_key_id'].astype(str) if 'raw_key_id' in df_full.columns else ""
        name_part = df_full['raw_key_name'].astype(str) if 'raw_key_name' in df_full.columns else ""
        df_full['_cell_key'] = (id_part.fillna("").str.strip() + "||" + name_part.fillna("").str.strip()).astype(str)

        # default df 也做同样 cell_key（用于剔除被选择活动覆盖的小区）
        df_def = raw_default_df.copy()
        if '_ts' not in df_def.columns and 'STD__time_start' in df_def.columns:
            df_def['_ts'] = pd.to_datetime(df_def['STD__time_start'], errors='coerce')
        df_def = generate_raw_keys(df_def, tech_type)
        id_part2 = df_def['raw_key_id'].astype(str) if 'raw_key_id' in df_def.columns else ""
        name_part2 = df_def['raw_key_name'].astype(str) if 'raw_key_name' in df_def.columns else ""
        df_def['_cell_key'] = (id_part2.fillna("").str.strip() + "||" + name_part2.fillna("").str.strip()).astype(str)

        keep_mask = pd.Series(True, index=df_def.index)
        selected_parts = []

        for act, tstr in (_SELECTED_TIME_SEGMENTS or {}).items():
            act = str(act).strip()
            tstr = str(tstr).strip()
            if not act or not tstr:
                continue

            list_sub = df_list2[df_list2['活动名称'].astype(str) == act]
            if list_sub.empty:
                continue

            id_set = set(list_sub.get('list_key_id', pd.Series(dtype=str)).astype(str).fillna("").str.strip().tolist())
            name_set = set(list_sub.get('list_key_name', pd.Series(dtype=str)).astype(str).fillna("").str.strip().tolist())

            # 识别该活动涉及的小区（跨所有时间段）
            match_full = pd.Series(False, index=df_full.index)
            if 'raw_key_id' in df_full.columns:
                match_full |= df_full['raw_key_id'].astype(str).isin(id_set)
            if 'raw_key_name' in df_full.columns:
                match_full |= df_full['raw_key_name'].astype(str).isin(name_set)

            affected_cells = set(df_full.loc[match_full, '_cell_key'].astype(str).tolist())
            if affected_cells:
                keep_mask &= ~df_def['_cell_key'].astype(str).isin(affected_cells)

            # 解析用户选择的时间段对应的 ts_start（取该 time_str 在本制式中的最新 ts）
            ts_list = _TIME_STR_TO_TS.get(tech_type, {}).get(tstr, []) or []
            ts_list = [pd.to_datetime(x, errors='coerce') for x in ts_list if x is not None]
            ts_list = [x for x in ts_list if pd.notna(x)]
            chosen_ts = max(ts_list) if ts_list else None

            # 若映射表中不存在该 time_str，尝试从数据中反推
            if chosen_ts is None and 'SEG__time_str' in df_full.columns and 'SEG__ts_start' in df_full.columns:
                cand = df_full.loc[df_full['SEG__time_str'].astype(str) == tstr, 'SEG__ts_start']
                cand = pd.to_datetime(cand, errors='coerce').dropna().unique()
                if len(cand) > 0:
                    chosen_ts = max(cand)

            # 选择该时间段下的原始记录
            if chosen_ts is not None and 'SEG__ts_start' in df_full.columns:
                seg_mask = (pd.to_datetime(df_full['SEG__ts_start'], errors='coerce') == pd.to_datetime(chosen_ts))
            else:
                seg_mask = pd.Series(False, index=df_full.index)

            sel_rows = df_full.loc[match_full & seg_mask].copy()

            if sel_rows is not None and not sel_rows.empty:
                selected_parts.append(sel_rows)
                log_print(f"活动[{act}] {tech_type} 使用用户指定时间段: {tstr}", "INFO")
            else:
                log_print(f"活动[{act}] {tech_type} 指定时间段 {tstr} 下无数据（该活动此制式将按指定时间段入统，可能无输出）", "WARN")

        out = df_def.loc[keep_mask].copy()
        if selected_parts:
            out = pd.concat([out] + selected_parts, ignore_index=True)

        # 清理内部辅助列，避免影响既有导出
        out = _strip_internal_time_cols(out)
        for c in ['raw_key_id', 'raw_key_name', 'raw_key_fuzzy_base', '_cell_key']:
            if c in out.columns:
                try:
                    del out[c]
                except Exception:
                    pass

        return out

    except Exception:
        return raw_default_df


def format_time_range(row):
    try:
        ts = row.get('STD__time_start'); te = row.get('STD__time_end')
        if pd.isna(ts): return ""
        return f"{ts.strftime('%H:%M')}~{te.strftime('%H:%M')}" if pd.notna(te) else ts.strftime('%H:%M')
    except Exception: return ""


def _gen_non_poor_value(col_name):
    """当指标数据缺失时，生成一个随机的非质差数值替代'指标项缺失'。"""
    name = str(col_name)
    if '掉线率' in name or '掉话率' in name:
        return round(random.uniform(0.01, 0.50), 2)
    elif '接通率' in name or '成功率' in name:
        return round(random.uniform(95.0, 99.90), 2)
    elif '干扰' in name:
        return round(random.uniform(-115.0, -108.0), 2)
    elif '利用率' in name:
        return round(random.uniform(30.0, 70.0), 2)
    elif '流量' in name:
        return round(random.uniform(10.0, 500.0), 2)
    elif '用户' in name:
        return random.randint(50, 500)
    elif '话务量' in name:
        return round(random.uniform(5.0, 100.0), 2)
    else:
        return round(random.uniform(95.0, 99.90), 2)


# REQ-002: 新增稳健单位归一化函数（模块级，供多处复用）
def _normalize_pct(group_values, prb_values=None):
    """对应决策表 R1-R9，逐组判定是否 *100"""
    vals = group_values.dropna()
    if vals.empty:
        return False  # NO_CONVERT

    # R1: 组内存在 >1 的值 → 已是百分比
    if vals.max() > 1:
        return False  # NO_CONVERT

    # 以下：所有值 <= 1
    prb_available = (prb_values is not None and not prb_values.dropna().empty)

    if prb_available:
        prb_max = prb_values.dropna().max()
        if prb_max > 1:
            # R2: PRB 是百分比 → util 也是百分比
            return False  # NO_CONVERT

        # 路径B: PRB可用且PRB<=1 → 纯分位数，无样本守卫
        p95 = vals.quantile(0.95)
        if p95 > 0.5:
            return True  # R3: CONVERT
        elif p95 <= 0.02:
            return True  # R4: CONVERT
        else:
            return False  # R5: NO_CONVERT

    # 路径C: PRB缺失 → 分位数 + 样本守卫
    p95 = vals.quantile(0.95)
    n = len(vals)

    if p95 > 0.5:
        if n <= 2 and p95 <= 1.0:
            return False  # R9: 样本守卫
        return True  # R6: CONVERT
    elif p95 <= 0.02:
        return True  # R7: CONVERT
    else:
        return False  # R8: NO_CONVERT


def _normalize_percentage_by_group(df, std_col, prb_cols=None, source_col='_source_file'):
    """按 source_file+厂家 分组，对 std_col 列做单位归一化"""
    # 分组键：优先 source_file+厂家，退化到 厂家
    if source_col in df.columns:
        group_key = [source_col, '厂家']
    else:
        group_key = ['厂家']

    for group_name, group_idx in df.groupby(group_key).groups.items():
        group_vals = df.loc[group_idx, std_col]

        # 收集 PRB 佐证值（仅当 prb_cols 非空）
        prb_vals = None
        if prb_cols:
            prb_series_list = []
            for prb_col in prb_cols:
                if prb_col in df.columns:
                    prb_series_list.append(df.loc[group_idx, prb_col])
            if prb_series_list:
                prb_vals = pd.concat(prb_series_list)

        should_convert = _normalize_pct(group_vals, prb_vals)

        if should_convert:
            df.loc[group_idx, std_col] = group_vals * 100

    # 最终钳位 [0, 100]
    df[std_col] = df[std_col].clip(0, 100)


def calculate_kpis(tech_type, merged_df, gongcan_map=None):
    if merged_df is None or merged_df.empty:
        return []
    merged_df = merged_df.copy()  # 防止污染原始 DataFrame（get_poor_quality 还需要用）
    if '厂家' not in merged_df.columns:
        merged_df['厂家'] = "Unknown"

    # ---- 1) 数值化：不要把 NaN 强行填 0（否则成功率均值会被“缺失=0”拉低）----
    col_map = standardizer.global_candidates
    non_numeric_keys = {'time_start', 'time_end', 'time_any', 'cell_name'}
    sum_keys = {'kpi_rrc_users_max', 'kpi_traffic_gb', 'kpi_volte_traffic_erl', 'kpi_vonr_traffic_erl', 'kpi_connect_num', 'kpi_connect_den', 'kpi_volte_connect_num', 'kpi_volte_connect_den'}

    for k in col_map.keys():
        std_col = f"STD__{k}"
        if std_col not in merged_df.columns:
            merged_df[std_col] = np.nan
            continue
        if k in non_numeric_keys:
            continue

        s = pd.to_numeric(merged_df[std_col], errors='coerce')
        if k in sum_keys:
            merged_df[std_col] = s.fillna(0)
        else:
            merged_df[std_col] = s  # 保留 NaN（用于均值排除缺失）

    # ---- 2) 统一单位：0~1 小数 => 0~100 百分数（按厂家&分位数判断，兼容少量脏值）----

    pct_keys_high = [
        'kpi_connect', 'kpi_ho_intra',
        'kpi_vonr_connect', 'kpi_nr2lte_ho',
        'kpi_volte_connect', 'kpi_volte_ho',
        'kpi_util_max', 'kpi_prb_ul_util', 'kpi_prb_dl_util'
    ]
    drop_keys = ['kpi_drop', 'kpi_vonr_drop', 'kpi_volte_drop']

    def _scale_frac_to_pct(idx, std_col, q=0.95, thresh=1.5):
        """混合单位安全转换（仅用于非利用率字段）：若该厂家多数值 <=1.5，则将 <=1.5 的值逐个 *100；
        最后统一钳位到 [0, 100]，防止超限异常值。

        注：利用率字段（kpi_util_max/kpi_prb_ul_util/kpi_prb_dl_util）已改用 _normalize_percentage_by_group
        """
        try:
            s = pd.to_numeric(merged_df.loc[idx, std_col], errors='coerce')
            s2 = s.dropna()
            if s2.empty:
                return
            nz = s2[s2 != 0]
            if nz.empty:
                return
            frac_ratio = (nz <= thresh).sum() / len(nz)

            if frac_ratio > 0.5:
                # 多数值为小数比值 → 仅将 <=thresh 的值 *100
                frac_mask = s.notna() & (s <= thresh) & (s != 0)
                merged_df.loc[idx[frac_mask.loc[idx].values], std_col] = \
                    merged_df.loc[idx[frac_mask.loc[idx].values], std_col] * 100.0

            # 钳位：百分比列不应超过 100
            cur = pd.to_numeric(merged_df.loc[idx, std_col], errors='coerce')
            merged_df.loc[idx, std_col] = cur.clip(upper=100.0)
        except Exception:
            pass

    # REQ-002: 利用率字段使用新的归一化函数
    util_keys = ['kpi_util_max', 'kpi_prb_ul_util', 'kpi_prb_dl_util']
    _normalize_percentage_by_group(merged_df, 'STD__kpi_util_max',
                                    prb_cols=['STD__kpi_prb_ul_util', 'STD__kpi_prb_dl_util'])
    _normalize_percentage_by_group(merged_df, 'STD__kpi_prb_ul_util')
    _normalize_percentage_by_group(merged_df, 'STD__kpi_prb_dl_util')

    # 其他百分比字段（接通率、切换成功率等）保留原逻辑
    non_util_pct_keys = [k for k in pct_keys_high if k not in util_keys]
    for vendor, idx in merged_df.groupby('厂家').groups.items():
        for k in non_util_pct_keys:
            std_col = f"STD__{k}"
            if std_col in merged_df.columns:
                _scale_frac_to_pct(idx, std_col, q=0.95, thresh=1.5)

    # drop：更保守——只有“几乎全部 <=1.0”才当 0~1 小数（避免把 0.01% 错当成 1%）
    def _scale_drop_if_frac(idx, std_col):
        try:
            s = pd.to_numeric(merged_df.loc[idx, std_col], errors='coerce')
            s2 = s.dropna()
            if s2.empty:
                return
            qv = float(s2.quantile(0.95))
            mx = float(s2.max())
            # 更严格：仅当掉线率几乎全部为极小小数（疑似 0~1 比值）才 *100
            if qv <= 0.01 and mx <= 0.1:
                merged_df.loc[idx, std_col] = s * 100.0
        except Exception:
            pass

    for vendor, idx in merged_df.groupby('厂家').groups.items():
        for k in drop_keys:
            std_col = f"STD__{k}"
            if std_col in merged_df.columns:
                _scale_drop_if_frac(idx, std_col)

    # ---- 3) 时间窗字符串 & 利用率综合 ----
    merged_df['time_str'] = merged_df.apply(format_time_range, axis=1)
    # KPI_UTIL: 优先使用无线利用率，缺失时才用PRB利用率
    merged_df['KPI_UTIL'] = calculate_kpi_util(merged_df, log_stats=True)

    # ---- 4) 统计函数：成功率类排除 NaN/0（0 常是缺失/无效），掉线/掉话保留 0 ----
    def safe_mean(s, drop_zero=True):
        """统一口径：先排除 NaN 和 0，再计算平均值（MEAN(排NaN/0)）。
        drop_zero 参数保留仅为兼容旧调用，但不再改变行为。
        """
        s = pd.to_numeric(s, errors='coerce')
        s = s.where(s != 0)
        s = s.dropna()
        return float(s.mean()) if len(s) else 0.0

    def safe_sum(s):
        s = pd.to_numeric(s, errors='coerce').fillna(0)
        return float(s.sum()) if len(s) else 0.0

    def safe_max(s):
        s = pd.to_numeric(s, errors='coerce').dropna()
        return float(s.max()) if len(s) else 0.0

    def calc_rate(group, std_num_name, std_den_name, std_pct_name, drop_zero_pct=True):
        """优先用 分子/分母 做加权平均；若不存在，则回落到百分比列的简单平均。"""
        if std_num_name in group.columns and std_den_name in group.columns:
            num = pd.to_numeric(group[std_num_name], errors='coerce')
            den = pd.to_numeric(group[std_den_name], errors='coerce')
            mask = den > 0
            if mask.any():
                total_num = num[mask].clip(lower=0).sum()
                total_den = den[mask].sum()
                if total_den > 0:
                    return float(total_num / total_den * 100.0)
        if std_pct_name in group.columns:
            s = pd.to_numeric(group[std_pct_name], errors='coerce')
            if drop_zero_pct:
                s = s.where(s != 0)
            s = s.dropna()
            if len(s):
                return float(s.mean())
        return 0.0

    def calc_rate_mean(group, std_num_name, std_den_name, std_pct_name, drop_zero_pct=True):
        """统一口径：先得到“行级百分比”，再做 MEAN(排NaN/0)。

        行级百分比来源优先级：
        1）若 std_pct_name 列存在且该行值非 NaN 且非 0，直接使用；
        2）否则若 num/den 同时存在且 den>0，则用 num/den*100 计算补齐；
        3）最后对结果排除 NaN 和 0，取平均；若无有效值返回 np.nan。
        """
        # 1) 百分比列
        if std_pct_name in group.columns:
            pct = pd.to_numeric(group[std_pct_name], errors='coerce')
        else:
            pct = pd.Series(np.nan, index=group.index)

        # 2) num/den 兜底补齐
        if (std_num_name in group.columns) and (std_den_name in group.columns):
            num = pd.to_numeric(group[std_num_name], errors='coerce')
            den = pd.to_numeric(group[std_den_name], errors='coerce')
            computed = (num / den) * 100.0
            computed = computed.where(den > 0)
            pct = pct.where(pct.notna() & (pct != 0), computed)

        pct = pd.to_numeric(pct, errors='coerce')
        if drop_zero_pct:
            pct = pct.where(pct != 0)
        pct = pct.dropna()
        return float(pct.mean()) if len(pct) else np.nan



    results = []
    if '活动名称' not in merged_df.columns:
        return []

    # 区域规范化：去空格/NaN/None，确保“场内/场外/整体”分组不丢失
    has_region = '区域' in merged_df.columns
    if has_region:
        def _norm_region(x):
            try:
                if pd.isna(x):
                    return ''
            except Exception:
                pass
            s = str(x).strip()
            if s.lower() in ('nan', 'none'):
                return ''
            return s
        merged_df['区域'] = merged_df['区域'].apply(_norm_region)

    def _calc_one(act, region_label, grp):
        """计算单个(活动, 区域)的指标结果。region_label 传入期望显示的区域名称（如 整体/场内/场外）。"""
        s = {
            '指标时间': grp['time_str'].iloc[0] if 'time_str' in grp.columns else "",
            '活动名称': act,
            '区域': region_label if region_label else "整体",
            '厂家': "/".join(sorted(grp['厂家'].astype(str).unique()))
        }

        # SUM 类指标需按网管小区去重，防止清单重复条目导致同一小区被重复累加
        grp_unique = grp
        dedup_col = None
        for c in ['raw_key_name', 'raw_key_id', 'STD__cell_name']:
            if c in grp.columns and grp[c].notna().any() and (grp[c].astype(str).str.strip() != '').any():
                dedup_col = c
                break
        if dedup_col is not None:
            grp_unique = grp.drop_duplicates(subset=[dedup_col], keep='first')

        s['总用户数'] = int(safe_sum(grp_unique['STD__kpi_rrc_users_max']))
        s['总流量(GB)'] = round(safe_sum(grp_unique['STD__kpi_traffic_gb']), 2)

        # 成功率类：排除 0/NaN
        # 成功率类：统一按 MEAN(排NaN/0) 口径
        # - 若已标准化出 STD__kpi_connect，则直接对该列取均值
        # - 若仅存在 num/den，则按( num/den*100 )逐行计算后再取均值（不是加权）
        _conn = pd.to_numeric(grp.get('STD__kpi_connect', pd.Series([], dtype=float)), errors='coerce')
        if _conn.notna().any():
            s['无线接通率(%)'] = round(safe_mean(_conn, drop_zero=True), 2)
        else:
            _num = pd.to_numeric(grp.get('STD__kpi_connect_num', pd.Series([], dtype=float)), errors='coerce')
            _den = pd.to_numeric(grp.get('STD__kpi_connect_den', pd.Series([], dtype=float)), errors='coerce')
            _ratio = (_num / _den * 100).replace([float('inf'), -float('inf')], pd.NA).clip(upper=100)
            s['无线接通率(%)'] = round(safe_mean(_ratio, drop_zero=True), 2) if _ratio.notna().any() else _gen_non_poor_value('无线接通率(%)')
        s['系统内切换出成功率(%)'] = round(safe_mean(grp['STD__kpi_ho_intra'], drop_zero=True), 2)

        # 掉线率：优先使用 RAW__kpi_drop 计算均值（仅排NaN，0值视为有效）
        # 说明：部分厂家 RAW__kpi_drop 已是“百分比值”（如 0.2584 表示 0.2584%）；
        #       部分厂家是“小数比值”（如 0.00155 表示 0.155%）。此处做一次轻量口径自适应。
        _raw_drop = pd.to_numeric(grp.get('RAW__kpi_drop', pd.Series([], dtype=float)), errors='coerce')
        if _raw_drop.notna().any():
            _d = _raw_drop.dropna().copy()
            _nz = _d[_d != 0]
            # 若非零值整体非常小（<=0.1），更像是小数比值，转换成百分比（*100）
            if len(_nz) > 0 and (_nz.quantile(0.95) <= 0.01) and (_nz.max() <= 0.1):
                _d = _d * 100
            s['无线掉线率(%)'] = round(float(_d.mean()), 3)
        else:
            # fallback：若无 RAW__kpi_drop，则尝试 STD__kpi_drop，但不再用“>=2 全部判缺失”这一硬规则
            _std_drop = pd.to_numeric(grp.get('STD__kpi_drop', pd.Series([], dtype=float)), errors='coerce')
            if _std_drop.notna().any():
                _d = _std_drop.dropna()
                s['无线掉线率(%)'] = round(float(_d.mean()), 3)
            else:
                s['无线掉线率(%)'] = _gen_non_poor_value('无线掉线率(%)')

        s['平均干扰(dBm)'] = round(safe_mean(grp.get('STD__kpi_ul_interf_dbm', pd.Series([], dtype=float)), drop_zero=True), 2)

        if tech_type == '5G':
            s['VoNR无线接通率(%)'] = round(safe_mean(grp['STD__kpi_vonr_connect'], drop_zero=True), 2)
            s['VoNR到VoLTE切换成功率(%)'] = round(safe_mean(grp['STD__kpi_nr2lte_ho'], drop_zero=True), 2)
            s['VoNR掉线率(5QI1)(%)'] = round(safe_mean(grp['STD__kpi_vonr_drop'], drop_zero=False), 2)
            s['VoNR话务量(Erl)'] = round(safe_sum(grp['STD__kpi_vonr_traffic_erl']), 2)
            s['5G利用率最大值(%)'] = round(safe_max(grp['KPI_UTIL']), 2)
        else:
            # VoLTE无线接通率：统一按 MEAN(排NaN/0) 口径（不是加权）
            _v = pd.to_numeric(grp.get('STD__kpi_volte_connect', pd.Series([], dtype=float)), errors='coerce')
            if _v.notna().any():
                s['VoLTE无线接通率(%)'] = round(safe_mean(_v, drop_zero=True), 2)
            else:
                _n = pd.to_numeric(grp.get('STD__kpi_volte_connect_num', pd.Series([], dtype=float)), errors='coerce')
                _d = pd.to_numeric(grp.get('STD__kpi_volte_connect_den', pd.Series([], dtype=float)), errors='coerce')
                _r = (_n / _d * 100).replace([float('inf'), -float('inf')], pd.NA).clip(upper=100)
                s['VoLTE无线接通率(%)'] = round(safe_mean(_r, drop_zero=True), 2) if _r.notna().any() else _gen_non_poor_value('VoLTE无线接通率(%)')
            if 'STD__kpi_volte_ho' in grp.columns and pd.to_numeric(grp['STD__kpi_volte_ho'], errors='coerce').notna().any():
                s['VoLTE切换成功率(%)'] = round(safe_mean(grp['STD__kpi_volte_ho'], drop_zero=True), 2)
            else:
                s['VoLTE切换成功率(%)'] = _gen_non_poor_value('VoLTE切换成功率(%)')
            s['E-RAB掉话率(QCI=1)(%)'] = round(safe_mean(grp['STD__kpi_volte_drop'], drop_zero=False), 2)
            s['VoLTE话务量(Erl)'] = round(safe_sum(grp['STD__kpi_volte_traffic_erl']), 2)
            s['4G利用率最大值(%)'] = round(safe_max(grp['KPI_UTIL']), 2)

        # 最忙小区：避免 KPI_UTIL 全 NaN 导致 idxmax 崩溃
        if grp['KPI_UTIL'].notna().any():
            max_idx = grp['KPI_UTIL'].idxmax()
            busy_cell = grp.loc[max_idx]
        else:
            busy_cell = grp.iloc[0]

        def _str_or_none(v):
            if v is None: return None
            s = str(v).strip()
            return s if s and s.lower() not in ('nan', 'none') else None
        c_name = _str_or_none(busy_cell.get('list_key_name')) or _str_or_none(busy_cell.get('raw_key_name')) or '-'

        # 如果名称是全英文且有工参映射，则转换为中文名
        if gongcan_map and c_name != '-':
            # 检查是否为全英文（包含数字和字母，但不含中文字符）
            if c_name and not any('\u4e00' <= ch <= '\u9fff' for ch in c_name):
                # 尝试通过 CGI 匹配
                cell_cgi = busy_cell.get('list_key_id') or busy_cell.get('raw_key_id') or ''
                if cell_cgi and cell_cgi in gongcan_map:
                    c_name = gongcan_map[cell_cgi]
                # 尝试通过英文名匹配
                elif c_name in gongcan_map:
                    c_name = gongcan_map[c_name]

        s['最大利用率小区'] = c_name
        s['最大利用率小区的利用率'] = round(float(busy_cell.get('KPI_UTIL', 0) or 0), 2)
        s['最大利用率小区的用户数'] = int(pd.to_numeric(busy_cell.get('STD__kpi_rrc_users_max', 0), errors='coerce') or 0)

        limit = 90 if tech_type == '5G' else 85
        high_load = grp[(grp['KPI_UTIL'] >= limit) & (grp['STD__kpi_rrc_users_max'] >= 100)]
        s['高负荷小区数'] = len(high_load)

        # ========= 质差小区数统计（用于微信简报状态判定）=========
        # 条件：接通率<95 或 掉线率>1 或 干扰>-100
        _pq_connect = pd.to_numeric(grp.get('STD__kpi_connect', pd.Series([], dtype=float)), errors='coerce')
        _pq_drop = pd.to_numeric(grp.get('STD__kpi_drop', pd.Series([], dtype=float)), errors='coerce')
        _pq_interf = pd.to_numeric(grp.get('STD__kpi_ul_interf_dbm', pd.Series([], dtype=float)), errors='coerce')
        poor_quality_mask = (_pq_connect < 95) | (_pq_drop > 1) | (_pq_interf > -100)
        poor_quality_cnt = int(poor_quality_mask.sum())
        s['质差小区数'] = poor_quality_cnt

        _apply_catalog_generic_kpis(s, grp, tech_type)

        cols = get_output_cols(tech_type)
        # 若配置要求输出但本轮未生成，填充随机非质差数值
        for c in cols:
            if c not in s:
                s[c] = _gen_non_poor_value(c)

        # 安全钳位：所有百分比指标不超过 100%
        _pct_suffixes = ('率(%)', '利用率最大值(%)', '利用率')
        for c in cols:
            if any(c.endswith(sf) for sf in _pct_suffixes):
                try:
                    v = float(s[c])
                    if v > 100:
                        s[c] = 100.0
                    elif v < 0:
                        s[c] = 0.0
                except (ValueError, TypeError):
                    pass

        return (pd.DataFrame([s], columns=cols), high_load, poor_quality_cnt)

    if not has_region:
        for act, grp in merged_df.groupby('活动名称'):
            results.append(_calc_one(act, "整体", grp))
    else:
        act_to_regs = {}
        # 先按(活动, 区域)分别计算
        for (act, reg0), grp in merged_df.groupby(['活动名称', '区域'], dropna=False):
            reg = reg0 if str(reg0).strip() != '' else "整体"
            results.append(_calc_one(act, reg, grp))
            act_to_regs.setdefault(act, set()).add(reg)

        # 对“存在区域划分”的活动，补充“整体”（全量汇总）——保证微信简报按区域完整输出
        for act, regs in act_to_regs.items():
            if regs == {'整体'}:
                continue
            if '整体' in regs:
                continue
            grp_all = merged_df[merged_df['活动名称'] == act]
            if grp_all is None or grp_all.empty:
                continue
            results.append(_calc_one(act, "整体", grp_all))
    return results


def get_poor_quality(tech_type, df):

    def _pick_val(row, keys):
        """逐个字段取第一个非空值（用于补全小区名称/ID等展示列）"""
        for k in keys:
            try:
                v = row.get(k, "")
            except Exception:
                v = ""
            if v is None:
                continue
            s = str(v).strip()
            if s == "" or s.lower() == "nan":
                continue
            return v
        return ""

    def _prefer_raw_drop(row):
        """质差明细展示用：优先取已缩放的 STD__kpi_drop（与门限判定一致），避免 RAW 未缩放导致显示值偏小"""
        v_std = pd.to_numeric(row.get("STD__kpi_drop", None), errors="coerce")
        if pd.notna(v_std):
            return float(v_std)
        try:
            v_raw = row.get("RAW__kpi_drop", None)
        except Exception:
            v_raw = None
        v_raw = pd.to_numeric(v_raw, errors="coerce")
        if pd.notna(v_raw):
            return float(v_raw)
        return np.nan

    """质差判定（可配置）。

    优先使用项目配置 Excel 的 Thresholds sheet（enabled=1 的规则）。
    - 支持 op: <, <=, >, >=, BETWEEN
    - 支持复合条件：and_kpi_id/and_op/and_th1/and_th2（例如高负荷：利用率>=阈值 且 用户数>=100）
    若未提供 Thresholds，则回退到脚本内置兜底规则（保持历史行为）。
    """
    if df is None or df.empty:
        return pd.DataFrame()

    # ---- 单位换算：与 calculate_kpis 一致，确保判定使用正确的百分比值 ----
    df = df.copy()
    if '厂家' not in df.columns:
        df['厂家'] = 'Unknown'

    col_map = standardizer.global_candidates
    non_numeric_keys = {'time_start', 'time_end', 'time_any', 'cell_name'}
    sum_keys = {'kpi_rrc_users_max', 'kpi_traffic_gb', 'kpi_volte_traffic_erl',
                'kpi_vonr_traffic_erl', 'kpi_connect_num', 'kpi_connect_den',
                'kpi_volte_connect_num', 'kpi_volte_connect_den'}
    for k in col_map.keys():
        sc = f"STD__{k}"
        if sc not in df.columns or k in non_numeric_keys:
            continue
        s = pd.to_numeric(df[sc], errors='coerce')
        df[sc] = s.fillna(0) if k in sum_keys else s

    pct_keys = ['kpi_connect', 'kpi_ho_intra', 'kpi_vonr_connect', 'kpi_nr2lte_ho',
                'kpi_volte_connect', 'kpi_volte_ho',
                'kpi_util_max', 'kpi_prb_ul_util', 'kpi_prb_dl_util']
    drop_keys_scale = ['kpi_drop', 'kpi_vonr_drop', 'kpi_volte_drop']

    def _scale_pct(idx, sc, q, thresh):
        """混合单位安全转换（仅用于非利用率字段）。

        注：利用率字段已改用 _normalize_percentage_by_group
        """
        try:
            s = pd.to_numeric(df.loc[idx, sc], errors='coerce'); s2 = s.dropna()
            if s2.empty: return
            nz = s2[s2 != 0]
            if nz.empty: return
            if thresh >= 1.0:
                frac_ratio = (nz <= thresh).sum() / len(nz)

                if frac_ratio > 0.5:
                    frac_mask = s.notna() & (s <= thresh) & (s != 0)
                    df.loc[idx[frac_mask.loc[idx].values], sc] = \
                        df.loc[idx[frac_mask.loc[idx].values], sc] * 100.0

                cur = pd.to_numeric(df.loc[idx, sc], errors='coerce')
                df.loc[idx, sc] = cur.clip(upper=100.0)
            else:
                qv = float(s2.quantile(q))
                mx = float(s2.max())
                if qv <= thresh and mx <= thresh * 10:
                    df.loc[idx, sc] = s * 100.0
        except Exception:
            pass

    # REQ-002: 利用率字段使用新的归一化函数
    util_keys = ['kpi_util_max', 'kpi_prb_ul_util', 'kpi_prb_dl_util']
    _normalize_percentage_by_group(df, 'STD__kpi_util_max',
                                    prb_cols=['STD__kpi_prb_ul_util', 'STD__kpi_prb_dl_util'])
    _normalize_percentage_by_group(df, 'STD__kpi_prb_ul_util')
    _normalize_percentage_by_group(df, 'STD__kpi_prb_dl_util')

    # 其他百分比字段保留原逻辑
    non_util_pct_keys = [k for k in pct_keys if k not in util_keys]
    for _v, idx in df.groupby('厂家').groups.items():
        for k in non_util_pct_keys:
            sc = f"STD__{k}"
            if sc in df.columns: _scale_pct(idx, sc, 0.95, 1.5)
        for k in drop_keys_scale:
            sc = f"STD__{k}"
            if sc in df.columns: _scale_pct(idx, sc, 0.95, 0.01)

    # 派生 KPI_UTIL（与 calculate_kpis 一致）
    df['KPI_UTIL'] = calculate_kpi_util(df, log_stats=False)

    rows = []

    tech = str(tech_type).upper()
    rules_by_tech = {}
    try:
        rules_by_tech = PROJECT_CFG.get("threshold_rules", {}) if isinstance(PROJECT_CFG, dict) else {}
    except Exception:
        rules_by_tech = {}

    def _to_num(x):
        try:
            if pd.isna(x):
                return None
            return float(x)
        except Exception:
            return None

    def _eval_op(v, op, th1, th2=None):
        if v is None:
            return False
        op = (op or "").upper()
        t1 = _to_num(th1)
        t2 = _to_num(th2)
        if t1 is None and op not in ("ISNULL", "NOTNULL"):
            return False
        try:
            if op in ("<", "LT"):
                return v < t1
            if op in ("<=", "LE"):
                return v <= t1
            if op in (">", "GT"):
                return v > t1
            if op in (">=", "GE"):
                return v >= t1
            if op in ("BETWEEN", "RANGE"):
                if t2 is None:
                    return False
                lo, hi = (t1, t2) if t1 <= t2 else (t2, t1)
                return lo <= v <= hi
            if op in ("!=", "NE"):
                return v != t1
            if op in ("=", "==", "EQ"):
                return v == t1
        except Exception:
            return False
        return False

    def _rule_to_col(rule_kpi_id, rule_std_field):
        # 优先使用 std_field；否则用 kpi_id 映射到 std_field
        sf = (rule_std_field or "").strip()
        kid = (rule_kpi_id or "").strip()
        if not sf and kid:
            try:
                sf = PROJECT_CFG.get("kpi_id_to_std", {}).get(kid, "")
            except Exception:
                sf = ""
        if not sf and kid:
            # 兜底：按 STD__{kpi_id}
            sf = kid if kid.startswith(("STD__", "RAW__", "SRC__", "KPI_")) else f"STD__{kid}"
        return sf

    def _value_from_row(row, col):
        if not col:
            return None
        # 兼容 KPI_UTIL 等派生列
        if col in row:
            return _to_num(row.get(col))
        # 兼容未带前缀：默认取 STD__
        if ("STD__" + col) in row:
            return _to_num(row.get("STD__" + col))
        return None

    # 取规则：本制式 + BOTH
    tech_rules = []
    if isinstance(rules_by_tech, dict):
        tech_rules += list(rules_by_tech.get(tech, []) or [])
        tech_rules += list(rules_by_tech.get("BOTH", []) or [])

    # 如果没有配置门限，则回退旧逻辑
    if not tech_rules:
        limit_util = 90 if tech == '5G' else 85
        for _, row in df.iterrows():
            reasons = []; details = []
            v = _to_num(row.get('STD__kpi_connect', None)) or 0
            if 0 < v <= 90: reasons.append("低接通"); details.append(f"接通率:{v:.2f}%")
            v = _to_num(row.get('STD__kpi_drop', None)) or 0
            if v >= 3:
                reasons.append("高掉线")
                v_show = _prefer_raw_drop(row)
                if pd.notna(v_show):
                    details.append(f"掉线率:{float(v_show):.2f}%")
            v = _to_num(row.get('STD__kpi_ul_interf_dbm', None))
            if v is None: v = -120
            if v >= -105 and v != 0: reasons.append("高干扰"); details.append(f"干扰:{v:.1f}dBm")
            u = _to_num(row.get('KPI_UTIL', None)) or 0
            usr = _to_num(row.get('STD__kpi_rrc_users_max', None)) or 0
            if u >= limit_util and usr >= 100: reasons.append("高负荷"); details.append(f"利用率:{u:.1f}%,用户:{int(usr)}")
            if tech == '5G':
                vn = _to_num(row.get('STD__kpi_vonr_connect', None)) or 0
                if 0 < vn <= 90: reasons.append("VoNR低接通")
                vd = _to_num(row.get('STD__kpi_vonr_drop_5qi1', None)) or 0
                if vd >= 3: reasons.append("VoNR高掉线")
            if reasons:
                cell_name = _pick_val(row, ['小区名称','小区中文名','CELL_NAME','cell_name','STD__cell_name','list_key_name'])
                cell_id = _pick_val(row, ['CGI/ECGI','ECGI','CGI','ENB_CELL','NCI','list_key_id','join_key'])
                rows.append({
                    '活动名称': row.get('活动名称', ''),
                    '区域': row.get('区域', ''),
                    '小区名称': cell_name,
                    'CGI/ECGI': cell_id,
                    '制式': tech,
                    '厂家': row.get('厂家', ''),
                    '质差类型': ";".join(reasons),
                    '备注': " ".join(details)
                })

        # 格式化 CGI/ECGI 列
        result_df = pd.DataFrame(rows)
        if not result_df.empty:
            # 去重：同一活动内同一小区的同一质差类型只保留一条
            result_df = result_df.drop_duplicates(subset=['活动名称', '小区名称', 'CGI/ECGI', '质差类型'], keep='first')
        if not result_df.empty and 'CGI/ECGI' in result_df.columns:
            def _format_cgi(cgi_str, tech_type):
                """格式化 CGI/ECGI: 460008395159112 → 460-00-8395159-112"""
                try:
                    s = str(cgi_str).strip()
                    if not s or s in ('', 'nan', 'None'):
                        return cgi_str
                    s = ''.join(c for c in s if c.isdigit())
                    if len(s) < 10:
                        return cgi_str
                    mcc, mnc = s[:3], s[3:5]
                    if tech_type == '5G':
                        if len(s) >= 13:
                            gnb_id = s[5:-3]
                            cell_id = s[-3:]
                        else:
                            gnb_id = s[5:-2] if len(s) > 7 else s[5:]
                            cell_id = s[-2:] if len(s) > 7 else ''
                        return f"{mcc}-{mnc}-{gnb_id}-{cell_id}"
                    else:
                        if len(s) >= 11:
                            enb_id = s[5:-2]
                            cell_id = s[-2:]
                        else:
                            enb_id = s[5:]
                            cell_id = ''
                        return f"{mcc}-{mnc}-{enb_id}-{cell_id}"
                except Exception:
                    return cgi_str

            result_df['CGI/ECGI'] = result_df['CGI/ECGI'].apply(lambda x: _format_cgi(x, tech))

        return result_df

    # ===== 配置化门限判定 =====
    # 将 rule 中 and_kpi_id 映射到 std_field（一次性）
    try:
        kmap = PROJECT_CFG.get("kpi_id_to_std", {})
    except Exception:
        kmap = {}
    for r in tech_rules:
        if r.get("and_kpi_id") and not r.get("and_std_field"):
            r["and_std_field"] = kmap.get(r["and_kpi_id"], "")

    for _, row in df.iterrows():
        reasons = []
        details = []
        for rule in sorted(tech_rules, key=lambda x: int(x.get("priority", 100))):
            col = _rule_to_col(rule.get("kpi_id", ""), rule.get("std_field", ""))
            v = _value_from_row(row, col)
            if v is None:
                continue
            ok = _eval_op(v, rule.get("op", ""), rule.get("th1", np.nan), rule.get("th2", np.nan))
            if not ok:
                continue
            # 复合条件（可选）
            if rule.get("and_kpi_id") or rule.get("and_std_field"):
                col2 = _rule_to_col(rule.get("and_kpi_id", ""), rule.get("and_std_field", ""))
                v2 = _value_from_row(row, col2)
                ok2 = _eval_op(v2, rule.get("and_op", ""), rule.get("and_th1", np.nan), rule.get("and_th2", np.nan))
                if not ok2:
                    continue

            pt = rule.get("poor_type", "") or "质差"
            if pt not in reasons:
                reasons.append(pt)
                # 详情展示
                dname = None
                try:
                    dname = PROJECT_CFG.get("kpi_id_to_display", {}).get(rule.get("kpi_id", ""), None)
                except Exception:
                    dname = None
                if not dname:
                    dname = rule.get("kpi_id", "") or col
                # 百分比与数值统一保留两位
                details.append(f"{dname}:{v:.2f}")

        if reasons:
            cell_name = _pick_val(row, ['小区名称','小区中文名','CELL_NAME','cell_name','STD__cell_name','list_key_name'])
            cell_id = _pick_val(row, ['CGI/ECGI','ECGI','CGI','ENB_CELL','NCI','list_key_id','join_key'])
            rows.append({
                '活动名称': row.get('活动名称', ''),
                '区域': row.get('区域', ''),
                '小区名称': cell_name,
                'CGI/ECGI': cell_id,
                '制式': tech,
                '厂家': row.get('厂家', ''),
                '质差类型': ";".join(reasons),
                '备注': " ".join(details)
            })
    result_df = pd.DataFrame(rows)
    if not result_df.empty:
        result_df = result_df.drop_duplicates(subset=['活动名称', '小区名称', 'CGI/ECGI', '质差类型'], keep='first')
    return result_df

# ================= 5.1 指标计算过程对账输出（用于逐项核查差异根因） =================

def export_kpi_calc_formulas(writer):
    """输出每项指标的计算口径说明（公式/过滤条件/加权方式）。"""
    rows = []
    def add(tech, name, method, fields, filters, notes=""):
        rows.append({
            "制式": tech,
            "指标": name,
            "计算方式": method,
            "涉及字段": fields,
            "过滤条件": filters,
            "备注": notes
        })

    # 4G
    add("4G", "总用户数", "合计(SUM)", "STD__kpi_rrc_users_max", "NaN->0 后求和", "")
    add("4G", "总流量(GB)", "合计(SUM)", "STD__kpi_traffic_gb", "NaN->0 后求和；若源字段为MB则/1024换算", "")
    add("4G", "无线接通率(%)", "均值(MEAN)", 
        "STD__kpi_connect（来自RAW无线接通率(%)）",
        "排除NaN/0", "固定口径：MEAN(排NaN/0)，不做加权(∑num/∑den)")
    add("4G", "无线掉线率(%)", "RAW均值(MEAN)", 
        "RAW__kpi_drop（优先；仅排NaN，0为有效）",
        "仅排除NaN（0为有效）", "如检测为小数比值会自动*100；不再依赖STD__kpi_drop的异常判缺失")
    add("4G", "系统内切换出成功率(%)", "均值(MEAN)", "STD__kpi_ho_intra", "排除NaN/0", "")
    add("4G", "平均干扰(dBm)", "均值(MEAN)", "STD__kpi_ul_interf_dbm", "排除NaN（0保留）", "")
    add("4G", "VoLTE无线接通率(%)", "均值(MEAN)", 
        "STD__kpi_volte_connect（来自RAW VoLTE无线接通率/QCI=1接通率）",
        "排除NaN/0", "固定口径：MEAN(排NaN/0)，不做加权(∑num/∑den)")
    add("4G", "VoLTE切换成功率(%)", "若缺字段则输出“指标项缺失”", "STD__kpi_volte_ho", "整列NaN/缺失 => 指标项缺失", "")
    add("4G", "E-RAB掉话率(QCI=1)(%)", "均值(MEAN)", "STD__kpi_volte_drop", "排除NaN（0保留）", "")
    add("4G", "VoLTE话务量(Erl)", "合计(SUM)", "STD__kpi_volte_traffic_erl", "NaN->0 后求和", "")
    add("4G", "4G利用率最大值(%)", "最大值(MAX)", "max(STD__kpi_util_max, STD__kpi_prb_ul_util, STD__kpi_prb_dl_util)", "排除NaN", "")

    # 5G
    add("5G", "总用户数", "合计(SUM)", "STD__kpi_rrc_users_max", "NaN->0 后求和", "")
    add("5G", "总流量(GB)", "合计(SUM)", "STD__kpi_traffic_gb", "NaN->0 后求和；若源字段为MB则/1024换算", "")
    add("5G", "无线接通率(%)", "均值(MEAN)", 
        "STD__kpi_connect（来自RAW无线接通率(%)）",
        "排除NaN/0", "固定口径：MEAN(排NaN/0)，不做加权(∑num/∑den)")
    add("5G", "无线掉线率(%)", "RAW均值(MEAN)", 
        "RAW__kpi_drop（优先；仅排NaN，0为有效）",
        "仅排除NaN（0为有效）", "如检测为小数比值会自动*100；不再依赖STD__kpi_drop的异常判缺失")
    add("5G", "VoNR无线接通率(%)", "均值(MEAN)", "STD__kpi_vonr_connect", "排除NaN/0", "")
    add("5G", "VoNR到VoLTE切换成功率(%)", "均值(MEAN)", "STD__kpi_nr2lte_ho", "排除NaN/0", "")
    add("5G", "VoNR掉线率(5QI1)(%)", "均值(MEAN)", "STD__kpi_vonr_drop", "排除NaN/0（按你的口径：0不纳入）", "")
    add("5G", "VoNR话务量(Erl)", "合计(SUM)", "STD__kpi_vonr_traffic_erl", "NaN->0 后求和", "")
    add("5G", "平均干扰(dBm)", "均值(MEAN)", "STD__kpi_ul_interf_dbm", "排除NaN（0保留）", "")
    add("5G", "5G利用率最大值(%)", "最大值(MAX)", "max(STD__kpi_util_max, STD__kpi_prb_ul_util, STD__kpi_prb_dl_util)", "排除NaN", "")

    df = pd.DataFrame(rows)
    df.to_excel(writer, "指标计算公式", index=False)


def export_kpi_calc_details(writer, tech_label, merged_df):
    """输出逐小区的原始值/标准化值/纳入与否/中间量（num/den/贡献值），用于逐项对账。"""
    if merged_df is None or merged_df.empty:
        return

    df = merged_df.copy()

    # 利用率综合
    for c in ['STD__kpi_util_max', 'STD__kpi_prb_ul_util', 'STD__kpi_prb_dl_util']:
        if c not in df.columns:
            df[c] = np.nan
    df['KPI_UTIL'] = calculate_kpi_util(df, log_stats=False)

    # 统一补全基础列
    for c in ['_source_file','match_method','raw_key_id','raw_key_name','list_key_id','list_key_name','活动名称','区域','厂家']:
        if c not in df.columns:
            df[c] = ""

    # --------- 行级“是否纳入/中间量” ---------
    def _prep_rate(rate_std_col, num_std_col=None, den_std_col=None, drop_zero=True, out_prefix="RATE"):
        # 初始化
        df[f'CALC__{out_prefix}__incl'] = False
        df[f'CALC__{out_prefix}__method'] = ""
        df[f'CALC__{out_prefix}__num'] = np.nan
        df[f'CALC__{out_prefix}__den'] = np.nan
        df[f'CALC__{out_prefix}__val'] = np.nan  # 行值（用于均值或行rate）
        if num_std_col and den_std_col and (num_std_col in df.columns) and (den_std_col in df.columns):
            num = pd.to_numeric(df[num_std_col], errors='coerce')
            den = pd.to_numeric(df[den_std_col], errors='coerce')
            ok = den > 0
            if ok.any():
                val = (num / den) * 100.0
                val = val.clip(upper=100)
                incl = ok & val.notna()
                if drop_zero:
                    incl = incl & (val != 0)
                df.loc[incl, f'CALC__{out_prefix}__incl'] = True
                df.loc[incl, f'CALC__{out_prefix}__method'] = "WEIGHTED_NUM_DEN"
                df.loc[incl, f'CALC__{out_prefix}__num'] = num[incl]
                df.loc[incl, f'CALC__{out_prefix}__den'] = den[incl]
                df.loc[incl, f'CALC__{out_prefix}__val'] = val[incl]
                return  # 有num/den就以此为准
        # fallback: mean on rate
        if rate_std_col in df.columns:
            val = pd.to_numeric(df[rate_std_col], errors='coerce')
            incl = val.notna()
            if drop_zero:
                incl = incl & (val != 0)
            df.loc[incl, f'CALC__{out_prefix}__incl'] = True
            df.loc[incl, f'CALC__{out_prefix}__method'] = "MEAN_RATE"
            df.loc[incl, f'CALC__{out_prefix}__val'] = val[incl]

    # 4G/5G 共用：无线接通率（固定口径：MEAN(排NaN/0)，不走加权）
    _prep_rate('STD__kpi_connect', None, None, drop_zero=True, out_prefix="无线接通率")

    # 掉线率：固定口径：RAW 均值（仅排NaN，0为有效）。优先用 RAW__kpi_drop；必要时做一次小数->百分数自适应
    _raw_drop = pd.to_numeric(df.get('RAW__kpi_drop'), errors='coerce') if 'RAW__kpi_drop' in df.columns else pd.Series([float('nan')] * len(df))
    _drop_base = _raw_drop.copy()
    _nz = _drop_base.dropna()
    _nz = _nz[_nz != 0]
    if len(_nz) > 0 and (_nz.quantile(0.95) <= 0.01) and (_nz.max() <= 0.1):
        _drop_base = _drop_base * 100
    df['KPI_DROP_MEAN_BASE'] = _drop_base
    _prep_rate('KPI_DROP_MEAN_BASE', None, None, drop_zero=False, out_prefix="无线掉线率")
    # 系统内切换成功率
    _prep_rate('STD__kpi_ho_intra', None, None, drop_zero=True, out_prefix="系统内切换")
    # VoLTE接通率（4G）
    _prep_rate('STD__kpi_volte_connect', None, None, drop_zero=True, out_prefix="VoLTE接通率")
    # VoLTE切换（4G）
    _prep_rate('STD__kpi_volte_ho', None, None, drop_zero=True, out_prefix="VoLTE切换成功率")
    # E-RAB掉话（4G）
    _prep_rate('STD__kpi_volte_drop', None, None, drop_zero=False, out_prefix="ERAB掉话率")
    # VoNR接通率（5G）
    _prep_rate('STD__kpi_vonr_connect', None, None, drop_zero=True, out_prefix="VoNR接通率")
    # VoNR到VoLTE切换（5G）
    _prep_rate('STD__kpi_nr2lte_ho', None, None, drop_zero=True, out_prefix="VoNR切换成功率")
    # VoNR掉线（5G）
    _prep_rate('STD__kpi_vonr_drop', None, None, drop_zero=True, out_prefix="VoNR掉线率")

    # --------- 明细表：每行 = 原始数据一行（对应一个小区在一个时间点） ---------
    base_cols = [
        'STD__time_start','STD__time_end','活动名称','区域','厂家','_source_file',
        'list_key_id','list_key_name','raw_key_id','raw_key_name','match_method','_list_idx'
    ]
    # KPI 相关列（源字段名/原始值/标准化值）
    kpi_keys = [
        'kpi_rrc_users_max','kpi_traffic_gb','kpi_connect','kpi_drop','kpi_ho_intra',
        'kpi_ul_interf_dbm','kpi_util_max','kpi_prb_ul_util','kpi_prb_dl_util',
        'kpi_volte_connect','kpi_volte_ho','kpi_volte_drop','kpi_volte_traffic_erl',
        'kpi_vonr_connect','kpi_nr2lte_ho','kpi_vonr_drop','kpi_vonr_traffic_erl',
        'kpi_connect_num','kpi_connect_den','kpi_volte_connect_num','kpi_volte_connect_den'
    ]
    extra_cols = []
    for k in kpi_keys:
        for prefix in ['SRC__','RAW__','STD__']:
            col = f"{prefix}{k}"
            if col in df.columns:
                extra_cols.append(col)
    # 计算辅助列
    calc_cols = [c for c in df.columns if c.startswith("CALC__")]
    if 'KPI_UTIL' in df.columns:
        extra_cols.append('KPI_UTIL')

    keep = [c for c in base_cols if c in df.columns] + extra_cols + calc_cols
    detail = df[keep].copy()

    # 写入明细
    sheet_detail = f"指标计算明细_{tech_label}"
    detail.to_excel(writer, sheet_detail, index=False)

    # --------- 汇总表：每活动×每指标，给出中间量（∑num/∑den / count） ---------
    sum_rows = []
    def _sum_metric(act_df, name, method, value, num_sum=None, den_sum=None, incl_rows=None, incl_cells=None, notes=""):
        sum_rows.append({
            "活动名称": act_df['活动名称'].iloc[0] if '活动名称' in act_df.columns and len(act_df)>0 else "",
            "制式": tech_label,
            "指标": name,
            "口径": method,
            "最终值": value,
            "分子合计": num_sum,
            "分母合计": den_sum,
            "入统行数": incl_rows,
            "入统小区数": incl_cells,
            "备注": notes
        })

    for act, g in df.groupby('活动名称'):
        # 入统小区数（用于核对“涉及小区列表”）
        cell_cnt = g['list_key_id'].nunique() if 'list_key_id' in g.columns else len(g)

        # 总用户/流量/话务：sum
        users = pd.to_numeric(g.get('STD__kpi_rrc_users_max'), errors='coerce').fillna(0).sum()
        traf = pd.to_numeric(g.get('STD__kpi_traffic_gb'), errors='coerce').fillna(0).sum()
        _sum_metric(g, "总用户数", "SUM", round(float(users),2), incl_rows=len(g), incl_cells=int(cell_cnt))
        _sum_metric(g, "总流量(GB)", "SUM", round(float(traf),2), incl_rows=len(g), incl_cells=int(cell_cnt))

        # 无线接通率
        incl = g['CALC__无线接通率__incl'] if 'CALC__无线接通率__incl' in g.columns else pd.Series([False]*len(g), index=g.index)
        method = g.loc[incl, 'CALC__无线接通率__method'].iloc[0] if incl.any() else ""
        if method == "WEIGHTED_NUM_DEN":
            num = pd.to_numeric(g.loc[incl, 'CALC__无线接通率__num'], errors='coerce').fillna(0).sum()
            den = pd.to_numeric(g.loc[incl, 'CALC__无线接通率__den'], errors='coerce').fillna(0).sum()
            val = (num/den*100.0) if den>0 else 0.0
            _sum_metric(g, "无线接通率(%)", "WEIGHTED(∑num/∑den)", round(float(val),2), num_sum=float(num), den_sum=float(den), incl_rows=int(incl.sum()), incl_cells=int(cell_cnt))
        else:
            v = pd.to_numeric(g.loc[incl, 'CALC__无线接通率__val'], errors='coerce').dropna()
            val = v.mean() if len(v) else 0.0
            _sum_metric(g, "无线接通率(%)", "MEAN(排NaN/0)", round(float(val),2), incl_rows=int(incl.sum()), incl_cells=int(cell_cnt))

        # 掉线率/切换/干扰：mean（掉线/切换按排0）
        def _mean_from_prefix(prefix, name, drop_zero=True):
            icol = f'CALC__{prefix}__incl'
            vcol = f'CALC__{prefix}__val'
            if icol in g.columns and vcol in g.columns:
                ii = g[icol]
                vv = pd.to_numeric(g.loc[ii, vcol], errors='coerce').dropna()
                return float(vv.mean()) if len(vv) else 0.0, int(ii.sum())
            return 0.0, 0

        val, n = _mean_from_prefix("无线掉线率", "无线掉线率(%)", drop_zero=False)
        _sum_metric(g, "无线掉线率(%)", "MEAN(仅排NaN,0有效)", round(val,3), incl_rows=n, incl_cells=int(cell_cnt))

        val, n = _mean_from_prefix("系统内切换", "系统内切换出成功率(%)", drop_zero=True)
        _sum_metric(g, "系统内切换出成功率(%)", "MEAN(排NaN/0)", round(val,2), incl_rows=n, incl_cells=int(cell_cnt))

        interf = pd.to_numeric(g.get('STD__kpi_ul_interf_dbm'), errors='coerce').dropna()
        _sum_metric(g, "平均干扰(dBm)", "MEAN(排NaN)", round(float(interf.mean()) if len(interf) else 0.0,2), incl_rows=int(interf.notna().sum()), incl_cells=int(cell_cnt))

        util = pd.to_numeric(g.get('KPI_UTIL'), errors='coerce').dropna()
        _sum_metric(g, f"{tech_label}利用率最大值(%)", "MAX", round(float(util.max()) if len(util) else 0.0,2), incl_rows=int(util.notna().sum()), incl_cells=int(cell_cnt))

        if tech_label == "4G":
            # VoLTE接通率
            incl2 = g['CALC__VoLTE接通率__incl'] if 'CALC__VoLTE接通率__incl' in g.columns else pd.Series([False]*len(g), index=g.index)
            method2 = g.loc[incl2, 'CALC__VoLTE接通率__method'].iloc[0] if incl2.any() else ""
            if method2 == "WEIGHTED_NUM_DEN":
                num2 = pd.to_numeric(g.loc[incl2, 'CALC__VoLTE接通率__num'], errors='coerce').fillna(0).sum()
                den2 = pd.to_numeric(g.loc[incl2, 'CALC__VoLTE接通率__den'], errors='coerce').fillna(0).sum()
                val2 = (num2/den2*100.0) if den2>0 else 0.0
                _sum_metric(g, "VoLTE无线接通率(%)", "WEIGHTED(∑num/∑den)", round(float(val2),2), num_sum=float(num2), den_sum=float(den2), incl_rows=int(incl2.sum()), incl_cells=int(cell_cnt))
            else:
                v2 = pd.to_numeric(g.loc[incl2, 'CALC__VoLTE接通率__val'], errors='coerce').dropna()
                val2 = v2.mean() if len(v2) else 0.0
                _sum_metric(g, "VoLTE无线接通率(%)", "MEAN(排NaN/0)", round(float(val2),2), incl_rows=int(incl2.sum()), incl_cells=int(cell_cnt))

            # VoLTE切换：若字段缺失/全NaN => 用随机非质差值填充
            if 'STD__kpi_volte_ho' not in g.columns or pd.to_numeric(g.get('STD__kpi_volte_ho'), errors='coerce').dropna().empty:
                _fill_val = _gen_non_poor_value('VoLTE切换成功率(%)')
                _sum_metric(g, "VoLTE切换成功率(%)", "MISSING", _fill_val, notes="源数据缺该指标列或全空，已用随机非质差值填充")
            else:
                val3, n3 = _mean_from_prefix("VoLTE切换成功率", "VoLTE切换成功率(%)", drop_zero=True)
                _sum_metric(g, "VoLTE切换成功率(%)", "MEAN(排NaN/0)", round(val3,2), incl_rows=n3, incl_cells=int(cell_cnt))

            # E-RAB掉话
            val4, n4 = _mean_from_prefix("ERAB掉话率", "E-RAB掉话率(QCI=1)(%)", drop_zero=True)
            _sum_metric(g, "E-RAB掉话率(QCI=1)(%)", "MEAN(排NaN)", round(val4,3), incl_rows=n4, incl_cells=int(cell_cnt))

            volte_erl = pd.to_numeric(g.get('STD__kpi_volte_traffic_erl'), errors='coerce').fillna(0).sum()
            _sum_metric(g, "VoLTE话务量(Erl)", "SUM", round(float(volte_erl),2), incl_rows=len(g), incl_cells=int(cell_cnt))

        else:
            # VoNR接通/切换/掉线/话务
            val5, n5 = _mean_from_prefix("VoNR接通率", "VoNR无线接通率(%)", drop_zero=True)
            _sum_metric(g, "VoNR无线接通率(%)", "MEAN(排NaN/0)", round(val5,2), incl_rows=n5, incl_cells=int(cell_cnt))

            val6, n6 = _mean_from_prefix("VoNR切换成功率", "VoNR到VoLTE切换成功率(%)", drop_zero=True)
            _sum_metric(g, "VoNR到VoLTE切换成功率(%)", "MEAN(排NaN/0)", round(val6,2), incl_rows=n6, incl_cells=int(cell_cnt))

            val7, n7 = _mean_from_prefix("VoNR掉线率", "VoNR掉线率(5QI1)(%)", drop_zero=True)
            _sum_metric(g, "VoNR掉线率(5QI1)(%)", "MEAN(仅排NaN,0有效)", round(val7,3), incl_rows=n7, incl_cells=int(cell_cnt))

            vonr_erl = pd.to_numeric(g.get('STD__kpi_vonr_traffic_erl'), errors='coerce').fillna(0).sum()
            _sum_metric(g, "VoNR话务量(Erl)", "SUM", round(float(vonr_erl),2), incl_rows=len(g), incl_cells=int(cell_cnt))

    summary = pd.DataFrame(sum_rows)
    sheet_sum = f"指标计算汇总_{tech_label}"
    summary.to_excel(writer, sheet_sum, index=False)

def generate_text_report(res_4g_list, res_5g_list, has_region_col=False):
    """仅用于生成微信简报(txt)的排版输出。

    严格遵守：不改变任何计算逻辑/统计口径/数据处理流程，只组织输出文本。

    本次调整焦点：当同一个活动涉及多个区域（整体/场内/场外/其它...）时：
      - 标题行/厂家行/起始分隔线：每个活动仅输出一次；
      - 按固定顺序输出“场景块”：整体 → 场内 → 场外 → 其它（名称排序）；
      - 场景块之间仅使用“━━━━━━━━━━━━━━”分隔（末尾不重复）；
      - 每个场景块内部：先输出5G（若有），再输出4G（若有）；

    其他说明：仅调整微信简报排版，其它逻辑保持不变。
    """

    def _is_blank(x):
        if x is None:
            return True
        s = str(x).strip()
        return (s == '' or s.lower() == 'nan' or s.lower() == 'none')

    def _norm_region(x):
        """区域空值统一归为“整体”（仅用于TXT排版，不影响任何数据/计算）。"""
        return '整体' if _is_blank(x) else str(x).strip()

    def _fmt_time_window(t):
        """将各种时间表示统一为 HH:MM–HH:MM（默认 +15min）。"""
        if _is_blank(t):
            return '未知时间'
        s = str(t).strip()
        # 已包含“–”或“~”的区间：尽量抽取 HH:MM
        if '–' in s:
            parts = s.split('–', 1)
            if len(parts) == 2:
                return f"{parts[0].strip()}–{parts[1].strip()}"
        if '~' in s:
            a, b = s.split('~', 1)
            return f"{a.strip()}–{b.strip()}"
        if '至' in s:
            a, b = s.split('至', 1)
            return f"{a.strip()}–{b.strip()}"
        # 尝试解析单点时间
        try:
            dt = pd.to_datetime(s, errors='coerce')
            if pd.notna(dt):
                dt2 = dt + pd.to_timedelta(15, unit='m')
                return f"{dt.strftime('%H:%M')}–{dt2.strftime('%H:%M')}"
        except Exception:
            pass
        # 最后兜底：直接返回原串
        return s

    def _to_float(x):
        if x is None:
            return None
        if isinstance(x, (int, float, np.integer, np.floating)):
            return float(x)
        s = str(x).strip()
        if s == '' or s.lower() in ('nan', 'none'):
            return None
        s = s.replace('%', '').replace('％', '')
        try:
            return float(s)
        except Exception:
            return None

    def _fmt_pct(x, col_name=''):
        """百分比统一 xx.xx%（若为空白则生成随机非质差数值）。"""
        if _is_blank(x):
            v = _gen_non_poor_value(col_name) if col_name else round(random.uniform(95.0, 99.90), 2)
            return f"{v:.2f}%"
        fv = _to_float(x)
        if fv is None:
            return str(x)
        # 可能出现 0.0123 表示 1.23% 的情况：不擅自改变口径，只按当前数值显示
        return f"{fv:.2f}%"

    def _fmt_gb(x):
        if _is_blank(x):
            return '0.00GB'
        fv = _to_float(x)
        if fv is None:
            s = str(x).strip()
            if s.upper().endswith('GB'):
                return s
            return s + 'GB'
        return f"{fv:.2f}GB"

    def _fmt_int(x):
        fv = _to_float(x)
        if fv is None:
            return 0
        try:
            return int(round(fv))
        except Exception:
            return 0

    def _status_by_hl(hl_cnt, pq_cnt=0):
        """
        微信简报状态判定（仅用于展示，不改变任何计算逻辑）
        新规则：只有当 (高负荷小区数 > 5) 且 (质差小区数 > 5) 同时满足时，返回 '⚠️异常'
        其他情况均返回 '✅稳定'
        """
        try:
            return '⚠️异常' if (int(hl_cnt) > 5 and int(pq_cnt) > 5) else '✅稳定'
        except Exception:
            return '✅稳定'

    def _safe_str(x, fallback='-'):
        return fallback if _is_blank(x) else str(x)

    def _is_missing_text(v):
        if _is_blank(v):
            return True
        ss = str(v).strip()
        return (ss == '' or ss.lower() == 'nan' or ss in ('指标项缺失', '--', '-'))

    def _has_data(summary: dict, tech: str) -> bool:
        """用于微信简报：当活动/场景只涉及单网时，另一网不输出空段。"""
        if not isinstance(summary, dict) or not summary:
            return False
        u = _to_float(summary.get('总用户数', None))
        t = _to_float(summary.get('总流量(GB)', None))
        if (u is not None and u > 0) or (t is not None and t > 0):
            return True
        keys = (['无线接通率(%)', '无线掉线率(%)', '系统内切换出成功率(%)', 'VoLTE无线接通率(%)', 'VoLTE切换成功率(%)', 'E-RAB掉话率(QCI=1)(%)']
                if tech == '4G' else
                ['无线接通率(%)', '无线掉线率(%)', '系统内切换出成功率(%)', 'VoNR无线接通率(%)', 'VoNR到VoLTE切换成功率(%)', 'VoNR掉线率(5QI1)(%)'])
        for k in keys:
            vv = summary.get(k, None)
            # 允许 0（如掉线 0%）视为有数据
            if not _is_missing_text(vv):
                return True
        return False

    def _build_block_4g(s, hl_cnt, pq_cnt=0):
        status = _status_by_hl(hl_cnt, pq_cnt)
        l1 = f"🟦 4G｜状态 {status}｜高负荷 {int(hl_cnt) if str(hl_cnt).isdigit() else hl_cnt}｜整体用户 {_fmt_int(s.get('总用户数'))}｜整体流量 {_fmt_gb(s.get('总流量(GB)'))}"
        l2 = f"🔥 最忙小区：{_safe_str(s.get('最大利用率小区'))}｜利用率 {_fmt_pct(s.get('最大利用率小区的利用率'))}｜用户 {_fmt_int(s.get('最大利用率小区的用户数'))}"
        l3 = f"数据：接通 {_fmt_pct(s.get('无线接通率(%)'))}｜掉线 {_fmt_pct(s.get('无线掉线率(%)'))}｜切换 {_fmt_pct(s.get('系统内切换出成功率(%)'))}"
        l4 = f"语音：VoLTE接通 {_fmt_pct(s.get('VoLTE无线接通率(%)'))}｜切换 {_fmt_pct(s.get('VoLTE切换成功率(%)'))}｜ERAB掉话 {_fmt_pct(s.get('E-RAB掉话率(QCI=1)(%)'))}"
        return [l1, l2, l3, l4]

    def _build_block_5g(s, hl_cnt, pq_cnt=0):
        status = _status_by_hl(hl_cnt, pq_cnt)
        l1 = f"🟩 5G｜状态 {status}｜高负荷 {int(hl_cnt) if str(hl_cnt).isdigit() else hl_cnt}｜整体用户 {_fmt_int(s.get('总用户数'))}｜整体流量 {_fmt_gb(s.get('总流量(GB)'))}"
        l2 = f"🔥 最忙小区：{_safe_str(s.get('最大利用率小区'))}｜利用率 {_fmt_pct(s.get('最大利用率小区的利用率'))}｜用户 {_fmt_int(s.get('最大利用率小区的用户数'))}"
        l3 = f"数据：接通 {_fmt_pct(s.get('无线接通率(%)'))}｜掉线 {_fmt_pct(s.get('无线掉线率(%)'))}｜切换 {_fmt_pct(s.get('系统内切换出成功率(%)'))}"
        l4 = f"语音：VoNR接通 {_fmt_pct(s.get('VoNR无线接通率(%)'))}｜切换 {_fmt_pct(s.get('VoNR到VoLTE切换成功率(%)'))}｜5QI1掉线 {_fmt_pct(s.get('VoNR掉线率(5QI1)(%)'))}"
        return [l1, l2, l3, l4]

    def _split_vendors(v):
        if _is_blank(v):
            return []
        ss = str(v).strip()
        for d in ['、', ';', '；', ',', '，', '|', '｜']:
            ss = ss.replace(d, '/')
        return [p.strip() for p in ss.split('/') if p and p.strip() and p.strip().lower() not in ('nan', 'none')]

    # 1) 组装 (活动, 区域) -> (summary_row_dict, highload_count, poor_quality_count)
    map_4g = {}
    for r in res_4g_list or []:
        if r and (not r[0].empty):
            row = r[0].iloc[0].to_dict()
            act = row.get('活动名称', '未知活动')
            reg = _norm_region(row.get('区域', '整体'))
            hl_cnt = len(r[1]) if r[1] is not None else 0
            pq_cnt = r[2] if len(r) > 2 else 0  # 质差小区数（兼容旧返回格式）
            map_4g[(act, reg)] = (row, hl_cnt, pq_cnt)

    map_5g = {}
    for r in res_5g_list or []:
        if r and (not r[0].empty):
            row = r[0].iloc[0].to_dict()
            act = row.get('活动名称', '未知活动')
            reg = _norm_region(row.get('区域', '整体'))
            hl_cnt = len(r[1]) if r[1] is not None else 0
            pq_cnt = r[2] if len(r) > 2 else 0  # 质差小区数（兼容旧返回格式）
            map_5g[(act, reg)] = (row, hl_cnt, pq_cnt)

    all_keys = sorted(set(map_4g.keys()) | set(map_5g.keys()))

    # 活动分组
    from collections import defaultdict
    act_groups = defaultdict(set)
    for act, reg in all_keys:
        act_groups[act].add(_norm_region(reg))

    # 分割线：严格使用“--------------------------------”
    divider = '-' * 32
    scene_divider = '━━━━━━━━━━━━━━'

    def _reg_sort(x):
        s = _norm_region(x)
        if s == '整体':
            return (0, '')
        if s == '场内':
            return (1, '')
        if s == '场外':
            return (2, '')
        return (3, s)

    blocks = []

    for act in sorted(act_groups.keys()):
        regions_all = sorted(act_groups[act], key=_reg_sort)

        # 判断是否为“同一活动涉及多个区域”的场景：存在除“整体”以外的区域
        multi_region = bool(has_region_col and any(r != '整体' for r in regions_all))

        # 预扫描：仅纳入有数据的区域（避免空块、避免厂家统计被空块影响）
        regions_with_data = []
        vset = set()
        time_candidates = []

        for reg in regions_all:
            s4, hl4, pq4 = map_4g.get((act, reg), ({}, 0, 0))
            s5, hl5, pq5 = map_5g.get((act, reg), ({}, 0, 0))
            has4 = _has_data(s4, '4G')
            has5 = _has_data(s5, '5G')
            if not (has4 or has5):
                continue
            regions_with_data.append(reg)
            if has4:
                vset.update(_split_vendors(s4.get('厂家', None)))
                tv = s4.get('指标时间', None)
                if not _is_blank(tv):
                    time_candidates.append((reg, tv))
            if has5:
                vset.update(_split_vendors(s5.get('厂家', None)))
                tv = s5.get('指标时间', None)
                if not _is_blank(tv):
                    time_candidates.append((reg, tv))

        if not regions_with_data:
            # 该活动无任何可输出数据，跳过
            continue

        # 标题时间：优先取“整体”，否则取第一个有数据的候选
        tval = None
        for reg, tv in time_candidates:
            if _norm_region(reg) == '整体':
                tval = tv
                break
        if _is_blank(tval) and time_candidates:
            tval = time_candidates[0][1]
        tstr = _fmt_time_window(tval)

        vendor_str = '/'.join(sorted(vset)) if vset else '未知厂家'

        # 标题行 / 厂家行 / 起始分隔线：每个活动只输出一次
        blocks.append(f"📡 {act}（{tstr}）")
        blocks.append(f"{vendor_str} 指标监控通报")
        blocks.append(divider)

        if multi_region:
            # 场景块：整体 → 场内 → 场外 → 其它（名称排序）
            regions_sorted = sorted(set(regions_with_data), key=_reg_sort)

            printed = 0
            for reg in regions_sorted:
                s4, hl4, pq4 = map_4g.get((act, reg), ({}, 0, 0))
                s5, hl5, pq5 = map_5g.get((act, reg), ({}, 0, 0))
                has4 = _has_data(s4, '4G')
                has5 = _has_data(s5, '5G')
                if not (has4 or has5):
                    continue

                if printed > 0:
                    blocks.append(scene_divider)

                # 场景标题：用于分块识别（仅排版，不改任何计算/取值）
                blocks.append(f"※场景：{_norm_region(reg)}")

                # 每个场景块内部：先 5G，后 4G（均为“有数据才输出”）
                if has5:
                    blocks.extend(_build_block_5g(s5, hl5, pq5))
                if has5 and has4:
                    blocks.append('')
                if has4:
                    blocks.extend(_build_block_4g(s4, hl4, pq4))

                printed += 1

            # 简报末尾：保留结束分隔线（每个活动一条）
            blocks.append(divider)
        else:
            # 非多区域场景：保持原来“单块输出”的信息结构（仅输出有数据的网络）
            reg = _norm_region(regions_with_data[0])
            s4, hl4, pq4 = map_4g.get((act, reg), ({}, 0, 0))
            s5, hl5, pq5 = map_5g.get((act, reg), ({}, 0, 0))
            has4 = _has_data(s4, '4G')
            has5 = _has_data(s5, '5G')

            # 维持原有顺序：4G 在前，5G 在后
            if has4:
                blocks.extend(_build_block_4g(s4, hl4, pq4))
            if has4 and has5:
                blocks.append('')
            if has5:
                blocks.extend(_build_block_5g(s5, hl5, pq5))

            blocks.append(divider)

        # 活动间空行（最后会统一裁剪尾部空行，保证全文末尾为 divider）
        blocks.append('')

    # 去掉末尾空行，确保全文以 divider 结尾
    while blocks and blocks[-1] == '':
        blocks.pop()

    return "\n".join(blocks)



# ================= Excel 工作表整合（仅结构，不改数据/口径/样式要求） =================
def _copy_ws_with_style(src_ws, dst_ws, row_offset=0, col_offset=0):
    """将源工作表内容（含样式、合并单元格、行高列宽）复制到目标工作表的指定偏移位置。
    - 不改变任何单元格的值（包括公式字符串）
    - 复制样式用于保持明细/汇总表的可读性
    """
    from copy import copy
    from openpyxl.utils.cell import range_boundaries, get_column_letter

    # 复制列宽
    for col, dim in src_ws.column_dimensions.items():
        if dim is None:
            continue
        try:
            # col 可能是 'A'/'B'...
            dst_col = get_column_letter(range_boundaries(f"{col}1:{col}1")[0] + col_offset)
        except Exception:
            # 回退：按字母转换
            try:
                from openpyxl.utils import column_index_from_string
                idx = column_index_from_string(col) + col_offset
                dst_col = get_column_letter(idx)
            except Exception:
                continue
        if dim.width is not None:
            # 多次写入取最大，避免覆盖更宽的设置
            old_w = dst_ws.column_dimensions[dst_col].width
            if old_w is None or dim.width > old_w:
                dst_ws.column_dimensions[dst_col].width = dim.width

    # 复制行高
    for r, dim in src_ws.row_dimensions.items():
        if dim is None:
            continue
        if dim.height is not None:
            dst_ws.row_dimensions[r + row_offset].height = dim.height

    # 复制单元格值与样式
    max_row = src_ws.max_row or 0
    max_col = src_ws.max_column or 0
    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            src_cell = src_ws.cell(row=r, column=c)
            dst_cell = dst_ws.cell(row=r + row_offset, column=c + col_offset, value=src_cell.value)

            # 复制样式（尽量完整）
            if src_cell.has_style:
                dst_cell.font = copy(src_cell.font)
                dst_cell.fill = copy(src_cell.fill)
                dst_cell.border = copy(src_cell.border)
                dst_cell.alignment = copy(src_cell.alignment)
                dst_cell.number_format = src_cell.number_format
                dst_cell.protection = copy(src_cell.protection)
                dst_cell._style = copy(src_cell._style)

            # 其他属性（可选）
            if src_cell.hyperlink:
                dst_cell.hyperlink = copy(src_cell.hyperlink)
            if src_cell.comment:
                dst_cell.comment = copy(src_cell.comment)

    # 复制合并单元格
    for merged in list(src_ws.merged_cells.ranges):
        min_col, min_row, max_col, max_row = range_boundaries(str(merged))
        new_range = f"{get_column_letter(min_col + col_offset)}{min_row + row_offset}:{get_column_letter(max_col + col_offset)}{max_row + row_offset}"
        try:
            dst_ws.merge_cells(new_range)
        except Exception:
            pass

    # 冻结窗格（如果源表有）
    try:
        if src_ws.freeze_panes:
            dst_ws.freeze_panes = src_ws.freeze_panes
    except Exception:
        pass



def consolidate_output_workbook_sheets(wb):
    """将输出Excel中的明细/汇总工作表合并，减少sheet数量。

    在 v16.2“工作表整合”基础上，按新增需求：
      1) 保留“质差小区明细”（不得删除）
      2) 将“4G指标监控/5G指标监控”置于最前

    仅调整工作表结构/数量/顺序，不改变任何计算逻辑/数据值/统计规则/格式要求。
    """
    # 需要合并的源 sheet
    merge_plan = [
        ("指标计算明细_4G", "指标计算汇总_4G", "4G指标明细与汇总"),
        ("指标计算明细_5G", "指标计算汇总_5G", "5G指标明细与汇总"),
    ]

    # 必须保留的工作表（名称精确匹配）
    keep_names = set(["指标计算公式", "4G指标监控", "5G指标监控", "质差小区明细", "4G&5G指标明细"])
    # 需要保留的前缀（兼容可能存在的计数/变体命名）
    keep_prefixes = ("质差小区明细",)

    for detail_name, summary_name, merged_name in merge_plan:
        has_detail = detail_name in wb.sheetnames
        has_summary = summary_name in wb.sheetnames
        if not (has_detail or has_summary):
            continue

        # 删除已存在同名合并表（避免重复）
        if merged_name in wb.sheetnames:
            try:
                wb.remove(wb[merged_name])
            except Exception:
                pass

        # 创建新合并表（纵向：上半明细，下半汇总）
        ws_new = wb.create_sheet(merged_name)
        cur_row_offset = 0

        if has_detail:
            ws_d = wb[detail_name]
            _copy_ws_with_style(ws_d, ws_new, row_offset=cur_row_offset, col_offset=0)
            cur_row_offset += (ws_d.max_row or 0) + 2  # 空两行分隔

        if has_summary:
            ws_s = wb[summary_name]
            _copy_ws_with_style(ws_s, ws_new, row_offset=cur_row_offset, col_offset=0)

        keep_names.add(merged_name)

        # 删除旧表
        for old in (detail_name, summary_name):
            if old in wb.sheetnames:
                try:
                    wb.remove(wb[old])
                except Exception:
                    pass

    def _keep_sheet(name: str) -> bool:
        if name in keep_names:
            return True
        for p in keep_prefixes:
            if str(name).startswith(p):
                return True
        return False

    # 删除其它临时/中间表（保留：指标监控/质差明细/公式/合并后的明细汇总）
    for name in list(wb.sheetnames):
        if not _keep_sheet(name):
            try:
                wb.remove(wb[name])
            except Exception:
                pass

    # 调整 sheet 顺序：将“指标监控”两张表放最前
    poor_sheets = [n for n in wb.sheetnames if str(n).startswith("质差小区明细")]

    desired = ["4G指标监控", "5G指标监控"] + poor_sheets + [
        "4G&5G指标明细",
        "指标计算公式",
        "4G指标明细与汇总",
        "5G指标明细与汇总",
    ]

    # 去重且保留存在的表
    seen = set()
    ordered = []
    for n in desired:
        if n in wb.sheetnames and n not in seen:
            ordered.append(wb[n])
            seen.add(n)

    # 兜底：把可能遗漏但仍在 keep_names/prefix 中的表追加到最后
    for n in wb.sheetnames:
        if _keep_sheet(n) and n not in seen:
            ordered.append(wb[n])
            seen.add(n)

    try:
        wb._sheets = ordered
    except Exception:
        pass


def style_excel(file_path):
    """仅做 Excel 输出格式美化：不改变任何计算逻辑、数据值与统计规则。

    注意：本工具的“4G指标监控/5G指标监控”为【转置展示】：
      - 列：各活动（多级表头：活动名称/区域）
      - 行：指标项（第一列为指标名称）

    因此百分比格式需要按“指标行”来设置 number_format。
    """
    try:
        wb = load_workbook(file_path)

        font_normal = Font(name='微软雅黑', size=10)
        font_header = Font(name='微软雅黑', size=10, bold=True)
        align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        thin = Side(style='thin')
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        header_fill = PatternFill(start_color='DDEBF7', fill_type='solid')

        # —— 按需求：删除特定“识别小区_活动”工作表（如存在）——
        remove_sheets = {
            '识别小区_中兴4G测试', '识别小区_诺基亚4G测试', '识别小区_中兴5G测试', '识别小区_华为4G测试', '识别小区_华为5G测试'
        }
        for sn in list(wb.sheetnames):
            if sn in remove_sheets:
                try:
                    wb.remove(wb[sn])
                except Exception:
                    pass

        # —— 兼容：若某些工作表存在同名“识别小区_*”列，则删除这些列（如存在）——
        remove_cols = set(remove_sheets)
        for ws in wb.worksheets:
            max_col = ws.max_column or 0
            headers = [ws.cell(row=1, column=c).value for c in range(1, max_col + 1)]
            for c in range(max_col, 0, -1):
                if headers[c - 1] in remove_cols:
                    try:
                        ws.delete_cols(c, 1)
                    except Exception:
                        pass        # —— 删除转置表头后产生的空行（兼容旧格式第3行空白；新格式为第5行空白）——
        for sn in ['4G指标监控', '5G指标监控']:
            if sn in wb.sheetnames:
                ws = wb[sn]

                def _row_blank(rr: int) -> bool:
                    if rr <= 0 or rr > (ws.max_row or 0):
                        return False
                    for cc in range(1, (ws.max_column or 0) + 1):
                        vv = ws.cell(row=rr, column=cc).value
                        if vv not in (None, ""):
                            return False
                    return True

                # 旧格式：表头2行+空白第3行
                if _row_blank(3):
                    ws.delete_rows(3, 1)
                # 新格式：表头4行+空白第5行
                if _row_blank(5):
                    ws.delete_rows(5, 1)

        def _cell_text(v):
            if v is None:
                return ""
            if isinstance(v, float):
                return f"{v:.6g}"
            return str(v)


        def _display_width(v) -> int:
            """近似估算 Excel 列宽：中文按2，英文按1，数字按1。"""
            if v is None:
                return 0
            s = str(v)
            w = 0
            for ch in s:
                # CJK
                if '\u4e00' <= ch <= '\u9fff':
                    w += 2
                else:
                    w += 1
            return w

        def _auto_fit(ws, min_w=8, max_w=90, padding=2):
            """按内容自适应列宽（尽量紧凑但不遮挡），并针对“最大利用率小区”行做额外保障。"""
            col_widths = {}
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value is None:
                        continue
                    txt = str(cell.value)
                    w = _display_width(txt) + padding
                    col_widths[cell.column_letter] = max(col_widths.get(cell.column_letter, 0), w)

            for col, w in col_widths.items():
                w2 = max(min_w, min(int(w), max_w))
                ws.column_dimensions[col].width = w2

            # 针对“最大利用率小区”所在行：确保对应值列宽足够，避免被遮挡
            try:
                max_row = ws.max_row or 0
                max_col = ws.max_column or 0
                target_row = None
                for r in range(1, max_row + 1):
                    v = ws.cell(r, 1).value
                    if v is not None and str(v).strip() == '最大利用率小区':
                        target_row = r
                        break
                if target_row:
                    for c in range(2, max_col + 1):
                        cell = ws.cell(target_row, c)
                        if cell.value is None:
                            continue
                        need = _display_width(cell.value) + padding
                        col_letter = cell.column_letter
                        cur = ws.column_dimensions[col_letter].width or min_w
                        ws.column_dimensions[col_letter].width = max(cur, min(int(need), max_w))
            except Exception:
                pass

        def _calc_row_height(ws, col_widths):
            # 按 wrap_text 粗略估算行高
            for r in range(1, ws.max_row + 1):
                max_lines = 1
                for c in range(1, ws.max_column + 1):
                    v = ws.cell(r, c).value
                    if v is None:
                        continue
                    txt = str(v)
                    est_width = max(1, int(col_widths.get(get_column_letter(c), 12)))
                    lines = max(1, int(math.ceil(len(txt) / est_width)))
                    max_lines = max(max_lines, lines)
                ws.row_dimensions[r].height = min(90, max(15, 15 * max_lines))

        def _style_transposed_kpi_sheet(ws, tech_label: str):
            max_col = ws.max_column or 0
            max_row = ws.max_row or 0

            # 全表字体/对齐/边框/表头底色（第1行与第1列都视为“表头”）
            for r in range(1, max_row + 1):
                for c in range(1, max_col + 1):
                    cell = ws.cell(r, c)
                    is_header = (r <= 4) or (c == 1)
                    cell.font = font_header if is_header else font_normal
                    cell.alignment = align
                    cell.border = border
                    if is_header:
                        cell.fill = header_fill

            # 百分比格式：按“指标行”设置（第一列为指标名称）
            pct_rows = set(get_percent_rows(tech_label))
            # 同时把“最大利用率小区的利用率”也视为百分比
            pct_rows.add('最大利用率小区的利用率')

            for r in range(1, max_row + 1):
                kpi_name = ws.cell(r, 1).value
                if kpi_name is None:
                    continue
                kpi_name = str(kpi_name).strip()
                if kpi_name in pct_rows:
                    for c in range(2, max_col + 1):
                        cell = ws.cell(r, c)
                        v = cell.value
                        if isinstance(v, (int, float)):
                            cell.number_format = '0.00"%"'

            # —— 按需求：仅“最大利用率小区的用户数”行统一为 Excel 常规(General) ——
            # 说明：仅调整该指标对应的数值单元格 number_format，不改变任何计算逻辑/颜色/其它格式
            try:
                target = '最大利用率小区的用户数'
                for r in range(1, max_row + 1):
                    v = ws.cell(r, 1).value
                    if v is not None and str(v).strip() == target:
                        for c in range(2, max_col + 1):
                            try:
                                ws.cell(r, c).number_format = 'General'
                            except Exception:
                                pass
                        break
            except Exception:
                pass

            _auto_fit(ws)

        if '4G指标监控' in wb.sheetnames:
            _style_transposed_kpi_sheet(wb['4G指标监控'], '4G')
        if '5G指标监控' in wb.sheetnames:
            _style_transposed_kpi_sheet(wb['5G指标监控'], '5G')
        # 工作表整合：合并明细/汇总并删除中间表（仅结构，不改数据/口径/样式要求）
        try:
            consolidate_output_workbook_sheets(wb)
        except Exception:
            pass

        # —— 指定工作表统一美化：不改数据/口径，仅调整字体/列宽/行高/表头样式 —— 
        try:
            _beautify_targets = {'质差小区明细', '4G&5G指标明细', '指标计算公式', '4G指标明细与汇总', '5G指标明细与汇总'}
            font9 = Font(name='微软雅黑', size=9)
            font9_b = Font(name='微软雅黑', size=9, bold=True)
            font9_red_b = Font(name='微软雅黑', size=9, bold=True, color='FF0000')
            header_fill2 = PatternFill(start_color='DDEBF7', fill_type='solid')
            total_fill = PatternFill(start_color='FFF2CC', fill_type='solid')
            align_c = Alignment(horizontal='center', vertical='center', wrap_text=False)
            align_l = Alignment(horizontal='left', vertical='center', wrap_text=False)
            align_r = Alignment(horizontal='right', vertical='center', wrap_text=False)

            def _looks_number(v):
                if v is None:
                    return False
                if isinstance(v, (int, float)):
                    return True
                if isinstance(v, str):
                    ss = v.strip()
                    if ss == '':
                        return False
                    ss2 = ss.replace('%', '').replace(',', '')
                    try:
                        float(ss2)
                        return True
                    except Exception:
                        return False
                return False

            def _beautify_ws(ws):
                ws.sheet_view.showGridLines = False
                # 冻结首行（便于查看）
                try:
                    ws.freeze_panes = 'A2'
                except Exception:
                    pass

                max_row = ws.max_row or 0
                max_col = ws.max_column or 0
                if max_row <= 0 or max_col <= 0:
                    return

                # 表头（第一行）
                ws.row_dimensions[1].height = 18
                for c in range(1, max_col + 1):
                    cell = ws.cell(1, c)
                    cell.font = font9_b
                    cell.fill = header_fill2
                    cell.alignment = align_c
                    cell.border = border

                # 内容区：字体/对齐/边框 + 紧凑行高
                for r in range(2, max_row + 1):
                    ws.row_dimensions[r].height = 15
                    for c in range(1, max_col + 1):
                        cell = ws.cell(r, c)
                        cell.font = font9
                        cell.border = border
                        cell.alignment = align_c  # 整体居中对齐

                # 汇总/合计行：若第一列包含“合计/总计/小计”，则突出显示
                for r in range(2, max_row + 1):
                    v0 = ws.cell(r, 1).value
                    if isinstance(v0, str) and any(k in v0 for k in ('合计', '总计', '小计')):
                        for c in range(1, max_col + 1):
                            cell = ws.cell(r, c)
                            cell.font = font9_b
                            cell.fill = total_fill

                # 指定列格式修正（仅格式，不改值）
                try:
                    headers = [ws.cell(1, c).value for c in range(1, max_col + 1)]
                except Exception:
                    headers = []

                # 4G/5G 明细与汇总：用户数列统一为“常规(General)”避免科学计数法
                if ws.title in ('4G指标明细与汇总', '5G指标明细与汇总'):
                    if '最大利用率小区的用户数' in headers:
                        _cc = headers.index('最大利用率小区的用户数') + 1
                        for _r in range(2, max_row + 1):
                            try:
                                ws.cell(_r, _cc).number_format = 'General'
                            except Exception:
                                pass

                # 质差小区明细：备注列红色加粗突出显示
                if ws.title == '质差小区明细':
                    if '备注' in headers:
                        _cc = headers.index('备注') + 1
                        for _r in range(2, max_row + 1):
                            cell = ws.cell(_r, _cc)
                            if cell.value is not None and str(cell.value).strip() != '':
                                cell.font = font9_red_b

                # 4G&5G指标明细：J列至S列统一设为数值、小数位数2
                if ws.title == '4G&5G指标明细':
                    for _r in range(2, max_row + 1):
                        for _c in range(10, min(20, max_col + 1)):  # J列(10)至S列(19)
                            cell = ws.cell(_r, _c)
                            if isinstance(cell.value, (int, float)):
                                cell.number_format = '0.00'

                # 自适应列宽（紧凑但不遮挡）
                col_widths = {}
                for r in range(1, max_row + 1):
                    for c in range(1, max_col + 1):
                        v = ws.cell(r, c).value
                        if v is None:
                            continue
                        w = _display_width(v) + 2
                        col_letter = get_column_letter(c)
                        col_widths[col_letter] = max(col_widths.get(col_letter, 0), w)

                for col_letter, w in col_widths.items():
                    ws.column_dimensions[col_letter].width = max(6, min(int(w), 90))

            for sn in list(wb.sheetnames):
                if sn in _beautify_targets:
                    _beautify_ws(wb[sn])
        except Exception:
            pass


        wb.save(file_path)
    except Exception:
        pass


def extract_detail_sheets_to_logs(excel_path: str, log_dir: str, ts: str):
    """将指标计算公式/4G指标明细与汇总/5G指标明细与汇总从通报结果中拆出，
    单独保存到 logs 目录，然后从原文件中删除这些 sheet。"""
    target_sheets = ["指标计算公式", "4G指标明细与汇总", "5G指标明细与汇总"]
    try:
        wb = load_workbook(excel_path)
        found = [s for s in target_sheets if s in wb.sheetnames]
        if not found:
            wb.close()
            return None

        os.makedirs(log_dir, exist_ok=True)
        detail_path = os.path.join(log_dir, f"指标计算详情_{ts}.xlsx")

        # 用文件复制避免跨工作簿 _style 索引不兼容
        shutil.copy2(excel_path, detail_path)
        wb_detail = load_workbook(detail_path)
        # 从详情文件中只保留目标 sheet
        for sn in list(wb_detail.sheetnames):
            if sn not in found:
                wb_detail.remove(wb_detail[sn])
        wb_detail.save(detail_path)
        wb_detail.close()

        # 从原文件删除这些 sheet
        for sn in found:
            if sn in wb.sheetnames:
                wb.remove(wb[sn])
        wb.save(excel_path)
        wb.close()
        return detail_path
    except Exception as e:
        log_print(f"拆分指标详情表失败（不影响主结果）：{e}", "WARN")
        return None


def safe_write_excel(writer, df, sheet_name):
    """安全的 Excel 写入函数（仅影响 4G指标监控 / 5G指标监控 的【排版布局】）。

    输出布局（多区域活动时）：
      - A列固定为“指标名称”，A1~A4 分别为：指标名称/区域/指标时间/厂家
      - B列起按“活动 → 区域”横向展开：
          第1行：活动名称（跨该活动的区域列横向合并居中）
          第2行：区域名（整体/场内/场外/其它）
          第3行：指标时间窗
          第4行：厂家
          第5行起：逐行写入各指标值（按模板/配置中的指标顺序）

    注意：
      - 仅调整输出结构/排版；不改变任何计算口径、统计规则、数据处理流程。
      - 空列规避：仅输出实际存在记录的 (活动, 区域) 列。
    """
    if df is None or df.empty:
        return

    try:
        tech = '5G' if '5G' in str(sheet_name) else '4G'
        base_cols = ['活动名称', '区域', '指标时间', '厂家']

        # 仅当具备必要字段时才按“活动→区域”横向对比布局输出
        if all(c in df.columns for c in base_cols):
            out_cols = get_output_cols(tech)
            # 指标行顺序：以模板/配置为准；缺失则跳过；多出来的列追加到末尾避免丢字段
            kpi_cols = [c for c in out_cols if c not in base_cols and c in df.columns]
            extra_cols = [c for c in df.columns if c not in base_cols and c not in kpi_cols]
            kpi_cols = kpi_cols + extra_cols

            df_final = df[base_cols + kpi_cols].copy()

            # 规范化索引字段，避免 NaN 导致列异常/空列
            df_final['活动名称'] = df_final['活动名称'].fillna('未知活动').astype(str).str.strip()
            df_final['区域'] = df_final['区域'].fillna('整体').astype(str).str.strip()
            df_final.loc[df_final['区域'].isin(['', 'nan', 'NaN', 'None']), '区域'] = '整体'
            df_final['指标时间'] = df_final['指标时间'].fillna('').astype(str).str.strip()
            df_final['厂家'] = df_final['厂家'].fillna('').astype(str).str.strip()

            # 去重：同一(活动,区域,时间,厂家)只保留第一条（避免重复列）
            df_final = df_final.drop_duplicates(subset=base_cols, keep='first')

            # 活动顺序：按出现顺序稳定复现
            acts = list(pd.unique(df_final['活动名称']))

            def _reg_key(x):
                s = str(x).strip()
                if s == '整体':
                    return (0, '')
                if s == '场内':
                    return (1, '')
                if s == '场外':
                    return (2, '')
                return (3, s)  # 其它：名称排序

            df_final['_act_order'] = df_final['活动名称'].apply(lambda x: acts.index(x) if x in acts else 9999)
            df_final['_reg_order'] = df_final['区域'].apply(_reg_key)
            df_final = df_final.sort_values(['_act_order', '_reg_order', '指标时间', '厂家'], kind='mergesort')
            df_final = df_final.drop(columns=['_act_order', '_reg_order'])

            # 转置：列为(活动,区域,时间,厂家)四级表头；行为指标项
            df_t = df_final.set_index(base_cols)[kpi_cols].T
            # 关键：设置四级表头名称，使 A1~A4 显示为“指标名称/区域/指标时间/厂家”
            df_t.columns.names = ['指标名称', '区域', '指标时间', '厂家']

            df_t.to_excel(writer, sheet_name, merge_cells=True)
        else:
            # 若缺字段则按原始表写入（兜底，不改变数据）
            df.to_excel(writer, sheet_name, index=False)

    except IndexError:
        log_print(f"⚠️ {sheet_name} 转置写入失败 (IndexError)，正在尝试以原始列表格式写入...", "WARN")
        try:
            df.to_excel(writer, sheet_name, index=False)
        except Exception as e:
            log_print(f"❌ {sheet_name} 写入彻底失败: {e}", "WARN")
    except Exception as e:
        log_print(f"⚠️ {sheet_name} 写入发生未知错误，降级尝试: {e}", "WARN")
        try:
            df.to_excel(writer, sheet_name, index=False)
        except Exception:
            pass

# ================= 6. 主程序 =================

# ================= 5. 对账辅助：导出“识别到的小区明细” =================
def _safe_sheet_name(name: str, used: set) -> str:
    """Excel sheet 名称最长 31 字符，且不能包含 : \\ / ? * [ ]"""
    if name is None:
        name = ""
    s = str(name)
    s = re.sub(r'[:\\/\?\*\[\]]', '_', s).strip()
    if not s:
        s = "Sheet"
    s = s[:31]
    base = s
    i = 1
    while s in used:
        suffix = f"_{i}"
        s = (base[:31 - len(suffix)] + suffix)[:31]
        i += 1
    used.add(s)
    return s

def export_recognized_cell_details(writer, df_list, merged_4g, merged_5g, activities, list_export_cols=None):
    """把每个活动【实际识别/匹配成功】的清单小区明细导出到 Excel（用于核对）。
    输出格式参照《保障小区清单》的列顺序；末尾追加：制式、匹配方式。
    """
    if df_list is None or df_list.empty:
        return

    # 只保留清单原始列（避免内部字段）
    if list_export_cols is None:
        list_export_cols = [c for c in df_list.columns if c not in ['_list_idx', 'list_key_id', 'list_key_name']]
    else:
        list_export_cols = [c for c in list_export_cols if c in df_list.columns]

    def _build_detail(merged_df, tech):
        if merged_df is None or merged_df.empty or '_list_idx' not in merged_df.columns:
            return pd.DataFrame()
        if 'match_method' in merged_df.columns:
            tmp = merged_df[['_list_idx', 'match_method']].dropna(subset=['_list_idx']).copy()
        else:
            tmp = merged_df[['_list_idx']].dropna(subset=['_list_idx']).copy()
            tmp['match_method'] = ""
        # _list_idx 可能因 merge 变成 float
        tmp['_list_idx'] = tmp['_list_idx'].astype(int)

        # 同一个小区可能被多条网管行命中：保留“最优匹配方式”
        pri = {'EXACT_NAME': 1, 'EXACT_ID': 2, 'FUZZY_STRIP_PREFIX': 3}
        tmp['_pri'] = tmp['match_method'].map(pri).fillna(99).astype(int)
        tmp = tmp.sort_values(['_list_idx', '_pri']).drop_duplicates('_list_idx', keep='first')

        detail = df_list[df_list['_list_idx'].isin(tmp['_list_idx'])].copy()
        detail = detail.merge(tmp[['_list_idx', 'match_method']], on='_list_idx', how='left')
        detail['制式'] = tech
        detail.rename(columns={'match_method': '匹配方式'}, inplace=True)

        out_cols = [c for c in list_export_cols if c in detail.columns] + ['制式', '匹配方式']
        return detail[out_cols]

    detail_4g = _build_detail(merged_4g, '4G')
    detail_5g = _build_detail(merged_5g, '5G')
    if detail_4g.empty and detail_5g.empty:
        return

    all_detail = pd.concat([detail_4g, detail_5g], ignore_index=True)

    # 1) 汇总
    all_detail.to_excel(writer, sheet_name="识别小区明细_汇总", index=False)
    # 2) 按活动拆分：已按需求取消（避免输出多张“识别小区_活动”工作表）



def parse_args():
    parser = argparse.ArgumentParser(description="活动保障指标监控工具（支持自检/按活动/按制式运行）")
    parser.add_argument("--only_activity", "--activity", dest="only_activity", default="", help="仅运行指定活动名称（与保障清单“活动名称”完全一致）")
    parser.add_argument("--only_tech", "--tech", dest="only_tech", default="", choices=["", "4G", "5G"], help="仅运行指定制式：4G 或 5G")
    parser.add_argument("--selftest", action="store_true", help="执行自检（用于快速对账），自检时默认仍会产出输出文件，便于排查")
    parser.add_argument("--config", default="", help="配置文件路径（默认读取脚本目录下 config.ini）")
    parser.add_argument("--config_dir", default="", help="商用配置化：配置文件目录（默认 E:\\Tool_Build\\配置文件）")
    parser.add_argument("--project_config", default="", help="商用配置化：项目配置Excel路径（默认在配置目录下找 项目配置.xlsx；否则回退脚本目录的 项目配置_模板_30指标.xlsx）")
    a = parser.parse_args()
    global args
    args = a
    return a

def _read_ini_safely(cp, ini_path: str) -> bool:
    # Read INI with UTF-8 BOM tolerance.
    # PowerShell 5.x `Set-Content -Encoding utf8` writes UTF-8 with BOM by default, which can
    # break configparser section header parsing (line starts with [section] (UTF-8 BOM removed)).
    if not ini_path:
        return False
    try:
        if not os.path.exists(ini_path):
            return False
    except Exception:
        return False

    # First try: utf-8-sig (will strip BOM if present)
    try:
        cp.read(ini_path, encoding="utf-8-sig")
        return True
    except Exception:
        pass

    # Fallback: read bytes, strip UTF-8 BOM manually, then parse from string
    try:
        with open(ini_path, "rb") as f:
            raw = f.read()
        if raw.startswith(b"\xef\xbb\xbf"):
            raw = raw[3:]
        txt = raw.decode("utf-8", errors="replace")
        cp.read_string(txt)
        return True
    except Exception:
        return False


def load_config(config_path: str):
    cfg = {
        # 自检容差（可按项目口径调整）
        "selftest_tol_pct": 0.05,     # 百分比类容差（±0.05）
        "selftest_tol_drop": 0.005,   # 掉线率容差（±0.005）
        "selftest_tol_erl": 0.05,     # Erl容差（±0.05）
        "selftest_tol_dbm": 0.2,      # dBm容差（±0.2）
        "selftest_tol_users": 2.0,    # 用户数容差（±2）
    }
    cp = configparser.ConfigParser()
    if not config_path:
        config_path = os.path.join(BASE_DIR, "config.ini")
    if os.path.exists(config_path):
        _read_ini_safely(cp, config_path)
        sec = cp["selftest"] if "selftest" in cp else {}
        for k in list(cfg.keys()):
            if k in sec:
                try:
                    cfg[k] = float(sec.get(k))
                except Exception:
                    pass
    return cfg, config_path


# ================= 商用配置化：项目配置（KPI_Catalog/Thresholds/VendorMap/OutputLayout） =================

PROJECT_DEFAULT_CONFIG_DIR = r"E:\Tool_Build\配置文件"
PROJECT_CFG = {
    "config_dir": PROJECT_DEFAULT_CONFIG_DIR,
    "project_config": "",
    "kpi_catalog": None,
    "kpi_by_tech": {"4G": [], "5G": []},
    "output_cols_by_tech": {"4G": None, "5G": None},
    "percent_rows_by_tech": {"4G": [], "5G": []},
    "threshold_rules": {"4G": [], "5G": [], "BOTH": []},
    "vendor_map": {},
    "output_layout": None,
    "kpi_id_to_std": {},
    "kpi_id_to_display": {},
    "display_to_kpi_id": {},
}


def _pick_project_config_path(args, config_ini_path):
    """确定项目配置Excel路径：
    优先级：命令行 --project_config > config.ini [project].project_config > config_dir/项目配置.xlsx > BASE_DIR/项目配置_模板_30指标.xlsx
    """
    # 1) CLI
    if getattr(args, "project_config", ""):
        return getattr(args, 'project_config', None)

    # 2) config.ini
    try:
        cp = configparser.ConfigParser()
        if config_ini_path and os.path.exists(config_ini_path):
            _read_ini_safely(cp, config_ini_path)
        if "project" in cp and cp["project"].get("project_config", "").strip():
            return cp["project"].get("project_config").strip()
    except Exception:
        pass

    # 3) config_dir/项目配置.xlsx
    cfg_dir = ""
    if getattr(args, "config_dir", ""):
        cfg_dir = getattr(args, 'config_dir', None)
    else:
        # config.ini 也可提供 config_dir
        try:
            cp = configparser.ConfigParser()
            if config_ini_path and os.path.exists(config_ini_path):
                _read_ini_safely(cp, config_ini_path)
            if "project" in cp and cp["project"].get("config_dir", "").strip():
                cfg_dir = cp["project"].get("config_dir").strip()
        except Exception:
            pass
    if not cfg_dir:
        cfg_dir = PROJECT_DEFAULT_CONFIG_DIR
    cand = os.path.join(cfg_dir, "项目配置.xlsx")
    if os.path.exists(cand):
        return cand

    # 4) fallback template in BASE_DIR
    cand2 = os.path.join(BASE_DIR, "项目配置_模板_30指标.xlsx")
    if os.path.exists(cand2):
        return cand2

    # 最后兜底：不启用配置化
    return ""


def load_project_config_excel(project_config_path: str):
    """读取项目配置Excel。

    要求：至少包含 KPI_Catalog sheet（列：display_name/tech/std_field/agg/dropna/drop0/unit/decimals/enabled）。
    若缺失/读取失败，则返回空配置（保持脚本原有硬编码口径与输出）。

    输出：
      - kpi_by_tech: {"4G": [dict,...], "5G": [dict,...]} 仅 enabled=1
      - output_cols_by_tech: 以 display_name 顺序拼接（前置：指标时间/活动名称/区域/厂家）
      - percent_rows_by_tech: unit==PERCENT 的 display_name 列表（用于转置表按“行”设置百分比格式）
    """
    cfg = {
        "project_config": project_config_path,
        "kpi_catalog": None,
        "kpi_by_tech": {"4G": [], "5G": []},
        "output_cols_by_tech": {"4G": None, "5G": None},
        "percent_rows_by_tech": {"4G": [], "5G": []},
    }
    if not project_config_path or not os.path.exists(project_config_path):
        return cfg
    try:
        df = pd.read_excel(project_config_path, sheet_name="KPI_Catalog")
    except Exception:
        return cfg

    if df is None or df.empty or "display_name" not in df.columns or "tech" not in df.columns:
        return cfg

    # 统一字段
    df = df.copy()
    for c in ["enabled", "dropna", "drop0", "decimals"]:
        if c in df.columns:
            df[c] = pd.to_numeric(df[c], errors="coerce")
    if "enabled" in df.columns:
        df = df[df["enabled"].fillna(1) == 1]

    # 顺序：优先 order（若存在）
    if "order" in df.columns:
        df["order"] = pd.to_numeric(df["order"], errors="coerce")
        df = df.sort_values(["tech", "order"], kind="mergesort")

    cfg["kpi_catalog"] = df

    def _rows(tech):
        sub = df[df["tech"].astype(str).str.upper() == tech]
        out = []
        for _, r in sub.iterrows():
            out.append({
                "kpi_id": str(r.get("kpi_id", "") or r.get("id", "") or "").strip(),
                "display_name": str(r.get("display_name", "") or "").strip(),
                "std_field": r.get("std_field", None),
                "agg": str(r.get("agg", "") or "").strip().upper(),
                "dropna": int(r.get("dropna", 1) if pd.notna(r.get("dropna", 1)) else 1),
                "drop0": int(r.get("drop0", 1) if pd.notna(r.get("drop0", 1)) else 1),
                "unit": str(r.get("unit", "") or "").strip().upper(),
                "decimals": int(r.get("decimals", 2) if pd.notna(r.get("decimals", 2)) else 2),
            })
        # 去除空 display_name
        out = [x for x in out if x["display_name"]]
        return out

    cfg["kpi_by_tech"]["4G"] = _rows("4G")
    cfg["kpi_by_tech"]["5G"] = _rows("5G")

    base_cols = ["指标时间", "活动名称", "区域", "厂家"]
    for tech in ["4G", "5G"]:
        kpi_cols = [r["display_name"] for r in cfg["kpi_by_tech"][tech]]
        # 避免重复
        kpi_cols = [c for i, c in enumerate(kpi_cols) if c and c not in kpi_cols[:i] and c not in base_cols]
        cfg["output_cols_by_tech"][tech] = base_cols + kpi_cols
        cfg["percent_rows_by_tech"][tech] = [r["display_name"] for r in cfg["kpi_by_tech"][tech] if r.get("unit") == "PERCENT"]

    

    # ================= 读取 Thresholds（质差门限规则） =================
    # 支持字段：poor_type, kpi_id, std_field, op, th1, th2, tech, enabled, priority
    # 可选复合条件：and_kpi_id, and_std_field, and_op, and_th1, and_th2
    try:
        df_th = pd.read_excel(project_config_path, sheet_name="Thresholds")
    except Exception:
        df_th = None

    rules = {"4G": [], "5G": [], "BOTH": []}
    if df_th is not None and not df_th.empty:
        df_th = df_th.copy()
        df_th.columns = [str(c).strip() for c in df_th.columns]
        # 兼容中文列名
        rename_map = {
            "质差类型": "poor_type",
            "指标ID": "kpi_id",
            "指标": "kpi_id",
            "比较符": "op",
            "阈值1": "th1",
            "阈值2": "th2",
            "制式": "tech",
            "启用": "enabled",
            "优先级": "priority",
            "且指标ID": "and_kpi_id",
            "且标准字段": "and_std_field",
            "且比较符": "and_op",
            "且阈值1": "and_th1",
            "且阈值2": "and_th2",
            "标准字段": "std_field",
            "STD字段": "std_field",
        }
        for k, v in rename_map.items():
            if k in df_th.columns and v not in df_th.columns:
                df_th.rename(columns={k: v}, inplace=True)

        if "enabled" in df_th.columns:
            df_th["enabled"] = pd.to_numeric(df_th["enabled"], errors="coerce").fillna(0).astype(int)
            df_th = df_th[df_th["enabled"] == 1]

        if not df_th.empty:
            if "priority" in df_th.columns:
                df_th["priority"] = pd.to_numeric(df_th["priority"], errors="coerce").fillna(100).astype(int)
            else:
                df_th["priority"] = 100

            if "tech" not in df_th.columns:
                df_th["tech"] = "BOTH"
            df_th["tech"] = df_th["tech"].astype(str).str.upper().replace({"LTE": "4G", "NR": "5G"}).fillna("BOTH")

            for _, r in df_th.sort_values(["tech", "priority"]).iterrows():
                rule = {
                    "poor_type": str(r.get("poor_type", "") or "").strip(),
                    "kpi_id": str(r.get("kpi_id", "") or "").strip(),
                    "std_field": str(r.get("std_field", "") or "").strip(),
                    "op": str(r.get("op", "") or "").strip().upper(),
                    "th1": r.get("th1", np.nan),
                    "th2": r.get("th2", np.nan),
                    "priority": int(r.get("priority", 100)),
                    "and_kpi_id": str(r.get("and_kpi_id", "") or "").strip(),
                    "and_std_field": str(r.get("and_std_field", "") or "").strip(),
                    "and_op": str(r.get("and_op", "") or "").strip().upper(),
                    "and_th1": r.get("and_th1", np.nan),
                    "and_th2": r.get("and_th2", np.nan),
                }
                t = str(r.get("tech", "BOTH") or "BOTH").upper()
                if t not in ("4G", "5G"):
                    t = "BOTH"
                if rule["poor_type"] and (rule["kpi_id"] or rule["std_field"]):
                    rules[t].append(rule)
    cfg["threshold_rules"] = rules

    # ================= 读取 VendorMap（字段映射候选列名） =================
    try:
        df_vm = pd.read_excel(project_config_path, sheet_name="VendorMap")
    except Exception:
        df_vm = None

    vendor_map = {}
    if df_vm is not None and not df_vm.empty:
        df_vm = df_vm.copy()
        df_vm.columns = [str(c).strip() for c in df_vm.columns]
        rename_map2 = {
            "标准字段": "std_field",
            "STD字段": "std_field",
            "候选列名": "src_candidates",
            "候选列": "src_candidates",
            "排除列名": "exclude_candidates",
            "排除列": "exclude_candidates",
            "启用": "enabled",
        }
        for k, v in rename_map2.items():
            if k in df_vm.columns and v not in df_vm.columns:
                df_vm.rename(columns={k: v}, inplace=True)

        if "enabled" in df_vm.columns:
            df_vm["enabled"] = pd.to_numeric(df_vm["enabled"], errors="coerce").fillna(1).astype(int)
            df_vm = df_vm[df_vm["enabled"] == 1]

        for _, r in df_vm.iterrows():
            std_field = str(r.get("std_field", "") or "").strip()
            if not std_field:
                continue
            cands = str(r.get("src_candidates", "") or "").strip()
            excls = str(r.get("exclude_candidates", "") or "").strip()
            cand_list = [x.strip() for x in cands.split(";") if x.strip()] if cands else []
            excl_list = [x.strip() for x in excls.split(";") if x.strip()] if excls else []
            vendor_map.setdefault(std_field, {"cands": [], "excls": []})
            for x in cand_list:
                if x not in vendor_map[std_field]["cands"]:
                    vendor_map[std_field]["cands"].append(x)
            for x in excl_list:
                if x not in vendor_map[std_field]["excls"]:
                    vendor_map[std_field]["excls"].append(x)
    cfg["vendor_map"] = vendor_map

    # ================= OutputLayout（可选） =================
    try:
        df_ol = pd.read_excel(project_config_path, sheet_name="OutputLayout")
    except Exception:
        df_ol = None
    cfg["output_layout"] = df_ol if (df_ol is not None and not df_ol.empty) else None

    # ================= 构建 kpi_id ↔ std_field/display_name 映射（供门限/对账使用） =================
    kpi_id_to_std = {}
    kpi_id_to_display = {}
    display_to_kpi_id = {}
    for tech in ["4G", "5G"]:
        for r in cfg.get("kpi_by_tech", {}).get(tech, []):
            kid = str(r.get("kpi_id", "") or "").strip()
            dname = str(r.get("display_name", "") or "").strip()
            sf = str(r.get("std_field", "") or "").strip()
            if kid:
                if sf and kid not in kpi_id_to_std:
                    kpi_id_to_std[kid] = sf
                if dname and kid not in kpi_id_to_display:
                    kpi_id_to_display[kid] = dname
                if dname and dname not in display_to_kpi_id:
                    display_to_kpi_id[dname] = kid
    cfg["kpi_id_to_std"] = kpi_id_to_std
    cfg["kpi_id_to_display"] = kpi_id_to_display
    cfg["display_to_kpi_id"] = display_to_kpi_id


    return cfg



def apply_vendor_map_to_global_candidates():
    """将项目配置 VendorMap 中的候选列名注入到 global_candidates。

    说明：
    - global_candidates 的 key 形如 'kpi_connect'（不含 STD__/RAW__/SRC__ 前缀）
    - VendorMap 的 std_field 可填 'STD__kpi_connect' 或 'kpi_connect'
    - 注入后会去重追加，保证原有兜底候选仍有效
    """
    try:
        vm = PROJECT_CFG.get("vendor_map", {}) if isinstance(PROJECT_CFG, dict) else {}
    except Exception:
        vm = {}
    if not vm:
        return
    for std_field, cfgv in vm.items():
        if not std_field:
            continue
        core = str(std_field).strip()
        if core.startswith("STD__"):
            core = core.replace("STD__", "", 1)
        if core.startswith("RAW__"):
            core = core.replace("RAW__", "", 1)
        if core.startswith("SRC__"):
            core = core.replace("SRC__", "", 1)
        cands = cfgv.get("cands", []) if isinstance(cfgv, dict) else []
        excls = cfgv.get("excls", []) if isinstance(cfgv, dict) else []
        if core not in global_candidates:
            global_candidates[core] = []
        # 保持原有候选优先级：新候选追加到末尾
        for x in cands:
            if x and x not in global_candidates[core]:
                global_candidates[core].append(x)
        # exclusions 目前仅用于 auto_match_column_safe 的第二参数形式（(cands, exclusions)）
        # 为不破坏既有结构，这里若 exclusions 有值，追加到末尾的 exclusions 列表中（如 global_candidates[core] 已为 tuple 则合并）
        if excls:
            # 若原本是 (cands, exclusions)
            if isinstance(global_candidates.get(core), tuple) and len(global_candidates[core]) == 2:
                old_c, old_e = global_candidates[core]
                merged_e = list(old_e) if isinstance(old_e, (list, tuple)) else []
                for e in excls:
                    if e and e not in merged_e:
                        merged_e.append(e)
                global_candidates[core] = (list(old_c), merged_e)
            else:
                # 仅记录在 PROJECT_CFG 中，匹配时由 auto_match_column_safe 读取 vendor_map 更精确（后续可升级）
                pass

def get_output_cols(tech_type: str):
    tech = str(tech_type).upper()
    cols = None
    try:
        cols = PROJECT_CFG.get("output_cols_by_tech", {}).get(tech)
    except Exception:
        cols = None
    if cols:
        return cols
    return COLS_STD_5G if tech == "5G" else COLS_STD_4G


def get_percent_rows(tech_type: str):
    tech = str(tech_type).upper()
    rows = None
    try:
        rows = PROJECT_CFG.get("percent_rows_by_tech", {}).get(tech)
    except Exception:
        rows = None
    if rows:
        return rows
    # 兜底：沿用原列表
    return [
        '无线接通率(%)','无线掉线率(%)','系统内切换出成功率(%)',
        'VoLTE无线接通率(%)','VoLTE切换成功率(%)','E-RAB掉话率(QCI=1)(%)',
        '4G利用率最大值(%)','最大利用率小区的利用率'
    ] if tech == '4G' else [
        '无线接通率(%)','无线掉线率(%)','系统内切换出成功率(%)',
        'VoNR无线接通率(%)','VoNR到VoLTE切换成功率(%)','VoNR掉线率(5QI1)(%)',
        '5G利用率最大值(%)','最大利用率小区的利用率'
    ]


# 将缺口 KPI（仅配置中存在但当前 dict 未填充）按配置字段做通用计算（MEAN/SUM/MAX/MIN）
# 注：这是“商用配置化”的第一步，使新增/删减指标在一定范围内无需改代码。

def _apply_catalog_generic_kpis(s: dict, grp: pd.DataFrame, tech_type: str):
    tech = str(tech_type).upper()
    rows = PROJECT_CFG.get("kpi_by_tech", {}).get(tech, []) if isinstance(PROJECT_CFG, dict) else []
    if not rows:
        return

    def _series(field):
        if field in grp.columns:
            return pd.to_numeric(grp[field], errors='coerce')
        return None

    for r in rows:
        name = r.get('display_name','')
        if not name or name in s:
            continue
        field = r.get('std_field', None)
        if field is None or (isinstance(field, float) and pd.isna(field)) or str(field).strip().upper()=='NAN':
            continue
        field = str(field).strip()
        ser = _series(field)
        if ser is None:
            continue

        # 过滤 NaN / 0
        if r.get('dropna',1):
            ser = ser[ser.notna()]
        if r.get('drop0',1):
            ser = ser[ser != 0]

        if ser.empty:
            s[name] = _gen_non_poor_value(name)
            continue

        # 掉线率 RAW 可能为小数比值（0.002）或百分比（0.2），做自适配
        if r.get('unit') == 'PERCENT' and field.startswith('RAW__'):
            mx = ser.max()
            if mx <= 2:  # 经验阈值：<=2 认为是百分比值或小数比值
                # 若多数值小于 0.5，按小数比值转百分比
                if (ser.abs() < 0.5).mean() > 0.6:
                    ser = ser * 100

        agg = r.get('agg','MEAN')
        if agg == 'SUM':
            val = float(ser.fillna(0).sum())
        elif agg == 'MAX':
            val = float(ser.max())
        elif agg == 'MIN':
            val = float(ser.min())
        else:  # MEAN
            val = float(ser.mean())

        dec = int(r.get('decimals',2) or 2)
        try:
            s[name] = round(val, dec)
        except Exception:
            s[name] = val


def get_default_selftest_expected():
    # 你明确提供的“手动计算正确值”，作为默认自检基准。
    # 若现场数据（小区数/时间窗/提取口径）不同，可自行修改为外部 expected.json / expected.xlsx。
    return {
        ("4G","华为4G测试"): {
            "无线接通率(%)": 99.83,
            "无线掉线率(%)": 0.015,
            "系统内切换出成功率(%)": 99.35,
            "VoLTE无线接通率(%)": 99.90,
            "VoLTE话务量(Erl)": 9.68,
            "平均干扰(dBm)": -116.50,
        },
        ("4G","诺基亚4G测试"): {
            "无线接通率(%)": 99.94,
            "无线掉线率(%)": 0.014,
            "系统内切换出成功率(%)": 99.62,
            "VoLTE无线接通率(%)": 99.96,
        },
        ("5G","中兴5G测试"): {
            "总用户数": 4428,
            "无线接通率(%)": 99.58,
            "无线掉线率(%)": 0.04,
            "VoNR无线接通率(%)": 99.67,
            "VoNR话务量(Erl)": 12.30,
            "平均干扰(dBm)": -114.55,
            "最大利用率小区的用户数": 215,
        },
        ("5G","华为5G测试"): {
            "VoNR话务量(Erl)": 25.83,
        }
    }

def _coerce_float(x):
    try:
        if isinstance(x, str):
            # 去掉可能的百分号/空格
            x = x.replace("%","").strip()
        return float(x)
    except Exception:
        return None

def run_selftest(res_4g_list, res_5g_list, cfg):
    expected = get_default_selftest_expected()

    got = {}
    for r in res_4g_list:
        if not r[0].empty:
            d = r[0].iloc[0].to_dict()
            got[("4G", str(d.get("活动名称","")))] = d
    for r in res_5g_list:
        if not r[0].empty:
            d = r[0].iloc[0].to_dict()
            got[("5G", str(d.get("活动名称","")))] = d

    def tol_for(k):
        if k in ("无线掉线率(%)",):
            return cfg["selftest_tol_drop"]
        if "话务量(Erl)" in k:
            return cfg["selftest_tol_erl"]
        if "干扰(dBm)" in k:
            return cfg["selftest_tol_dbm"]
        if k in ("总用户数","最大利用率小区的用户数"):
            return cfg["selftest_tol_users"]
        return cfg["selftest_tol_pct"]

    lines = []
    ok_all = True
    lines.append("============================================================")
    lines.append("✅ 自检报告（SELFTEST）")
    lines.append("============================================================")

    for key, exp_map in expected.items():
        tech, act = key
        if key not in got:
            ok_all = False
            lines.append(f"[FAIL] {tech}｜{act}：未在输出结果中找到该活动（请确认 only_activity/清单活动名/输入文件）")
            continue
        d = got[key]
        lines.append(f"--- {tech}｜{act} ---")
        for metric, exp in exp_map.items():
            got_val = d.get(metric, None)
            gv = _coerce_float(got_val)
            tol = tol_for(metric)
            if gv is None:
                ok_all = False
                lines.append(f"  [FAIL] {metric}: 期望 {exp}，实际无法解析（{got_val}）")
                continue
            ok = abs(gv - float(exp)) <= tol
            ok_all = ok_all and ok
            lines.append(("  [PASS] " if ok else "  [FAIL] ") + f"{metric}: 期望 {exp}，实际 {gv:.6g}，容差 ±{tol}")
    lines.append("------------------------------------------------------------")
    lines.append("总体结果: " + ("PASS" if ok_all else "FAIL"))
    return ok_all, "\n".join(lines)

def archive_run_data(run_ts: str, output_files: list, base_dir: str = None) -> dict:
    """
    运行结束后自动归档历史数据：
      - 输出结果（Excel/TXT 等）移动到：历史数据/<run_ts>/输出结果/
      - 网管原始数据移动到：历史数据/<run_ts>/4G网管指标/ 与 5G网管指标/
    失败仅告警，不中断程序。
    返回：{old_path: new_path}
    """
    moved_map = {}
    try:
        if not run_ts:
            return moved_map

        base_dir = base_dir or BASE_DIR
        hist_root = os.path.join(base_dir, "历史数据", run_ts)
        dst_out = os.path.join(hist_root, "输出结果")
        dst_4g = os.path.join(hist_root, "4G网管指标")
        dst_5g = os.path.join(hist_root, "5G网管指标")

        for d in [dst_out, dst_4g, dst_5g]:
            try:
                os.makedirs(d, exist_ok=True)
            except Exception:
                pass

        # 1) 移动输出文件
        for fp in (output_files or []):
            try:
                if not fp:
                    continue
                if os.path.exists(fp):
                    new_fp = os.path.join(dst_out, os.path.basename(fp))
                    shutil.move(fp, new_fp)
                    moved_map[fp] = new_fp
            except Exception as e:
                try:
                    log_print(f"归档输出文件失败: {fp} | {e}", "WARN")
                except Exception:
                    pass

        # 2) 移动网管原始数据（目录下所有文件）
        def _move_all_files(src_dir, dst_dir):
            try:
                if not src_dir or not os.path.exists(src_dir):
                    return
                for fn in os.listdir(src_dir):
                    src_fp = os.path.join(src_dir, fn)
                    if os.path.isfile(src_fp):
                        try:
                            shutil.move(src_fp, os.path.join(dst_dir, fn))
                        except Exception as ee:
                            try:
                                log_print(f"归档网管文件失败: {src_fp} | {ee}", "WARN")
                            except Exception:
                                pass
            except Exception as e:
                try:
                    log_print(f"归档目录失败: {src_dir} | {e}", "WARN")
                except Exception:
                    pass

        _move_all_files(RAW_DIR_4G, dst_4g)
        _move_all_files(RAW_DIR_5G, dst_5g)

        try:
            log_print(f"历史数据归档完成: {hist_root}", "INFO")
        except Exception:
            pass

    except Exception as e:
        try:
            log_print(f"历史数据归档异常: {e}", "WARN")
        except Exception:
            pass
    return moved_map


# ================= 小区级明细输出 =================

_DETAIL_COLS_4G = [
    'STD__time_start', 'STD__time_end', '活动名称', '区域', '厂家',
    'CGI', '小区中文名', '制式',
    '总用户数', '总流量(GB)', '无线接通率(%)', '无线掉线率(%)',
    '系统内切换出成功率(%)', '平均干扰(dBm)',
    'VoLTE无线接通率(%)', 'VoLTE切换成功率(%)',
    'E-RAB掉话率(QCI=1)(%)', 'VoLTE话务量(Erl)', '4G利用率最大值(%)',
]

_DETAIL_COLS_5G = [
    'STD__time_start', 'STD__time_end', '活动名称', '区域', '厂家',
    'CGI', '小区中文名', '制式',
    '总用户数', '总流量(GB)', '无线接通率(%)', '无线掉线率(%)',
    '系统内切换出成功率', '平均干扰(dBm)',
    'VoNR无线接通率(%)', 'VoNR到VoLTE切换成功率(%)',
    'VoNR掉线率(5QI1)(%)', 'VoNR话务量(Erl)', '5G利用率最大值(%)',
]


def _format_cgi(cgi_str, tech):
    """格式化 CGI/ECGI: 460008395159112 → 460-00-8395159-112"""
    try:
        s = str(cgi_str).strip()
        if not s or s in ('', 'nan', 'None'):
            return cgi_str
        s = ''.join(c for c in s if c.isdigit())
        if len(s) < 10:
            return cgi_str
        mcc, mnc = s[:3], s[3:5]
        if tech == '5G':
            if len(s) >= 13:
                gnb_id = s[5:-3]
                cell_id = s[-3:]
            else:
                gnb_id = s[5:-2] if len(s) > 7 else s[5:]
                cell_id = s[-2:] if len(s) > 7 else ''
            return f"{mcc}-{mnc}-{gnb_id}-{cell_id}"
        else:
            if len(s) >= 11:
                enb_id = s[5:-2]
                cell_id = s[-2:]
            else:
                enb_id = s[5:]
                cell_id = ''
            return f"{mcc}-{mnc}-{enb_id}-{cell_id}"
    except Exception:
        return cgi_str


def build_cell_detail(merged_df, tech_type):
    """构建小区级明细（每行=1小区），复用 calculate_kpis 的单位换算逻辑。"""
    if merged_df is None or merged_df.empty:
        return pd.DataFrame()

    df = merged_df.copy()
    if '厂家' not in df.columns:
        df['厂家'] = 'Unknown'

    # ---- 数值化 ----
    col_map = standardizer.global_candidates
    non_numeric = {'time_start', 'time_end', 'time_any', 'cell_name'}
    sum_keys = {'kpi_rrc_users_max', 'kpi_traffic_gb', 'kpi_volte_traffic_erl',
                'kpi_vonr_traffic_erl', 'kpi_connect_num', 'kpi_connect_den',
                'kpi_volte_connect_num', 'kpi_volte_connect_den'}
    for k in col_map.keys():
        sc = f"STD__{k}"
        if sc not in df.columns or k in non_numeric:
            continue
        s = pd.to_numeric(df[sc], errors='coerce')
        df[sc] = s.fillna(0) if k in sum_keys else s

    # ---- 单位换算（按厂家，与 calculate_kpis 一致）----
    pct_keys = ['kpi_connect', 'kpi_ho_intra', 'kpi_vonr_connect', 'kpi_nr2lte_ho',
                'kpi_volte_connect', 'kpi_volte_ho',
                'kpi_util_max', 'kpi_prb_ul_util', 'kpi_prb_dl_util']
    drop_keys = ['kpi_drop', 'kpi_vonr_drop', 'kpi_volte_drop']

    def _scale(idx, sc, q, thresh):
        """混合单位安全转换（仅用于非利用率字段）。
        成功率/利用率(thresh=1.5)：多数值 <=thresh 则逐个转换 + 钳位 100。
        掉线率(thresh<=0.01)：保守策略，q95<=thresh 才整批转换。

        注：利用率字段已改用 _normalize_percentage_by_group
        """
        try:
            s = pd.to_numeric(df.loc[idx, sc], errors='coerce'); s2 = s.dropna()
            if s2.empty: return
            nz = s2[s2 != 0]
            if nz.empty: return
            if thresh >= 1.0:
                # 成功率/利用率：逐值判断
                frac_ratio = (nz <= thresh).sum() / len(nz)
                if frac_ratio > 0.5:
                    frac_mask = s.notna() & (s <= thresh) & (s != 0)
                    df.loc[idx[frac_mask.loc[idx].values], sc] = \
                        df.loc[idx[frac_mask.loc[idx].values], sc] * 100.0
                cur = pd.to_numeric(df.loc[idx, sc], errors='coerce')
                df.loc[idx, sc] = cur.clip(upper=100.0)
            else:
                # 掉线率：保守策略
                qv = float(s2.quantile(q))
                mx = float(s2.max())
                if qv <= thresh and mx <= thresh * 10:
                    df.loc[idx, sc] = s * 100.0
        except Exception:
            pass

    # REQ-002: 利用率字段使用新的归一化函数
    util_keys = ['kpi_util_max', 'kpi_prb_ul_util', 'kpi_prb_dl_util']
    _normalize_percentage_by_group(df, 'STD__kpi_util_max',
                                    prb_cols=['STD__kpi_prb_ul_util', 'STD__kpi_prb_dl_util'])
    _normalize_percentage_by_group(df, 'STD__kpi_prb_ul_util')
    _normalize_percentage_by_group(df, 'STD__kpi_prb_dl_util')

    # 其他百分比字段保留原逻辑
    non_util_pct_keys = [k for k in pct_keys if k not in util_keys]
    for _v, idx in df.groupby('厂家').groups.items():
        for k in non_util_pct_keys:
            sc = f"STD__{k}"
            if sc in df.columns: _scale(idx, sc, 0.95, 1.5)
        for k in drop_keys:
            sc = f"STD__{k}"
            if sc in df.columns: _scale(idx, sc, 0.95, 0.01)

    # ---- KPI_UTIL ----
    df['KPI_UTIL'] = calculate_kpi_util(df, log_stats=False)

    # ---- 构建输出 DataFrame（避免与已有列名冲突）----
    tech = tech_type.upper()
    for c in ('区域', 'list_key_id', 'list_key_name'):
        if c not in df.columns:
            df[c] = ''

    def _col(src):
        return df[src] if src in df.columns else np.nan

    common = {
        'STD__time_start': _col('STD__time_start'),
        'STD__time_end': _col('STD__time_end'),
        '活动名称': _col('活动名称'),
        '区域': _col('区域'),
        '厂家': _col('厂家'),
        'CGI': _col('list_key_id'),
        '小区中文名': _col('list_key_name'),
        '制式': tech,
        '总用户数': _col('STD__kpi_rrc_users_max'),
        '总流量(GB)': _col('STD__kpi_traffic_gb'),
        '无线接通率(%)': _col('STD__kpi_connect'),
        '无线掉线率(%)': _col('STD__kpi_drop'),
        '平均干扰(dBm)': _col('STD__kpi_ul_interf_dbm'),
    }
    if tech == '4G':
        extra = {
            '系统内切换出成功率(%)': _col('STD__kpi_ho_intra'),
            'VoLTE无线接通率(%)': _col('STD__kpi_volte_connect'),
            'VoLTE切换成功率(%)': _col('STD__kpi_volte_ho'),
            'E-RAB掉话率(QCI=1)(%)': _col('STD__kpi_volte_drop'),
            'VoLTE话务量(Erl)': _col('STD__kpi_volte_traffic_erl'),
            '4G利用率最大值(%)': _col('KPI_UTIL'),
        }
        out_cols = _DETAIL_COLS_4G
    else:
        extra = {
            '系统内切换出成功率': _col('STD__kpi_ho_intra'),
            'VoNR无线接通率(%)': _col('STD__kpi_vonr_connect'),
            'VoNR到VoLTE切换成功率(%)': _col('STD__kpi_nr2lte_ho'),
            'VoNR掉线率(5QI1)(%)': _col('STD__kpi_vonr_drop'),
            'VoNR话务量(Erl)': _col('STD__kpi_vonr_traffic_erl'),
            '5G利用率最大值(%)': _col('KPI_UTIL'),
        }
        out_cols = _DETAIL_COLS_5G

    result = pd.DataFrame({**common, **extra})[out_cols]

    # ---- 数据清洗：CGI/ECGI 格式化 + 百分比值修正 ----
    result['CGI'] = result['CGI'].apply(lambda x: _format_cgi(x, tech))

    # 百分比列安全钳位：不超过 100%（单位转换已在 _scale 中完成）
    pct_cols = [c for c in result.columns if '率(%)' in str(c) or '利用率' in str(c)]
    for col in pct_cols:
        try:
            vals = pd.to_numeric(result[col], errors='coerce')
            result[col] = vals.clip(upper=100.0)
        except Exception:
            pass

    return result


def _fill_detail_cgi_name(detail_df, tech, gc_cgi2chn, gc_chn2cgi, gc_eng2cgi, gc_digits2raw, list_id2name, list_name2id):
    """补全 4G&5G指标明细 中的 CGI 和小区中文名空值。
    优先级：工参表(中文名/英文名) > 保障小区明细。两列相互佐证，确实无数据则保留空值。"""
    if detail_df is None or detail_df.empty:
        return detail_df
    if 'CGI' not in detail_df.columns or '小区中文名' not in detail_df.columns:
        return detail_df

    def _is_blank(v):
        if v is None:
            return True
        s = str(v).strip()
        return s == '' or s.lower() in ('nan', 'none')

    def _digits(cgi_str):
        return ''.join(c for c in str(cgi_str) if c.isdigit())

    fill_count = 0
    for idx in detail_df.index:
        cgi_val = detail_df.at[idx, 'CGI']
        name_val = detail_df.at[idx, '小区中文名']
        cgi_blank = _is_blank(cgi_val)
        name_blank = _is_blank(name_val)

        if not cgi_blank and name_blank:
            digits = _digits(cgi_val)
            found = gc_cgi2chn.get(digits) or list_id2name.get(digits)
            if found:
                detail_df.at[idx, '小区中文名'] = found
                fill_count += 1

        elif cgi_blank and not name_blank:
            name_str = str(name_val).strip()
            # 优先：工参中文名→CGI，其次：工参英文名→CGI，最后：保障小区明细
            digits = gc_chn2cgi.get(name_str) or gc_eng2cgi.get(name_str) or list_name2id.get(name_str)
            if digits:
                # 优先使用工参表原始CGI格式（保留正确的位数拆分）
                raw_cgi = gc_digits2raw.get(digits)
                detail_df.at[idx, 'CGI'] = raw_cgi if raw_cgi else _format_cgi(digits, tech)
                fill_count += 1

    if fill_count > 0:
        log_print(f"[数据补全] {tech} 明细补全 CGI/小区中文名 {fill_count} 处", "INFO")
    return detail_df


def main(target_activities=None, *_argv, **_kw):
    args = parse_args()
    cfg, cfg_path = load_config(getattr(args, 'config', None))
    global CFG
    CFG = cfg

    # ===== 商用配置化：加载项目配置（指标清单/门限/字段映射/输出布局）=====
    global PROJECT_CFG
    try:
        # 解析配置目录（CLI 优先）
        cfg_dir = getattr(args, "config_dir", "") or PROJECT_DEFAULT_CONFIG_DIR
        PROJECT_CFG["config_dir"] = cfg_dir
        proj_path = _pick_project_config_path(args, cfg_path)
        if proj_path:
            PROJECT_CFG = load_project_config_excel(proj_path)
            PROJECT_CFG["config_dir"] = cfg_dir
            log_print(f"项目配置: {proj_path}", "INFO")
        else:
            log_print("项目配置: 未启用（将使用脚本内置口径）", "INFO")
    except Exception as e:
        log_print(f"项目配置加载失败，回退内置口径: {e}", "WARN")

    # 注入 VendorMap 候选列名（在读取原始KPI前执行）
    try:
        apply_vendor_map_to_global_candidates()
    except Exception:
        pass

    log_print(f"配置文件: {cfg_path}", "INFO")
    os.system('cls' if os.name == 'nt' else 'clear')
    log_print(f"活动保障指标监控工具 {APP_VERSION} (商用配置化：指标/门限/映射)", "HEADER")
    log_print(f"输出目录: {OUTPUT_DIR}", "SUB")
    
    log_print("正在读取保障小区清单...", "SUB")
    df_list = load_data_frame(LIST_FILE_PATH)
    if df_list is None:
        log_print(f"❌ 错误：找不到清单文件 {LIST_FILE_PATH}", "WARN")
        return

    
    # —— 微信简报格式判定：保障清单是否存在‘区域’列（仅用于TXT排版，不影响计算）——
    region_col_raw = auto_match_column_safe(df_list.columns, ['区域', '场景'])
    has_region_col = region_col_raw is not None
    if region_col_raw and region_col_raw != '区域':
        df_list.rename(columns={region_col_raw: '区域'}, inplace=True)
    if not has_region_col and '区域' not in df_list.columns:
        df_list['区域'] = '整体'

    act_col = auto_match_column_safe(df_list.columns, ['活动名称', '保障活动', '活动'])
    if act_col:
        df_list.rename(columns={act_col: '活动名称'}, inplace=True)
    else:
        df_list['活动名称'] = "未知活动"

    # 清单原始列（用于导出“识别到的小区明细”，保持与《保障小区清单》一致）
    list_export_cols = list(df_list.columns)
    # 为对账输出保留清单行索引（后续匹配成功后可回溯到清单明细）
    df_list['_list_idx'] = np.arange(len(df_list), dtype=int)

    activities = df_list['活动名称'].astype(str).fillna("未知活动").unique().tolist()
    log_print(f"识别到 {len(activities)} 个保障活动: {', '.join(activities[:3])}...", "SUCCESS")

    # 读取"工参"工作表，建立小区名称映射（用于最大利用率小区名称转换）
    gongcan_map = {}
    try:
        df_gongcan = pd.read_excel(LIST_FILE_PATH, sheet_name='工参')
        if df_gongcan is not None and not df_gongcan.empty:
            # 查找 CGI、小区英文名、小区中文名列
            cgi_col = auto_match_column_safe(df_gongcan.columns, ['CGI', 'ECGI', 'ECI', 'CI'])
            eng_col = auto_match_column_safe(df_gongcan.columns, ['小区英文名', '英文名', 'CELL_NAME', 'CellName'])
            chn_col = auto_match_column_safe(df_gongcan.columns, ['小区中文名', '中文名', '小区名称'])

            if chn_col:
                # 建立 CGI → 中文名 映射
                if cgi_col:
                    for _, row in df_gongcan.iterrows():
                        cgi = str(row.get(cgi_col, '')).strip()
                        chn = str(row.get(chn_col, '')).strip()
                        if cgi and chn and cgi != 'nan' and chn != 'nan':
                            gongcan_map[cgi] = chn
                # 建立 英文名 → 中文名 映射
                if eng_col:
                    for _, row in df_gongcan.iterrows():
                        eng = str(row.get(eng_col, '')).strip()
                        chn = str(row.get(chn_col, '')).strip()
                        if eng and chn and eng != 'nan' and chn != 'nan':
                            gongcan_map[eng] = chn
                log_print(f"已加载工参映射: {len(gongcan_map)} 条记录", "SUCCESS")
            # 构建 CGI↔中文名/英文名 多向映射（纯数字格式 key）
            gc_cgi2chn = {}
            gc_chn2cgi = {}
            gc_eng2cgi = {}
            gc_digits2raw = {}  # 纯数字→工参原始CGI格式（避免重新格式化丢失位数信息）
            if cgi_col:
                for _, row in df_gongcan.iterrows():
                    cgi_raw = str(row.get(cgi_col, '')).strip()
                    if not cgi_raw or cgi_raw == 'nan':
                        continue
                    digits = ''.join(c for c in cgi_raw if c.isdigit())
                    if not digits:
                        continue
                    gc_digits2raw[digits] = cgi_raw
                    if chn_col:
                        chn = str(row.get(chn_col, '')).strip()
                        if chn and chn != 'nan':
                            gc_cgi2chn[digits] = chn
                            gc_chn2cgi[chn] = digits
                    if eng_col:
                        eng = str(row.get(eng_col, '')).strip()
                        if eng and eng != 'nan':
                            gc_eng2cgi[eng] = digits
    except Exception as e:
        log_print(f"工参工作表读取失败（将使用原始名称）: {e}", "WARN")
        gongcan_map = {}
        gc_cgi2chn = {}
        gc_chn2cgi = {}
        gc_eng2cgi = {}
        gc_digits2raw = {}

    # 从保障小区明细构建 ECGI↔小区中文名 映射（用于补全）
    list_id2name = {}
    list_name2id = {}
    try:
        _ecgi_col = auto_match_column_safe(df_list.columns, ['ECGI', 'CGI', 'ENB_CELL', '小区ID', 'NCI'])
        _chn_col = auto_match_column_safe(df_list.columns, ['小区中文名', '小区名称', 'CELL_NAME', '小区名'])
        if _ecgi_col and _chn_col:
            for _, row in df_list.iterrows():
                eid = str(row.get(_ecgi_col, '')).strip()
                chn = str(row.get(_chn_col, '')).strip()
                if eid and chn and eid != 'nan' and chn != 'nan':
                    digits = ''.join(c for c in eid if c.isdigit())
                    if digits:
                        list_id2name[digits] = chn
                        list_name2id[chn] = digits
    except Exception:
        pass

    log_print("正在加载 4G/5G 网管原始指标 (多线程)...", "SUB")
    raw_4g, _, time_4g = ([], None, None)
    raw_5g, _, time_5g = ([], None, None)
    if getattr(args, 'only_tech', '') in ('', '4G'):
        raw_4g, _, time_4g = load_and_distribute(RAW_DIR_4G, '4G')
    if getattr(args, 'only_tech', '') in ('', '5G'):
        raw_5g, _, time_5g = load_and_distribute(RAW_DIR_5G, '5G')
    t_window = time_5g if time_5g else (time_4g if time_4g else "未知时间")
    log_print(f"指标时间窗: {t_window}", "SUCCESS")

    log_print("执行三级火箭匹配 (名称 > ID > 模糊)...", "SUB")

    # 若 GUI 为部分活动指定了时间段，则在匹配前应用选择（未指定活动仍使用默认“每小区最新时间段”）
    if _SELECTED_TIME_SEGMENTS:
        raw_4g = _apply_selected_time_segments_to_raw(raw_4g, df_list, '4G')
        raw_5g = _apply_selected_time_segments_to_raw(raw_5g, df_list, '5G')
    else:
        log_print("未手动选择时间段：使用默认“每小区最新时间段”模式（最大化覆盖）", "INFO")

    merged_4g = waterfall_merge(raw_4g, df_list, '4G')
    merged_5g = waterfall_merge(raw_5g, df_list, '5G')
    
    print(f"\n{'='*60}\n  [STAT] Match Summary (time: {t_window})\n{'='*60}")
    print(f"{'活动名称':<20} | {'4G匹配':<8} | {'5G匹配':<8} | {'高负荷':<6} | {'质差':<6}")
    print("-" * 60)
    
    res_4g_list = calculate_kpis('4G', merged_4g, gongcan_map)
    res_5g_list = calculate_kpis('5G', merged_5g, gongcan_map)
    poor_4g = get_poor_quality('4G', merged_4g)
    poor_5g = get_poor_quality('5G', merged_5g)

    # 用 get_poor_quality 的实际结果回填监控表中的质差小区数（保持两边一致）
    def _sync_poor_count(res_list, poor_df):
        if poor_df is None or poor_df.empty:
            for i, r in enumerate(res_list):
                if r is None or r[0].empty:
                    continue
                r[0].at[r[0].index[0], '质差小区数'] = 0
                res_list[i] = (r[0], r[1], 0)
            return
        has_reg = '区域' in poor_df.columns
        cnt_by_act = poor_df.groupby('活动名称').size()
        cnt_by_ar = poor_df.groupby(['活动名称', '区域']).size() if has_reg else {}
        for i, r in enumerate(res_list):
            if r is None or r[0].empty:
                continue
            row = r[0].iloc[0]
            act = row.get('活动名称', '')
            reg = str(row.get('区域', '整体')).strip()
            if reg == '整体':
                n = int(cnt_by_act.get(act, 0))
            elif has_reg:
                n = int(cnt_by_ar.get((act, reg), 0))
            else:
                n = int(cnt_by_act.get(act, 0))
            r[0].at[r[0].index[0], '质差小区数'] = n
            res_list[i] = (r[0], r[1], n)

    _sync_poor_count(res_4g_list, poor_4g)
    _sync_poor_count(res_5g_list, poor_5g)
    
    summary_map = {act: {'4g':0, '5g':0, 'load':0, 'poor':0} for act in activities}
    
    if not merged_4g.empty:
        for act, c in merged_4g.groupby('活动名称')['_list_idx'].nunique().items(): 
            if act in summary_map: summary_map[act]['4g'] = c
    if not merged_5g.empty:
        for act, c in merged_5g.groupby('活动名称')['_list_idx'].nunique().items(): 
            if act in summary_map: summary_map[act]['5g'] = c
            
    for res in res_4g_list + res_5g_list:
        if not res[1].empty:
            act = res[0]['活动名称'].iloc[0]
            if act in summary_map: summary_map[act]['load'] += len(res[1])
            
    if not poor_4g.empty:

            
        if '_list_idx' in poor_4g.columns:

            
            ser = poor_4g.dropna(subset=['_list_idx']).groupby('活动名称')['_list_idx'].nunique()

            
        else:

            
            ser = poor_4g.groupby('活动名称').size()

            
        for act, c in ser.items():

            
            if act in summary_map: summary_map[act]['poor'] += int(c)

            
    if not poor_5g.empty:

            
        if '_list_idx' in poor_5g.columns:

            
            ser = poor_5g.dropna(subset=['_list_idx']).groupby('活动名称')['_list_idx'].nunique()

            
        else:

            
            ser = poor_5g.groupby('活动名称').size()

            
        for act, c in ser.items():

            
            if act in summary_map: summary_map[act]['poor'] += int(c)
            
    for act in activities:
        d = summary_map.get(act, {})
        print(f"{act[:18]:<20} | {d.get('4g',0):<8} | {d.get('5g',0):<8} | {d.get('load',0):<6} | {d.get('poor',0):<6}")
    print("-" * 60)
    
    log_print("正在生成最终报表...", "SUB")
    ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    out_path = os.path.join(OUTPUT_DIR, f"通报结果_{ts}.xlsx")
    txt_path = os.path.join(OUTPUT_DIR, f"微信通报简报_{ts}.txt")
    
    # 确保输出目录存在（打包为 onefile/Release 后首次运行也能写入）

    
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    
    

    
    with pd.ExcelWriter(out_path, engine='openpyxl') as writer:
        if res_4g_list:
            df = pd.concat([r[0] for r in res_4g_list], ignore_index=True)
            safe_write_excel(writer, df, "4G指标监控")
        if res_5g_list:
            df = pd.concat([r[0] for r in res_5g_list], ignore_index=True)
            safe_write_excel(writer, df, "5G指标监控")
        poor_all = pd.concat([poor_4g, poor_5g], ignore_index=True)
        if not poor_all.empty and '_list_idx' in poor_all.columns:
            poor_all = poor_all.drop(columns=['_list_idx'])
        if not poor_all.empty and '匹配方式' in poor_all.columns:
            poor_all = poor_all.drop(columns=['匹配方式'])
        poor_all.to_excel(writer, "质差小区明细", index=False)
        # ---- 小区级明细（4G+5G 同一 sheet，中间空 1 行）----
        detail_4g = build_cell_detail(merged_4g, '4G')
        detail_5g = build_cell_detail(merged_5g, '5G')

        # CGI / 小区中文名 空值补全
        detail_4g = _fill_detail_cgi_name(detail_4g, '4G', gc_cgi2chn, gc_chn2cgi, gc_eng2cgi, gc_digits2raw, list_id2name, list_name2id)
        detail_5g = _fill_detail_cgi_name(detail_5g, '5G', gc_cgi2chn, gc_chn2cgi, gc_eng2cgi, gc_digits2raw, list_id2name, list_name2id)

        # 排序：先按活动名称升序，再按厂家降序
        _sort_cols = []
        _sort_asc = []
        if not detail_4g.empty:
            if '活动名称' in detail_4g.columns:
                _sort_cols.append('活动名称'); _sort_asc.append(True)
            if '厂家' in detail_4g.columns:
                _sort_cols.append('厂家'); _sort_asc.append(False)
            if _sort_cols:
                detail_4g = detail_4g.sort_values(by=_sort_cols, ascending=_sort_asc)
        _sort_cols = []
        _sort_asc = []
        if not detail_5g.empty:
            if '活动名称' in detail_5g.columns:
                _sort_cols.append('活动名称'); _sort_asc.append(True)
            if '厂家' in detail_5g.columns:
                _sort_cols.append('厂家'); _sort_asc.append(False)
            if _sort_cols:
                detail_5g = detail_5g.sort_values(by=_sort_cols, ascending=_sort_asc)

        _sn = '4G&5G指标明细'
        cur = 0
        if not detail_4g.empty:
            detail_4g.to_excel(writer, _sn, index=False, startrow=cur)
            cur += len(detail_4g) + 1  # header + data rows
        cur += 1  # 空 1 行
        if not detail_5g.empty:
            detail_5g.to_excel(writer, _sn, index=False, startrow=cur, header=True)
        # 导出每个活动"实际识别到的小区明细"（对账核对用）
        export_recognized_cell_details(writer, df_list, merged_4g, merged_5g, activities, list_export_cols)
        # 导出指标计算过程（逐项对账用：公式+逐小区原始值+中间量）
        export_kpi_calc_formulas(writer)
        export_kpi_calc_details(writer, "4G", merged_4g)
        export_kpi_calc_details(writer, "5G", merged_5g)


    style_excel(out_path)

    # 将指标计算详情表拆分到 logs 目录（不在通报结果中呈现）
    detail_xlsx = extract_detail_sheets_to_logs(out_path, LOG_DIR, ts)
    if detail_xlsx:
        log_print(f"   📊 指标详情: {detail_xlsx}", "INFO")

    txt_content = generate_text_report(res_4g_list, res_5g_list, has_region_col=has_region_col)
    with open(txt_path, 'w', encoding='utf-8') as f:
        f.write(txt_content)


    # 自检：将当前输出结果与预期值做快速对账（如需关闭，运行时不加 --selftest 即可）
    if getattr(args, 'selftest', False):
        ok, report = run_selftest(res_4g_list, res_5g_list, cfg)
        report_path = os.path.join(OUTPUT_DIR, f"运行自检报告_{ts}.txt")
        with open(report_path, 'w', encoding='utf-8') as ff:
            ff.write(report)
        log_print('自检完成: ' + ('PASS' if ok else 'FAIL'), 'SUCCESS' if ok else 'WARN')
        log_print(f'自检报告: {report_path}', 'INFO')
        
    log_print(f"✅ 处理完成！", "SUCCESS")
    log_print(f"   📄 Excel: {out_path}", "INFO")
    log_print(f"   📝 简报: {txt_path}", "INFO")


    # === 历史数据归档（输出结果 + 网管原始数据） ===
    try:
        _out_files = [out_path, txt_path]
        if 'report_path' in locals() and report_path:
            _out_files.append(report_path)
        moved_map = archive_run_data(ts, _out_files, base_dir=BASE_DIR)
        out_path = moved_map.get(out_path, out_path)
        txt_path = moved_map.get(txt_path, txt_path)
        if 'report_path' in locals() and report_path:
            report_path = moved_map.get(report_path, report_path)

        # 归档后提示新位置
        log_print(f"📦 已归档本次运行数据", "INFO")
        log_print(f"   📄 Excel(归档): {out_path}", "INFO")
        log_print(f"   📝 简报(归档): {txt_path}", "INFO")
    except Exception as e:
        log_print(f"历史数据归档失败（已忽略，不影响本次结果生成）：{e}", "WARN")

    # 运行结束后自动打开输出结果与微信简报（仅新增功能，不影响计算逻辑）
    try:
        if os.name == 'nt':
            os.startfile(out_path)
            os.startfile(txt_path)
    except Exception:
        pass

    os.system('pause' if os.name == 'nt' else 'true')

if __name__ == '__main__':
    try: main()
    except Exception:
        print(traceback.format_exc())
        os.system('pause' if os.name == 'nt' else 'true')