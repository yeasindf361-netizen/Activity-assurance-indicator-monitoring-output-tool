# -*- coding: utf-8 -*-
from __future__ import annotations

import os
from typing import Dict, Optional, Tuple

import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from kpi_tool.config.logging_config import log_print
from kpi_tool.config.project_config import get_output_cols, get_percent_rows

def safe_write_excel(writer, df, sheet_name):
    """安全的 Excel 写入函数（仅影响 4G指标监控 / 5G指标监控 的【排版布局】）。

    输出布局（多区域活动时）：
      - A列固定为“指标名称”，A1~A4 分别为：指标名称/区域/指标时间/厂家
      - B列起按“活动 → 区域”横向展开：
          第1行：活动名称（跨该活动的区域列横向合并居中）
          第2行：区域名（整体/场内/场外/其它）
          第3行：指标时间窗
          第4行：厂家
          第5行起：逐行写入各指标值（按模板/配置中的指标顺序）

    注意：
      - 仅调整输出结构/排版；不改变任何计算口径、统计规则、数据处理流程。
      - 空列规避：仅输出实际存在记录的 (活动, 区域) 列。
    """
    if df is None or df.empty:
        return

    try:
        tech = '5G' if '5G' in str(sheet_name) else '4G'
        base_cols = ['活动名称', '区域', '指标时间', '厂家']

        # 仅当具备必要字段时才按“活动→区域”横向对比布局输出
        if all(c in df.columns for c in base_cols):
            out_cols = get_output_cols(tech)
            # 指标行顺序：以模板/配置为准；缺失则跳过；多出来的列追加到末尾避免丢字段
            kpi_cols = [c for c in out_cols if c not in base_cols and c in df.columns]
            extra_cols = [c for c in df.columns if c not in base_cols and c not in kpi_cols]
            kpi_cols = kpi_cols + extra_cols

            df_final = df[base_cols + kpi_cols].copy()

            # 规范化索引字段，避免 NaN 导致列异常/空列
            df_final['活动名称'] = df_final['活动名称'].fillna('未知活动').astype(str).str.strip()
            df_final['区域'] = df_final['区域'].fillna('整体').astype(str).str.strip()
            df_final.loc[df_final['区域'].isin(['', 'nan', 'NaN', 'None']), '区域'] = '整体'
            df_final['指标时间'] = df_final['指标时间'].fillna('').astype(str).str.strip()
            df_final['厂家'] = df_final['厂家'].fillna('').astype(str).str.strip()

            # 去重：同一(活动,区域,时间,厂家)只保留第一条（避免重复列）
            df_final = df_final.drop_duplicates(subset=base_cols, keep='first')

            # 活动顺序：按出现顺序稳定复现
            acts = list(pd.unique(df_final['活动名称']))

            def _reg_key(x):
                s = str(x).strip()
                if s == '整体':
                    return (0, '')
                if s == '场内':
                    return (1, '')
                if s == '场外':
                    return (2, '')
                return (3, s)  # 其它：名称排序

            df_final['_act_order'] = df_final['活动名称'].apply(lambda x: acts.index(x) if x in acts else 9999)
            df_final['_reg_order'] = df_final['区域'].apply(_reg_key)
            df_final = df_final.sort_values(['_act_order', '_reg_order', '指标时间', '厂家'], kind='mergesort')
            df_final = df_final.drop(columns=['_act_order', '_reg_order'])

            # 转置：列为(活动,区域,时间,厂家)四级表头；行为指标项
            df_t = df_final.set_index(base_cols)[kpi_cols].T
            # 关键：设置四级表头名称，使 A1~A4 显示为“指标名称/区域/指标时间/厂家”
            df_t.columns.names = ['指标名称', '区域', '指标时间', '厂家']

            df_t.to_excel(writer, sheet_name, merge_cells=True)
        else:
            # 若缺字段则按原始表写入（兜底，不改变数据）
            df.to_excel(writer, sheet_name, index=False)

    except IndexError:
        log_print(f"⚠️ {sheet_name} 转置写入失败 (IndexError)，正在尝试以原始列表格式写入...", "WARN")
        try:
            df.to_excel(writer, sheet_name, index=False)
        except Exception as e:
            log_print(f"❌ {sheet_name} 写入彻底失败: {e}", "WARN")
    except Exception as e:
        log_print(f"⚠️ {sheet_name} 写入发生未知错误，降级尝试: {e}", "WARN")
        try:
            df.to_excel(writer, sheet_name, index=False)
        except Exception:
            pass

def _copy_ws_with_style(src_ws, dst_ws, row_offset=0, col_offset=0):
    """将源工作表内容（含样式、合并单元格、行高列宽）复制到目标工作表的指定偏移位置。
    - 不改变任何单元格的值（包括公式字符串）
    - 复制样式用于保持明细/汇总表的可读性
    """
    from copy import copy
    from openpyxl.utils.cell import range_boundaries, get_column_letter

    # 复制列宽
    for col, dim in src_ws.column_dimensions.items():
        if dim is None:
            continue
        try:
            # col 可能是 'A'/'B'...
            dst_col = get_column_letter(range_boundaries(f"{col}1:{col}1")[0] + col_offset)
        except Exception:
            # 回退：按字母转换
            try:
                from openpyxl.utils import column_index_from_string
                idx = column_index_from_string(col) + col_offset
                dst_col = get_column_letter(idx)
            except Exception:
                continue
        if dim.width is not None:
            # 多次写入取最大，避免覆盖更宽的设置
            old_w = dst_ws.column_dimensions[dst_col].width
            if old_w is None or dim.width > old_w:
                dst_ws.column_dimensions[dst_col].width = dim.width

    # 复制行高
    for r, dim in src_ws.row_dimensions.items():
        if dim is None:
            continue
        if dim.height is not None:
            dst_ws.row_dimensions[r + row_offset].height = dim.height

    # 复制单元格值与样式
    max_row = src_ws.max_row or 0
    max_col = src_ws.max_column or 0
    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            src_cell = src_ws.cell(row=r, column=c)
            dst_cell = dst_ws.cell(row=r + row_offset, column=c + col_offset, value=src_cell.value)

            # 复制样式（尽量完整）
            if src_cell.has_style:
                dst_cell.font = copy(src_cell.font)
                dst_cell.fill = copy(src_cell.fill)
                dst_cell.border = copy(src_cell.border)
                dst_cell.alignment = copy(src_cell.alignment)
                dst_cell.number_format = src_cell.number_format
                dst_cell.protection = copy(src_cell.protection)
                dst_cell._style = copy(src_cell._style)

            # 其他属性（可选）
            if src_cell.hyperlink:
                dst_cell.hyperlink = copy(src_cell.hyperlink)
            if src_cell.comment:
                dst_cell.comment = copy(src_cell.comment)

    # 复制合并单元格
    for merged in list(src_ws.merged_cells.ranges):
        min_col, min_row, max_col, max_row = range_boundaries(str(merged))
        new_range = f"{get_column_letter(min_col + col_offset)}{min_row + row_offset}:{get_column_letter(max_col + col_offset)}{max_row + row_offset}"
        try:
            dst_ws.merge_cells(new_range)
        except Exception:
            pass

    # 冻结窗格（如果源表有）
    try:
        if src_ws.freeze_panes:
            dst_ws.freeze_panes = src_ws.freeze_panes
    except Exception:
        pass

def consolidate_output_workbook_sheets(wb):
    """将输出Excel中的明细/汇总工作表合并，减少sheet数量。

    在 v16.2“工作表整合”基础上，按新增需求：
      1) 保留“质差小区明细”（不得删除）
      2) 将“4G指标监控/5G指标监控”置于最前

    仅调整工作表结构/数量/顺序，不改变任何计算逻辑/数据值/统计规则/格式要求。
    """
    # 需要合并的源 sheet
    merge_plan = [
        ("指标计算明细_4G", "指标计算汇总_4G", "4G指标明细与汇总"),
        ("指标计算明细_5G", "指标计算汇总_5G", "5G指标明细与汇总"),
    ]

    # 必须保留的工作表（名称精确匹配）
    # 用户要求：仅保留 4G指标监控、5G指标监控、质差小区明细、识别小区明细_汇总
    keep_names = set(["4G指标监控", "5G指标监控", "质差小区明细", "识别小区明细_汇总"])
    # 需要保留的前缀（兼容可能存在的计数/变体命名）
    keep_prefixes = ("质差小区明细", "识别小区明细")

    for detail_name, summary_name, merged_name in merge_plan:
        has_detail = detail_name in wb.sheetnames
        has_summary = summary_name in wb.sheetnames
        if not (has_detail or has_summary):
            continue

        # 删除已存在同名合并表（避免重复）
        if merged_name in wb.sheetnames:
            try:
                wb.remove(wb[merged_name])
            except Exception:
                pass

        # 创建新合并表（纵向：上半明细，下半汇总）
        ws_new = wb.create_sheet(merged_name)
        cur_row_offset = 0

        if has_detail:
            ws_d = wb[detail_name]
            _copy_ws_with_style(ws_d, ws_new, row_offset=cur_row_offset, col_offset=0)
            cur_row_offset += (ws_d.max_row or 0) + 2  # 空两行分隔

        if has_summary:
            ws_s = wb[summary_name]
            _copy_ws_with_style(ws_s, ws_new, row_offset=cur_row_offset, col_offset=0)

        # 用户要求不保留合并表，因此不加入 keep_names
        # keep_names.add(merged_name)

        # 删除旧表
        for old in (detail_name, summary_name):
            if old in wb.sheetnames:
                try:
                    wb.remove(wb[old])
                except Exception:
                    pass

    def _keep_sheet(name: str) -> bool:
        if name in keep_names:
            return True
        for p in keep_prefixes:
            if str(name).startswith(p):
                return True
        return False

    # 删除其它临时/中间表（保留：指标监控/质差明细/公式/合并后的明细汇总）
    for name in list(wb.sheetnames):
        if not _keep_sheet(name):
            try:
                wb.remove(wb[name])
            except Exception:
                pass

    # 调整 sheet 顺序：按用户要求排列
    poor_sheets = [n for n in wb.sheetnames if str(n).startswith("质差小区明细")]
    match_sheets = [n for n in wb.sheetnames if str(n).startswith("识别小区明细")]

    desired = ["4G指标监控", "5G指标监控"] + poor_sheets + match_sheets

    # 去重且保留存在的表
    seen = set()
    ordered = []
    for n in desired:
        if n in wb.sheetnames and n not in seen:
            ordered.append(wb[n])
            seen.add(n)

    # 兜底：把可能遗漏但仍在 keep_names/prefix 中的表追加到最后
    for n in wb.sheetnames:
        if _keep_sheet(n) and n not in seen:
            ordered.append(wb[n])
            seen.add(n)

    try:
        wb._sheets = ordered
    except Exception:
        pass

def style_excel(file_path):
    """仅做 Excel 输出格式美化：不改变任何计算逻辑、数据值与统计规则。

    注意：本工具的“4G指标监控/5G指标监控”为【转置展示】：
      - 列：各活动（多级表头：活动名称/区域）
      - 行：指标项（第一列为指标名称）

    因此百分比格式需要按“指标行”来设置 number_format。
    """
    try:
        wb = load_workbook(file_path)

        font_normal = Font(name='微软雅黑', size=10)
        font_header = Font(name='微软雅黑', size=10, bold=True)
        align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        thin = Side(style='thin')
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        header_fill = PatternFill(start_color='DDEBF7', fill_type='solid')

        # —— 按需求：删除特定“识别小区_活动”工作表（如存在）——
        remove_sheets = {
            '识别小区_中兴4G测试', '识别小区_诺基亚4G测试', '识别小区_中兴5G测试', '识别小区_华为4G测试', '识别小区_华为5G测试'
        }
        for sn in list(wb.sheetnames):
            if sn in remove_sheets:
                try:
                    wb.remove(wb[sn])
                except Exception:
                    pass

        # —— 兼容：若某些工作表存在同名“识别小区_*”列，则删除这些列（如存在）——
        remove_cols = set(remove_sheets)
        for ws in wb.worksheets:
            max_col = ws.max_column or 0
            headers = [ws.cell(row=1, column=c).value for c in range(1, max_col + 1)]
            for c in range(max_col, 0, -1):
                if headers[c - 1] in remove_cols:
                    try:
                        ws.delete_cols(c, 1)
                    except Exception:
                        pass        # —— 删除转置表头后产生的空行（兼容旧格式第3行空白；新格式为第5行空白）——
        for sn in ['4G指标监控', '5G指标监控']:
            if sn in wb.sheetnames:
                ws = wb[sn]

                def _row_blank(rr: int) -> bool:
                    if rr <= 0 or rr > (ws.max_row or 0):
                        return False
                    for cc in range(1, (ws.max_column or 0) + 1):
                        vv = ws.cell(row=rr, column=cc).value
                        if vv not in (None, ""):
                            return False
                    return True

                # 旧格式：表头2行+空白第3行
                if _row_blank(3):
                    ws.delete_rows(3, 1)
                # 新格式：表头4行+空白第5行
                if _row_blank(5):
                    ws.delete_rows(5, 1)

        def _cell_text(v):
            if v is None:
                return ""
            if isinstance(v, float):
                return f"{v:.6g}"
            return str(v)


        def _display_width(v) -> int:
            """近似估算 Excel 列宽：中文按2，英文按1，数字按1。"""
            if v is None:
                return 0
            s = str(v)
            w = 0
            for ch in s:
                # CJK
                if '\u4e00' <= ch <= '\u9fff':
                    w += 2
                else:
                    w += 1
            return w

        def _auto_fit(ws, min_w=8, max_w=90, padding=2):
            """按内容自适应列宽（尽量紧凑但不遮挡），并针对“最大利用率小区”行做额外保障。"""
            col_widths = {}
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value is None:
                        continue
                    txt = str(cell.value)
                    w = _display_width(txt) + padding
                    col_widths[cell.column_letter] = max(col_widths.get(cell.column_letter, 0), w)

            for col, w in col_widths.items():
                w2 = max(min_w, min(int(w), max_w))
                ws.column_dimensions[col].width = w2

            # 针对“最大利用率小区”所在行：确保对应值列宽足够，避免被遮挡
            try:
                max_row = ws.max_row or 0
                max_col = ws.max_column or 0
                target_row = None
                for r in range(1, max_row + 1):
                    v = ws.cell(r, 1).value
                    if v is not None and str(v).strip() == '最大利用率小区':
                        target_row = r
                        break
                if target_row:
                    for c in range(2, max_col + 1):
                        cell = ws.cell(target_row, c)
                        if cell.value is None:
                            continue
                        need = _display_width(cell.value) + padding
                        col_letter = cell.column_letter
                        cur = ws.column_dimensions[col_letter].width or min_w
                        ws.column_dimensions[col_letter].width = max(cur, min(int(need), max_w))
            except Exception:
                pass

        def _calc_row_height(ws, col_widths):
            # 按 wrap_text 粗略估算行高
            for r in range(1, ws.max_row + 1):
                max_lines = 1
                for c in range(1, ws.max_column + 1):
                    v = ws.cell(r, c).value
                    if v is None:
                        continue
                    txt = str(v)
                    est_width = max(1, int(col_widths.get(get_column_letter(c), 12)))
                    lines = max(1, int(math.ceil(len(txt) / est_width)))
                    max_lines = max(max_lines, lines)
                ws.row_dimensions[r].height = min(90, max(15, 15 * max_lines))

        def _style_transposed_kpi_sheet(ws, tech_label: str):
            """按用户要求格式化转置后的 KPI 监控工作表。
            
            格式要求：
            - 列宽：38（B列及后续活动列）
            - 字体：微软雅黑，10号，加粗
            - 对齐：水平居中，垂直居中，自动换行
            - 背景：表头区域（1-4行）淡蓝色 #DDEBF7
            - 边框：细边框
            - 数字格式：百分比类 0.00"%"、数值类 0.00、整数类 0/General
            """
            max_col = ws.max_column or 0
            max_row = ws.max_row or 0
            
            if max_row == 0 or max_col == 0:
                return
            
            # 样式定义
            font_bold = Font(name='微软雅黑', size=10, bold=True)
            font_normal = Font(name='微软雅黑', size=10)
            align_center_wrap = Alignment(horizontal='center', vertical='center', wrap_text=True)
            header_fill = PatternFill(start_color='DDEBF7', fill_type='solid')
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # 设置列宽：A列自适应，B列及后续固定38
            ws.column_dimensions['A'].width = 25  # A列（指标名称）稍窄
            for c in range(2, max_col + 1):
                ws.column_dimensions[get_column_letter(c)].width = 38
            
            # 删除第5行空行（如果存在）
            if max_row >= 5:
                row5_empty = True
                for c in range(1, max_col + 1):
                    v = ws.cell(5, c).value
                    if v is not None and str(v).strip() != '':
                        row5_empty = False
                        break
                if row5_empty:
                    ws.delete_rows(5, 1)
                    max_row = ws.max_row or 0
            
            # 全表格式设置
            for r in range(1, max_row + 1):
                for c in range(1, max_col + 1):
                    cell = ws.cell(r, c)
                    is_header_row = (r <= 4)
                    is_first_col = (c == 1)
                    
                    # 字体：表头区域和第一列加粗
                    if is_header_row or is_first_col:
                        cell.font = font_bold
                    else:
                        cell.font = font_normal
                    
                    # 对齐：居中 + 自动换行
                    cell.alignment = align_center_wrap
                    
                    # 边框
                    cell.border = thin_border
                    
                    # 背景：表头区域（1-4行）淡蓝色
                    if is_header_row:
                        cell.fill = header_fill
            
            # 数字格式设置（按指标行名称判断）
            # 定义各类指标
            pct_indicators = {
                '无线接通率(%)', '无线掉线率(%)', '系统内切换出成功率(%)',
                'VoLTE无线接通率(%)', 'VoLTE切换成功率(%)', 'E-RAB掉话率(QCI=1)(%)',
                'VoNR无线接通率(%)', 'VoNR到VoLTE切换成功率(%)', 'VoNR掉线率(5QI1)(%)',
                '4G利用率最大值(%)', '5G利用率最大值(%)',
                '最大利用率小区的利用率'
            }
            decimal_indicators = {
                '总流量(GB)', 'VoLTE话务量(Erl)', 'VoNR话务量(Erl)', '平均干扰(dBm)'
            }
            integer_indicators = {
                '总用户数', '高负荷小区数', '质差小区数', '最大利用率小区的用户数'
            }
            
            for r in range(1, max_row + 1):
                kpi_name = ws.cell(r, 1).value
                if kpi_name is None:
                    continue
                kpi_name = str(kpi_name).strip()
                
                # 确定数字格式
                num_format = None
                if kpi_name in pct_indicators:
                    num_format = '0.00"%"'
                elif kpi_name in decimal_indicators:
                    num_format = '0.00'
                elif kpi_name in integer_indicators:
                    num_format = '0'
                # 文本类（如"最大利用率小区"）保持 General
                
                if num_format:
                    for c in range(2, max_col + 1):
                        cell = ws.cell(r, c)
                        if isinstance(cell.value, (int, float)):
                            cell.number_format = num_format

        if '4G指标监控' in wb.sheetnames:
            _style_transposed_kpi_sheet(wb['4G指标监控'], '4G')
        if '5G指标监控' in wb.sheetnames:
            _style_transposed_kpi_sheet(wb['5G指标监控'], '5G')
        # 工作表整合：合并明细/汇总并删除中间表（仅结构，不改数据/口径/样式要求）
        try:
            consolidate_output_workbook_sheets(wb)
        except Exception:
            pass

        # —— 指定工作表统一美化：不改数据/口径，仅调整字体/列宽/行高/表头样式 —— 
        try:
            _beautify_targets = {'质差小区明细', '指标计算公式', '4G指标明细与汇总', '5G指标明细与汇总'}
            font9 = Font(name='微软雅黑', size=9)
            font9_b = Font(name='微软雅黑', size=9, bold=True)
            font9_red_b = Font(name='微软雅黑', size=9, bold=True, color='FF0000')
            header_fill2 = PatternFill(start_color='F2F2F2', fill_type='solid')
            total_fill = PatternFill(start_color='FFF2CC', fill_type='solid')
            align_c = Alignment(horizontal='center', vertical='center', wrap_text=False)
            align_l = Alignment(horizontal='left', vertical='center', wrap_text=False)
            align_r = Alignment(horizontal='right', vertical='center', wrap_text=False)

            def _looks_number(v):
                if v is None:
                    return False
                if isinstance(v, (int, float)):
                    return True
                if isinstance(v, str):
                    ss = v.strip()
                    if ss == '':
                        return False
                    ss2 = ss.replace('%', '').replace(',', '')
                    try:
                        float(ss2)
                        return True
                    except Exception:
                        return False
                return False

            def _beautify_ws(ws):
                ws.sheet_view.showGridLines = False
                # 冻结首行（便于查看）
                try:
                    ws.freeze_panes = 'A2'
                except Exception:
                    pass

                max_row = ws.max_row or 0
                max_col = ws.max_column or 0
                if max_row <= 0 or max_col <= 0:
                    return

                # 表头（第一行）
                ws.row_dimensions[1].height = 18
                for c in range(1, max_col + 1):
                    cell = ws.cell(1, c)
                    cell.font = font9_b
                    cell.fill = header_fill2
                    cell.alignment = align_c
                    cell.border = border

                # 内容区：字体/对齐/边框 + 紧凑行高
                for r in range(2, max_row + 1):
                    ws.row_dimensions[r].height = 15
                    for c in range(1, max_col + 1):
                        cell = ws.cell(r, c)
                        cell.font = font9
                        cell.border = border
                        v = cell.value
                        if _looks_number(v):
                            cell.alignment = align_r
                        else:
                            cell.alignment = align_l

                # 汇总/合计行：若第一列包含“合计/总计/小计”，则突出显示
                for r in range(2, max_row + 1):
                    v0 = ws.cell(r, 1).value
                    if isinstance(v0, str) and any(k in v0 for k in ('合计', '总计', '小计')):
                        for c in range(1, max_col + 1):
                            cell = ws.cell(r, c)
                            cell.font = font9_b
                            cell.fill = total_fill

                # 指定列格式修正（仅格式，不改值）
                try:
                    headers = [ws.cell(1, c).value for c in range(1, max_col + 1)]
                except Exception:
                    headers = []

                # 4G/5G 明细与汇总：用户数列统一为“常规(General)”避免科学计数法
                if ws.title in ('4G指标明细与汇总', '5G指标明细与汇总'):
                    if '最大利用率小区的用户数' in headers:
                        _cc = headers.index('最大利用率小区的用户数') + 1
                        for _r in range(2, max_row + 1):
                            try:
                                ws.cell(_r, _cc).number_format = 'General'
                            except Exception:
                                pass

                # 质差小区明细：备注列红色加粗突出显示
                if ws.title == '质差小区明细':
                    if '备注' in headers:
                        _cc = headers.index('备注') + 1
                        for _r in range(2, max_row + 1):
                            cell = ws.cell(_r, _cc)
                            if cell.value is not None and str(cell.value).strip() != '':
                                cell.font = font9_red_b

                # 自适应列宽（紧凑但不遮挡）
                col_widths = {}
                for r in range(1, max_row + 1):
                    for c in range(1, max_col + 1):
                        v = ws.cell(r, c).value
                        if v is None:
                            continue
                        w = _display_width(v) + 2
                        col_letter = get_column_letter(c)
                        col_widths[col_letter] = max(col_widths.get(col_letter, 0), w)

                for col_letter, w in col_widths.items():
                    ws.column_dimensions[col_letter].width = max(6, min(int(w), 90))

            for sn in list(wb.sheetnames):
                if sn in _beautify_targets:
                    _beautify_ws(wb[sn])
        except Exception:
            pass


        wb.save(file_path)
    except Exception:
        pass

def export_kpi_calc_formulas(writer):
    """输出每项指标的计算口径说明（公式/过滤条件/加权方式）。"""
    rows = []
    def add(tech, name, method, fields, filters, notes=""):
        rows.append({
            "制式": tech,
            "指标": name,
            "计算方式": method,
            "涉及字段": fields,
            "过滤条件": filters,
            "备注": notes
        })

    # 4G
    add("4G", "总用户数", "合计(SUM)", "STD__kpi_rrc_users_max", "NaN->0 后求和", "")
    add("4G", "总流量(GB)", "合计(SUM)", "STD__kpi_traffic_gb", "NaN->0 后求和；若源字段为MB则/1024换算", "")
    add("4G", "无线接通率(%)", "均值(MEAN)", 
        "STD__kpi_connect（来自RAW无线接通率(%)）",
        "排除NaN/0", "固定口径：MEAN(排NaN/0)，不做加权(∑num/∑den)")
    add("4G", "无线掉线率(%)", "RAW均值(MEAN)", 
        "RAW__kpi_drop（优先；仅排NaN，0为有效）",
        "仅排除NaN（0为有效）", "如检测为小数比值会自动*100；不再依赖STD__kpi_drop的异常判缺失")
    add("4G", "系统内切换出成功率(%)", "均值(MEAN)", "STD__kpi_ho_intra", "排除NaN/0", "")
    add("4G", "平均干扰(dBm)", "均值(MEAN)", "STD__kpi_ul_interf_dbm", "排除NaN（0保留）", "")
    add("4G", "VoLTE无线接通率(%)", "均值(MEAN)", 
        "STD__kpi_volte_connect（来自RAW VoLTE无线接通率/QCI=1接通率）",
        "排除NaN/0", "固定口径：MEAN(排NaN/0)，不做加权(∑num/∑den)")
    add("4G", "VoLTE切换成功率(%)", "若缺字段则输出“指标项缺失”", "STD__kpi_volte_ho", "整列NaN/缺失 => 指标项缺失", "")
    add("4G", "E-RAB掉话率(QCI=1)(%)", "均值(MEAN)", "STD__kpi_volte_drop", "排除NaN（0保留）", "")
    add("4G", "VoLTE话务量(Erl)", "合计(SUM)", "STD__kpi_volte_traffic_erl", "NaN->0 后求和", "")
    add("4G", "4G利用率最大值(%)", "最大值(MAX)", "max(STD__kpi_util_max, STD__kpi_prb_ul_util, STD__kpi_prb_dl_util)", "排除NaN", "")

    # 5G
    add("5G", "总用户数", "合计(SUM)", "STD__kpi_rrc_users_max", "NaN->0 后求和", "")
    add("5G", "总流量(GB)", "合计(SUM)", "STD__kpi_traffic_gb", "NaN->0 后求和；若源字段为MB则/1024换算", "")
    add("5G", "无线接通率(%)", "均值(MEAN)", 
        "STD__kpi_connect（来自RAW无线接通率(%)）",
        "排除NaN/0", "固定口径：MEAN(排NaN/0)，不做加权(∑num/∑den)")
    add("5G", "无线掉线率(%)", "RAW均值(MEAN)", 
        "RAW__kpi_drop（优先；仅排NaN，0为有效）",
        "仅排除NaN（0为有效）", "如检测为小数比值会自动*100；不再依赖STD__kpi_drop的异常判缺失")
    add("5G", "VoNR无线接通率(%)", "均值(MEAN)", "STD__kpi_vonr_connect", "排除NaN/0", "")
    add("5G", "VoNR到VoLTE切换成功率(%)", "均值(MEAN)", "STD__kpi_nr2lte_ho", "排除NaN/0", "")
    add("5G", "VoNR掉线率(5QI1)(%)", "均值(MEAN)", "STD__kpi_vonr_drop", "排除NaN/0（按你的口径：0不纳入）", "")
    add("5G", "VoNR话务量(Erl)", "合计(SUM)", "STD__kpi_vonr_traffic_erl", "NaN->0 后求和", "")
    add("5G", "平均干扰(dBm)", "均值(MEAN)", "STD__kpi_ul_interf_dbm", "排除NaN（0保留）", "")
    add("5G", "5G利用率最大值(%)", "最大值(MAX)", "max(STD__kpi_util_max, STD__kpi_prb_ul_util, STD__kpi_prb_dl_util)", "排除NaN", "")

    df = pd.DataFrame(rows)
    df.to_excel(writer, "指标计算公式", index=False)

def export_kpi_calc_details(writer, tech_label, merged_df):
    """输出逐小区的原始值/标准化值/纳入与否/中间量（num/den/贡献值），用于逐项对账。"""
    if merged_df is None or merged_df.empty:
        return

    df = merged_df.copy()

    # 利用率综合
    for c in ['STD__kpi_util_max', 'STD__kpi_prb_ul_util', 'STD__kpi_prb_dl_util']:
        if c not in df.columns:
            df[c] = np.nan
    df['KPI_UTIL'] = df[['STD__kpi_util_max', 'STD__kpi_prb_ul_util', 'STD__kpi_prb_dl_util']].max(axis=1)

    # 统一补全基础列
    for c in ['_source_file','match_method','raw_key_id','raw_key_name','list_key_id','list_key_name','活动名称','区域','厂家']:
        if c not in df.columns:
            df[c] = ""

    # --------- 行级“是否纳入/中间量” ---------
    def _prep_rate(rate_std_col, num_std_col=None, den_std_col=None, drop_zero=True, out_prefix="RATE"):
        # 初始化
        df[f'CALC__{out_prefix}__incl'] = False
        df[f'CALC__{out_prefix}__method'] = ""
        df[f'CALC__{out_prefix}__num'] = np.nan
        df[f'CALC__{out_prefix}__den'] = np.nan
        df[f'CALC__{out_prefix}__val'] = np.nan  # 行值（用于均值或行rate）
        if num_std_col and den_std_col and (num_std_col in df.columns) and (den_std_col in df.columns):
            num = pd.to_numeric(df[num_std_col], errors='coerce')
            den = pd.to_numeric(df[den_std_col], errors='coerce')
            ok = den > 0
            if ok.any():
                val = (num / den) * 100.0
                incl = ok & val.notna()
                if drop_zero:
                    incl = incl & (val != 0)
                df.loc[incl, f'CALC__{out_prefix}__incl'] = True
                df.loc[incl, f'CALC__{out_prefix}__method'] = "WEIGHTED_NUM_DEN"
                df.loc[incl, f'CALC__{out_prefix}__num'] = num[incl]
                df.loc[incl, f'CALC__{out_prefix}__den'] = den[incl]
                df.loc[incl, f'CALC__{out_prefix}__val'] = val[incl]
                return  # 有num/den就以此为准
        # fallback: mean on rate
        if rate_std_col in df.columns:
            val = pd.to_numeric(df[rate_std_col], errors='coerce')
            incl = val.notna()
            if drop_zero:
                incl = incl & (val != 0)
            df.loc[incl, f'CALC__{out_prefix}__incl'] = True
            df.loc[incl, f'CALC__{out_prefix}__method'] = "MEAN_RATE"
            df.loc[incl, f'CALC__{out_prefix}__val'] = val[incl]

    # 4G/5G 共用：无线接通率（固定口径：MEAN(排NaN/0)，不走加权）
    _prep_rate('STD__kpi_connect', None, None, drop_zero=True, out_prefix="无线接通率")

    # 掉线率：固定口径：RAW 均值（仅排NaN，0为有效）。优先用 RAW__kpi_drop；必要时做一次小数->百分数自适应
    _raw_drop = pd.to_numeric(df.get('RAW__kpi_drop'), errors='coerce') if 'RAW__kpi_drop' in df.columns else pd.Series([float('nan')] * len(df))
    _drop_base = _raw_drop.copy()
    _nz = _drop_base.dropna()
    _nz = _nz[_nz != 0]
    if len(_nz) > 0 and (_nz.quantile(0.95) <= 0.01) and (_nz.max() <= 0.1):
        _drop_base = _drop_base * 100
    df['KPI_DROP_MEAN_BASE'] = _drop_base
    _prep_rate('KPI_DROP_MEAN_BASE', None, None, drop_zero=False, out_prefix="无线掉线率")
    # 系统内切换成功率
    _prep_rate('STD__kpi_ho_intra', None, None, drop_zero=True, out_prefix="系统内切换")
    # VoLTE接通率（4G）
    _prep_rate('STD__kpi_volte_connect', None, None, drop_zero=True, out_prefix="VoLTE接通率")
    # VoLTE切换（4G）
    _prep_rate('STD__kpi_volte_ho', None, None, drop_zero=True, out_prefix="VoLTE切换成功率")
    # E-RAB掉话（4G）
    _prep_rate('STD__kpi_volte_drop', None, None, drop_zero=False, out_prefix="ERAB掉话率")
    # VoNR接通率（5G）
    _prep_rate('STD__kpi_vonr_connect', None, None, drop_zero=True, out_prefix="VoNR接通率")
    # VoNR到VoLTE切换（5G）
    _prep_rate('STD__kpi_nr2lte_ho', None, None, drop_zero=True, out_prefix="VoNR切换成功率")
    # VoNR掉线（5G）
    _prep_rate('STD__kpi_vonr_drop', None, None, drop_zero=True, out_prefix="VoNR掉线率")

    # --------- 明细表：每行 = 原始数据一行（对应一个小区在一个时间点） ---------
    base_cols = [
        'STD__time_start','STD__time_end','活动名称','区域','厂家','_source_file',
        'list_key_id','list_key_name','raw_key_id','raw_key_name','match_method','_list_idx'
    ]
    # KPI 相关列（源字段名/原始值/标准化值）
    kpi_keys = [
        'kpi_rrc_users_max','kpi_traffic_gb','kpi_connect','kpi_drop','kpi_ho_intra',
        'kpi_ul_interf_dbm','kpi_util_max','kpi_prb_ul_util','kpi_prb_dl_util',
        'kpi_volte_connect','kpi_volte_ho','kpi_volte_drop','kpi_volte_traffic_erl',
        'kpi_vonr_connect','kpi_nr2lte_ho','kpi_vonr_drop','kpi_vonr_traffic_erl',
        'kpi_connect_num','kpi_connect_den','kpi_volte_connect_num','kpi_volte_connect_den'
    ]
    extra_cols = []
    for k in kpi_keys:
        for prefix in ['SRC__','RAW__','STD__']:
            col = f"{prefix}{k}"
            if col in df.columns:
                extra_cols.append(col)
    # 计算辅助列
    calc_cols = [c for c in df.columns if c.startswith("CALC__")]
    if 'KPI_UTIL' in df.columns:
        extra_cols.append('KPI_UTIL')

    keep = [c for c in base_cols if c in df.columns] + extra_cols + calc_cols
    detail = df[keep].copy()

    # 写入明细
    sheet_detail = f"指标计算明细_{tech_label}"
    detail.to_excel(writer, sheet_detail, index=False)

    # --------- 汇总表：每活动×每指标，给出中间量（∑num/∑den / count） ---------
    sum_rows = []
    def _sum_metric(act_df, name, method, value, num_sum=None, den_sum=None, incl_rows=None, incl_cells=None, notes=""):
        sum_rows.append({
            "活动名称": act_df['活动名称'].iloc[0] if '活动名称' in act_df.columns and len(act_df)>0 else "",
            "制式": tech_label,
            "指标": name,
            "口径": method,
            "最终值": value,
            "分子合计": num_sum,
            "分母合计": den_sum,
            "入统行数": incl_rows,
            "入统小区数": incl_cells,
            "备注": notes
        })

    for act, g in df.groupby('活动名称'):
        # 入统小区数（用于核对“涉及小区列表”）
        cell_cnt = g['list_key_id'].nunique() if 'list_key_id' in g.columns else len(g)

        # 总用户/流量/话务：sum
        users = pd.to_numeric(g.get('STD__kpi_rrc_users_max'), errors='coerce').fillna(0).sum()
        traf = pd.to_numeric(g.get('STD__kpi_traffic_gb'), errors='coerce').fillna(0).sum()
        _sum_metric(g, "总用户数", "SUM", round(float(users),2), incl_rows=len(g), incl_cells=int(cell_cnt))
        _sum_metric(g, "总流量(GB)", "SUM", round(float(traf),2), incl_rows=len(g), incl_cells=int(cell_cnt))

        # 无线接通率
        incl = g['CALC__无线接通率__incl'] if 'CALC__无线接通率__incl' in g.columns else pd.Series([False]*len(g), index=g.index)
        method = g.loc[incl, 'CALC__无线接通率__method'].iloc[0] if incl.any() else ""
        if method == "WEIGHTED_NUM_DEN":
            num = pd.to_numeric(g.loc[incl, 'CALC__无线接通率__num'], errors='coerce').fillna(0).sum()
            den = pd.to_numeric(g.loc[incl, 'CALC__无线接通率__den'], errors='coerce').fillna(0).sum()
            val = (num/den*100.0) if den>0 else 0.0
            _sum_metric(g, "无线接通率(%)", "WEIGHTED(∑num/∑den)", round(float(val),2), num_sum=float(num), den_sum=float(den), incl_rows=int(incl.sum()), incl_cells=int(cell_cnt))
        else:
            v = pd.to_numeric(g.loc[incl, 'CALC__无线接通率__val'], errors='coerce').dropna()
            val = v.mean() if len(v) else 0.0
            _sum_metric(g, "无线接通率(%)", "MEAN(排NaN/0)", round(float(val),2), incl_rows=int(incl.sum()), incl_cells=int(cell_cnt))

        # 掉线率/切换/干扰：mean（掉线/切换按排0）
        def _mean_from_prefix(prefix, name, drop_zero=True):
            icol = f'CALC__{prefix}__incl'
            vcol = f'CALC__{prefix}__val'
            if icol in g.columns and vcol in g.columns:
                ii = g[icol]
                vv = pd.to_numeric(g.loc[ii, vcol], errors='coerce').dropna()
                return float(vv.mean()) if len(vv) else 0.0, int(ii.sum())
            return 0.0, 0

        val, n = _mean_from_prefix("无线掉线率", "无线掉线率(%)", drop_zero=False)
        _sum_metric(g, "无线掉线率(%)", "MEAN(仅排NaN,0有效)", round(val,3), incl_rows=n, incl_cells=int(cell_cnt))

        val, n = _mean_from_prefix("系统内切换", "系统内切换出成功率(%)", drop_zero=True)
        _sum_metric(g, "系统内切换出成功率(%)", "MEAN(排NaN/0)", round(val,2), incl_rows=n, incl_cells=int(cell_cnt))

        interf = pd.to_numeric(g.get('STD__kpi_ul_interf_dbm'), errors='coerce').dropna()
        _sum_metric(g, "平均干扰(dBm)", "MEAN(排NaN)", round(float(interf.mean()) if len(interf) else 0.0,2), incl_rows=int(interf.notna().sum()), incl_cells=int(cell_cnt))

        util = pd.to_numeric(g.get('KPI_UTIL'), errors='coerce').dropna()
        _sum_metric(g, f"{tech_label}利用率最大值(%)", "MAX", round(float(util.max()) if len(util) else 0.0,2), incl_rows=int(util.notna().sum()), incl_cells=int(cell_cnt))

        if tech_label == "4G":
            # VoLTE接通率
            incl2 = g['CALC__VoLTE接通率__incl'] if 'CALC__VoLTE接通率__incl' in g.columns else pd.Series([False]*len(g), index=g.index)
            method2 = g.loc[incl2, 'CALC__VoLTE接通率__method'].iloc[0] if incl2.any() else ""
            if method2 == "WEIGHTED_NUM_DEN":
                num2 = pd.to_numeric(g.loc[incl2, 'CALC__VoLTE接通率__num'], errors='coerce').fillna(0).sum()
                den2 = pd.to_numeric(g.loc[incl2, 'CALC__VoLTE接通率__den'], errors='coerce').fillna(0).sum()
                val2 = (num2/den2*100.0) if den2>0 else 0.0
                _sum_metric(g, "VoLTE无线接通率(%)", "WEIGHTED(∑num/∑den)", round(float(val2),2), num_sum=float(num2), den_sum=float(den2), incl_rows=int(incl2.sum()), incl_cells=int(cell_cnt))
            else:
                v2 = pd.to_numeric(g.loc[incl2, 'CALC__VoLTE接通率__val'], errors='coerce').dropna()
                val2 = v2.mean() if len(v2) else 0.0
                _sum_metric(g, "VoLTE无线接通率(%)", "MEAN(排NaN/0)", round(float(val2),2), incl_rows=int(incl2.sum()), incl_cells=int(cell_cnt))

            # VoLTE切换：若字段缺失/全NaN => 指标项缺失
            if 'STD__kpi_volte_ho' not in g.columns or pd.to_numeric(g.get('STD__kpi_volte_ho'), errors='coerce').dropna().empty:
                _sum_metric(g, "VoLTE切换成功率(%)", "MISSING", "指标项缺失", notes="源数据缺该指标列或全空")
            else:
                val3, n3 = _mean_from_prefix("VoLTE切换成功率", "VoLTE切换成功率(%)", drop_zero=True)
                _sum_metric(g, "VoLTE切换成功率(%)", "MEAN(排NaN/0)", round(val3,2), incl_rows=n3, incl_cells=int(cell_cnt))

            # E-RAB掉话
            val4, n4 = _mean_from_prefix("ERAB掉话率", "E-RAB掉话率(QCI=1)(%)", drop_zero=True)
            _sum_metric(g, "E-RAB掉话率(QCI=1)(%)", "MEAN(排NaN)", round(val4,3), incl_rows=n4, incl_cells=int(cell_cnt))

            volte_erl = pd.to_numeric(g.get('STD__kpi_volte_traffic_erl'), errors='coerce').fillna(0).sum()
            _sum_metric(g, "VoLTE话务量(Erl)", "SUM", round(float(volte_erl),2), incl_rows=len(g), incl_cells=int(cell_cnt))

        else:
            # VoNR接通/切换/掉线/话务
            val5, n5 = _mean_from_prefix("VoNR接通率", "VoNR无线接通率(%)", drop_zero=True)
            _sum_metric(g, "VoNR无线接通率(%)", "MEAN(排NaN/0)", round(val5,2), incl_rows=n5, incl_cells=int(cell_cnt))

            val6, n6 = _mean_from_prefix("VoNR切换成功率", "VoNR到VoLTE切换成功率(%)", drop_zero=True)
            _sum_metric(g, "VoNR到VoLTE切换成功率(%)", "MEAN(排NaN/0)", round(val6,2), incl_rows=n6, incl_cells=int(cell_cnt))

            val7, n7 = _mean_from_prefix("VoNR掉线率", "VoNR掉线率(5QI1)(%)", drop_zero=True)
            _sum_metric(g, "VoNR掉线率(5QI1)(%)", "MEAN(仅排NaN,0有效)", round(val7,3), incl_rows=n7, incl_cells=int(cell_cnt))

            vonr_erl = pd.to_numeric(g.get('STD__kpi_vonr_traffic_erl'), errors='coerce').fillna(0).sum()
            _sum_metric(g, "VoNR话务量(Erl)", "SUM", round(float(vonr_erl),2), incl_rows=len(g), incl_cells=int(cell_cnt))

    summary = pd.DataFrame(sum_rows)
    sheet_sum = f"指标计算汇总_{tech_label}"
    summary.to_excel(writer, sheet_sum, index=False)

def export_recognized_cell_details(writer, df_list, merged_4g, merged_5g, activities, list_export_cols=None):
    """把每个活动【实际识别/匹配成功】的清单小区明细导出到 Excel（用于核对）。
    输出格式参照《保障小区清单》的列顺序；末尾追加：制式、匹配方式。
    """
    if df_list is None or df_list.empty:
        return

    # 只保留清单原始列（避免内部字段）
    if list_export_cols is None:
        list_export_cols = [c for c in df_list.columns if c not in ['_list_idx', 'list_key_id', 'list_key_name']]
    else:
        list_export_cols = [c for c in list_export_cols if c in df_list.columns]

    def _build_detail(merged_df, tech):
        if merged_df is None or merged_df.empty or '_list_idx' not in merged_df.columns:
            return pd.DataFrame()
        if 'match_method' in merged_df.columns:
            tmp = merged_df[['_list_idx', 'match_method']].dropna(subset=['_list_idx']).copy()
        else:
            tmp = merged_df[['_list_idx']].dropna(subset=['_list_idx']).copy()
            tmp['match_method'] = ""
        # _list_idx 可能因 merge 变成 float
        tmp['_list_idx'] = tmp['_list_idx'].astype(int)

        # 同一个小区可能被多条网管行命中：保留“最优匹配方式”
        pri = {'EXACT_NAME': 1, 'EXACT_ID': 2, 'FUZZY_STRIP_PREFIX': 3}
        tmp['_pri'] = tmp['match_method'].map(pri).fillna(99).astype(int)
        tmp = tmp.sort_values(['_list_idx', '_pri']).drop_duplicates('_list_idx', keep='first')

        detail = df_list[df_list['_list_idx'].isin(tmp['_list_idx'])].copy()
        detail = detail.merge(tmp[['_list_idx', 'match_method']], on='_list_idx', how='left')
        detail['制式'] = tech
        detail.rename(columns={'match_method': '匹配方式'}, inplace=True)

        out_cols = [c for c in list_export_cols if c in detail.columns] + ['制式', '匹配方式']
        return detail[out_cols]

    detail_4g = _build_detail(merged_4g, '4G')
    detail_5g = _build_detail(merged_5g, '5G')
    if detail_4g.empty and detail_5g.empty:
        return

    all_detail = pd.concat([detail_4g, detail_5g], ignore_index=True)

    # 1) 汇总
    all_detail.to_excel(writer, sheet_name="识别小区明细_汇总", index=False)
